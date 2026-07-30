import sys
sys.path.insert(0,'.')
from modules.excel_importer import generar_plantilla_excel_oficial_bytes
import openpyxl, io

data = generar_plantilla_excel_oficial_bytes()
wb = openpyxl.load_workbook(io.BytesIO(data))
print("HOJAS:", wb.sheetnames)

ws_gas = wb["GAS"]
print("GAS row1 title:", ws_gas.cell(1,1).value)
print("GAS row3 headers:", [ws_gas.cell(3, c).value for c in range(1, 7)])
print("GAS row4 sample:", [ws_gas.cell(4, c).value for c in range(1, 7)])
print("GAS row200 total:", ws_gas.cell(200, 1).value, ws_gas.cell(200, 6).value)

ws_res = wb["RESUMEN DE COSTO"]
print("RESUMEN row3 headers:", [ws_res.cell(3, c).value for c in range(1, 11)])
print("RESUMEN row4 % formula:", ws_res.cell(4, 8).value)
print("RESUMEN row4 GAS prorrateo:", ws_res.cell(4, 9).value)
print("RESUMEN row4 TOTAL CON GAS:", ws_res.cell(4, 10).value)
print("RESUMEN row7 totals: GAS=", ws_res.cell(7, 9).value, "TOTAL=", ws_res.cell(7, 10).value)

with open("TEST_plantilla_GAS.xlsx", "wb") as f:
    f.write(data)
print("TEST_plantilla_GAS.xlsx OK!")
