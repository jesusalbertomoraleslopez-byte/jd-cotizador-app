import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import io
import os
import pandas as pd
from datetime import datetime, date
from config import BRAND_ORANGE, BRAND_CHARCOAL, BRAND_CHARCOAL_MED, BRAND_GRAY_BG, BRAND_WHITE, BRAND_BORDER_LIGHT
from database.models import get_connection, init_db
from database.db_manager import sync_cotizacion_totals

# ─────────────────────────────────────────────────────────────────────────────
# 1. GENERADOR DE PLANTILLA EXCEL OFICIAL ROBUSTA (HOJAS: PROY, WBS, DETALLES, RESUMEN, BD)
# ─────────────────────────────────────────────────────────────────────────────

def generar_plantilla_excel_oficial_bytes():
    """
    Genera la Plantilla Oficial Inteligente de Excel (.xlsx):
    Hojas: PROY, WBS, MAT, MO, GAS, SUB, MAQ, RESUMEN DE COSTO, BD (al final).
    - Sueldo Base y FASAR se calculan automáticamente en MO via VLOOKUP a BD al cambiar el Puesto.
    - Catálogos COMPLETOS de la BD: MO, Gastos, Subcontratos y Maquinaria.
    - Validaciones por Dropdown ancladas a la hoja BD.
    """
    import sqlite3 as _sqlite3
    import os as _os
    _db_path = _os.path.join(_os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))), "database", "cotizador.db")
    _conn = _sqlite3.connect(_db_path, timeout=10)
    _cur = _conn.cursor()

    _cur.execute("SELECT categoria, sueldo_base_semanal, fasar FROM catalogo_mano_obra WHERE activo=1 ORDER BY id")
    mo_tuples = _cur.fetchall() or [("Diseñador de Controles", 5500.0, 1.45)]

    _cur.execute("SELECT concepto, costo_unitario_default, unidad FROM catalogo_gastos WHERE activo=1 ORDER BY clave, id")
    gastos_rows = _cur.fetchall() or [("Gastos de Viaje y Viáticos", 1500.0, "VJE")]
    gastos_cat = [r[0] for r in gastos_rows]

    _cur.execute("SELECT concepto, costo_referencia, proveedor_habitual FROM catalogo_subcontratos WHERE activo=1 ORDER BY concepto")
    subs_rows = _cur.fetchall() or [("Maquinado CNC Especializado", 4500.0, "General")]
    subs_cat = [r[0] for r in subs_rows]

    # Reutilizar catálogo de gastos para maquinaria (concepto de renta de equipo) o subconjunto
    maq_cat = [r for r in gastos_rows if any(kw in r[0].lower() for kw in ["maquinaria","grúa","elevador","renta","equipo"])]
    if not maq_cat:
        maq_cat = gastos_rows[:5]

    _cur.execute("SELECT nombre FROM clientes WHERE activo=1 ORDER BY nombre")
    clientes_cat = [r[0] for r in _cur.fetchall()] or ["TREBOTTI", "YESERA MONTERREY", "General Motors"]
    _conn.close()

    wb = openpyxl.Workbook()

    # Estilos corporativos J&D
    fill_header = PatternFill(start_color="434E62", end_color="434E62", fill_type="solid")
    fill_orange  = PatternFill(start_color="FE8C29", end_color="FE8C29", fill_type="solid")
    fill_auto    = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    fill_calc    = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    fill_total   = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    font_header  = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_bold    = Font(name="Segoe UI", size=10, bold=True)
    font_title   = Font(name="Segoe UI", size=12, bold=True, color="FE8C29")
    font_auto    = Font(name="Segoe UI", size=9, italic=True, color="475569")
    font_total   = Font(name="Segoe UI", size=11, bold=True, color="FE8C29")

    def apply_header_row(ws, headers, row=3):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row, c, h)
            cell.fill = fill_header; cell.font = font_header
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[row].height = 30

    def wbs_formula(ws, r):
        c = ws.cell(r, 2, f"=IFERROR(VLOOKUP(A{r},WBS!A$4:B$200,2,FALSE),\"\")")
        c.fill = fill_auto; c.font = font_auto
        return c

    # ─────────────────────────────────────
    # 1. PROY
    # ─────────────────────────────────────
    ws_proy = wb.active; ws_proy.title = "PROY"
    ws_proy.cell(1,1,"PASO 1 — SELECCIÓN DE CLIENTE Y DATOS DEL PROYECTO").font = font_title
    apply_header_row(ws_proy, ["Campo / Parámetro","Valor (Presupuestador)","Instrucciones de Captura"])
    gen_rows = [
        ("Cliente (Seleccionar de Lista)", clientes_cat[0], "Selecciona el cliente de la lista desplegable → BD!F"),
        ("Nombre del Proyecto", "AUTOMATIZACIÓN DE LÍNEA DE PROCESO", "Nombre técnico oficial del proyecto"),
        ("Ingeniero Responsable / Presupuestador", "DS", "Iniciales del presupuestador (ej: DS, RG, JM)"),
        ("Tipo de Cambio USD (FIX)", 18.00, "Tipo de cambio MXN por 1 USD para conversiones"),
    ]
    for r, row in enumerate(gen_rows, 4):
        ws_proy.cell(r,1,row[0]).font = font_bold
        ws_proy.cell(r,2,row[1])
        ws_proy.cell(r,3,row[2]).font = Font(name="Segoe UI", size=9, color="64748B")
    ws_proy.column_dimensions['A'].width = 40
    ws_proy.column_dimensions['B'].width = 55
    ws_proy.column_dimensions['C'].width = 55

    # ─────────────────────────────────────
    # 2. WBS
    # ─────────────────────────────────────
    ws_wbs = wb.create_sheet("WBS")
    ws_wbs.cell(1,1,"PASO 2 — DEFINICIÓN DE PARTIDAS Y ESTRUCTURA WBS").font = font_title
    apply_header_row(ws_wbs, ["N° Partida WBS","Nombre / Concepto de la Partida WBS","Alcance General y Especificaciones Técnicas"])
    partidas_sample = [
        (1,"Gabinete de Control Principal PLC","Suministro, ensamble y cableado de tablero principal NEMA 12"),
        (2,"Programación de PLC y Pantalla HMI","Desarrollo de lógica de control, pantallas HMI y pruebas FAT/SAT"),
        (3,"Instalación Eléctrica e Integración en Planta","Tendido de tubería conduit, cableado de campo y puesta en marcha"),
    ]
    for r, p in enumerate(partidas_sample, 4):
        ws_wbs.cell(r,1,p[0]).alignment = Alignment(horizontal='center')
        ws_wbs.cell(r,2,p[1]).font = font_bold
        ws_wbs.cell(r,3,p[2])
    ws_wbs.column_dimensions['A'].width = 16
    ws_wbs.column_dimensions['B'].width = 50
    ws_wbs.column_dimensions['C'].width = 65

    # ─────────────────────────────────────
    # 3. MAT (Materiales)
    # ─────────────────────────────────────
    ws_mat = wb.create_sheet("MAT")
    ws_mat.cell(1,1,"PASO 3.1 — DETALLE DE MATERIALES Y COMPONENTES POR PARTIDA").font = font_title
    apply_header_row(ws_mat, ["Partida N°","Partida WBS (Auto)","Concepto / Material","Especificación / Marca","Cantidad","Unidad","Precio Unitario","Moneda (MXN/USD)"])
    mat_sample = [
        (1,"PLC CompactLogix 5380","Allen Bradley 5069-L310ER",1,"PZA",2500.00,"USD"),
        (1,"Fuente de Alimentación 24VDC 10A","Phoenix Contact Quint Power",2,"PZA",1850.00,"MXN"),
        (2,"Licencia de Software Studio 5000","Rockwell Automation Professional",1,"LIC",3200.00,"USD"),
    ]
    for r, m in enumerate(mat_sample, 4):
        ws_mat.cell(r,1,m[0]).alignment = Alignment(horizontal='center')
        wbs_formula(ws_mat, r)
        ws_mat.cell(r,3,m[1]).font = font_bold
        ws_mat.cell(r,4,m[2])
        ws_mat.cell(r,5,m[3])
        ws_mat.cell(r,6,m[4])
        ws_mat.cell(r,7,m[5]).number_format = '$#,##0.00'
        ws_mat.cell(r,8,m[6]).alignment = Alignment(horizontal='center')
    for r in range(7, 200):
        wbs_formula(ws_mat, r)
    ws_mat.column_dimensions['A'].width = 12
    ws_mat.column_dimensions['B'].width = 38
    ws_mat.column_dimensions['C'].width = 38
    ws_mat.column_dimensions['D'].width = 38
    ws_mat.column_dimensions['G'].width = 16

    # ─────────────────────────────────────
    # 4. MO (Mano de Obra) con Sueldo+FASAR+CostoHora auto por Puesto
    # ─────────────────────────────────────
    ws_mo = wb.create_sheet("MO")
    ws_mo.cell(1,1,"PASO 3.2 — MANO DE OBRA (SUELDO BASE Y FASAR SE ACTUALIZAN AL CAMBIAR EL PUESTO)").font = font_title
    ws_mo.cell(2,1,"🔵 Las columnas azules (Sueldo Base, FASAR, Costo Hora) se calculan AUTOMÁTICAMENTE desde el Catálogo BD al seleccionar el Puesto").font = Font(name="Segoe UI", size=9, italic=True, color="1565C0")
    apply_header_row(ws_mo, ["Partida N°","Partida WBS (Auto)","Puesto / Categoría (Seleccionar ▼)","Personas","Horas/Día","Días","Sueldo Base Semanal (BD)","FASAR (BD)","Costo Hora MXN (Auto)","Importe Total MO (MXN)"], row=3)
    mo_sample = [
        (1, mo_tuples[0][0], 1, 8, 5),
        (1, mo_tuples[1][0] if len(mo_tuples)>1 else mo_tuples[0][0], 1, 8, 10),
        (2, mo_tuples[2][0] if len(mo_tuples)>2 else mo_tuples[0][0], 1, 8, 5),
    ]
    for r, mo in enumerate(mo_sample, 4):
        ws_mo.cell(r,1,mo[0]).alignment = Alignment(horizontal='center')
        wbs_formula(ws_mo, r)
        ws_mo.cell(r,3,mo[1]).font = font_bold
        ws_mo.cell(r,4,mo[2]); ws_mo.cell(r,5,mo[3]); ws_mo.cell(r,6,mo[4])
        # Sueldo Base Semanal → VLOOKUP desde BD col B
        c_sb = ws_mo.cell(r,7, f"=IFERROR(VLOOKUP(C{r},BD!$A$2:$C${len(mo_tuples)+1},2,FALSE),3500)")
        c_sb.number_format = '$#,##0.00'; c_sb.fill = fill_auto; c_sb.font = font_auto
        # FASAR → VLOOKUP desde BD col C
        c_fs = ws_mo.cell(r,8, f"=IFERROR(VLOOKUP(C{r},BD!$A$2:$C${len(mo_tuples)+1},3,FALSE),1.45)")
        c_fs.number_format = '0.00'; c_fs.fill = fill_auto; c_fs.font = font_auto
        # Costo Hora = (Sueldo*FASAR)/48
        c_ch = ws_mo.cell(r,9, f"=(G{r}*H{r})/48")
        c_ch.number_format = '$#,##0.00'; c_ch.fill = fill_calc; c_ch.font = font_bold
        # Importe Total = Personas * Horas/Dia * Dias * CostoHora
        c_tot = ws_mo.cell(r,10, f"=D{r}*E{r}*F{r}*I{r}")
        c_tot.number_format = '$#,##0.00'; c_tot.fill = fill_total; c_tot.font = font_bold

    for r in range(7, 200):
        wbs_formula(ws_mo, r)
        c_sb = ws_mo.cell(r,7, f"=IFERROR(VLOOKUP(C{r},BD!$A$2:$C${len(mo_tuples)+1},2,FALSE),\"\")")
        c_sb.number_format = '$#,##0.00'; c_sb.fill = fill_auto; c_sb.font = font_auto
        c_fs = ws_mo.cell(r,8, f"=IFERROR(VLOOKUP(C{r},BD!$A$2:$C${len(mo_tuples)+1},3,FALSE),\"\")")
        c_fs.number_format = '0.00'; c_fs.fill = fill_auto; c_fs.font = font_auto
        c_ch = ws_mo.cell(r,9, f"=IFERROR((G{r}*H{r})/48,\"\")")
        c_ch.number_format = '$#,##0.00'; c_ch.fill = fill_calc
        c_tot = ws_mo.cell(r,10, f"=IFERROR(D{r}*E{r}*F{r}*I{r},\"\")")
        c_tot.number_format = '$#,##0.00'; c_tot.fill = fill_total

    ws_mo.column_dimensions['A'].width = 12
    ws_mo.column_dimensions['B'].width = 36
    ws_mo.column_dimensions['C'].width = 34
    ws_mo.column_dimensions['G'].width = 24
    ws_mo.column_dimensions['H'].width = 12
    ws_mo.column_dimensions['I'].width = 26
    ws_mo.column_dimensions['J'].width = 26

    # ─────────────────────────────────────
    # 5. GAS (Gastos Generales — GLOBALES, se prorratean en RESUMEN)
    # ─────────────────────────────────────
    ws_gastos = wb.create_sheet("GAS")
    ws_gastos.cell(1,1,"PASO 3.3 — GASTOS GENERALES DE OBRA (GLOBALES — SE PRORRATEAN ENTRE PARTIDAS EN EL RESUMEN)").font = font_title
    ws_gastos.cell(2,1,"📌 No asignar a partida. El costo total se distribuirá automáticamente entre todas las partidas según su % de costo directo.").font = Font(name="Segoe UI", size=9, italic=True, color="C0392B")
    apply_header_row(ws_gastos, ["Concepto de Gasto (Seleccionar ▼)","Cantidad","Unidad","Tiempo","Costo Unitario","IMPORTE TOTAL (MXN)"], row=3)

    gastos_sample2 = [
        (gastos_rows[0][0], 1, gastos_rows[0][2], 5, float(gastos_rows[0][1]) or 1500.0),
        (gastos_rows[1][0] if len(gastos_rows)>1 else gastos_rows[0][0], 1, (gastos_rows[1][2] if len(gastos_rows)>1 else "VJE"), 1, float((gastos_rows[1][1] if len(gastos_rows)>1 else gastos_rows[0][1])) or 1200.0),
    ]
    for r, g in enumerate(gastos_sample2, 4):
        ws_gastos.cell(r,1,g[0]).font = font_bold
        ws_gastos.cell(r,2,g[1]); ws_gastos.cell(r,3,g[2]); ws_gastos.cell(r,4,g[3])
        ws_gastos.cell(r,5,g[4]).number_format = '$#,##0.00'
        c_tot = ws_gastos.cell(r,6, f"=B{r}*D{r}*E{r}")
        c_tot.number_format = '$#,##0.00'; c_tot.font = font_bold; c_tot.fill = fill_total

    for r in range(6, 200):
        c_tot = ws_gastos.cell(r,6, f"=IFERROR(B{r}*D{r}*E{r},\"\")")
        c_tot.number_format = '$#,##0.00'; c_tot.fill = fill_total

    # Fila de TOTAL GASTOS
    ws_gastos.cell(200,1,"TOTAL GASTOS GENERALES DEL PROYECTO:").font = font_total
    c_tgas = ws_gastos.cell(200,6,"=SUM(F4:F199)")
    c_tgas.number_format = '$#,##0.00'; c_tgas.font = font_total; c_tgas.fill = fill_orange
    ws_gastos.cell(200,6).font = Font(name="Segoe UI", size=13, bold=True, color="FFFFFF")

    ws_gastos.column_dimensions['A'].width = 50
    ws_gastos.column_dimensions['B'].width = 12
    ws_gastos.column_dimensions['C'].width = 12
    ws_gastos.column_dimensions['D'].width = 12
    ws_gastos.column_dimensions['E'].width = 18
    ws_gastos.column_dimensions['F'].width = 24

    # ─────────────────────────────────────
    # 6. SUB (Subcontratos)
    # ─────────────────────────────────────
    ws_subs = wb.create_sheet("SUB")
    ws_subs.cell(1,1,"PASO 3.4 — DETALLE DE SUBCONTRATOS POR PARTIDA").font = font_title
    apply_header_row(ws_subs, ["Partida N°","Partida WBS (Auto)","Concepto Subcontrato (Seleccionar ▼)","Proveedor Habitual","Cantidad","Unidad","Precio Unitario","Moneda","Importe Total MXN"])
    sub_sample = [(1, subs_rows[0][0], subs_rows[0][2] or "Metalmecánica del Norte", 1, "SERV", float(subs_rows[0][1]) or 4500.0, "MXN")]
    for r, s in enumerate(sub_sample, 4):
        ws_subs.cell(r,1,s[0]).alignment = Alignment(horizontal='center')
        wbs_formula(ws_subs, r)
        ws_subs.cell(r,3,s[1]).font = font_bold
        ws_subs.cell(r,4,s[2]); ws_subs.cell(r,5,s[3]); ws_subs.cell(r,6,s[4])
        ws_subs.cell(r,7,s[5]).number_format = '$#,##0.00'
        ws_subs.cell(r,8,s[6]).alignment = Alignment(horizontal='center')
        c_tot = ws_subs.cell(r,9, f"=IF(H{r}=\"USD\",E{r}*G{r}*PROY!$B$7,E{r}*G{r})")
        c_tot.number_format = '$#,##0.00'; c_tot.font = font_bold; c_tot.fill = fill_total
    for r in range(5, 200):
        wbs_formula(ws_subs, r)
        c_tot = ws_subs.cell(r,9, f"=IFERROR(IF(H{r}=\"USD\",E{r}*G{r}*PROY!$B$7,E{r}*G{r}),\"\")")
        c_tot.number_format = '$#,##0.00'; c_tot.fill = fill_total
    ws_subs.column_dimensions['A'].width = 12
    ws_subs.column_dimensions['B'].width = 38
    ws_subs.column_dimensions['C'].width = 42
    ws_subs.column_dimensions['G'].width = 16
    ws_subs.column_dimensions['I'].width = 22

    # ─────────────────────────────────────
    # 7. MAQ (Maquinaria y Equipo)
    # ─────────────────────────────────────
    ws_maq = wb.create_sheet("MAQ")
    ws_maq.cell(1,1,"PASO 3.5 — DETALLE DE MAQUINARIA Y EQUIPO POR PARTIDA").font = font_title
    apply_header_row(ws_maq, ["Partida N°","Partida WBS (Auto)","Concepto Maquinaria / Equipo (Seleccionar ▼)","Modelo / Capacidad","Tiempo","Unidad","Costo Unitario","Moneda","Importe Total MXN"])
    maq_sample = [(1, maq_cat[0][0] if maq_cat else "Maquinaria y Equipo Menor de Obra", "—", 5, "DÍA", float(maq_cat[0][1]) if maq_cat and maq_cat[0][1] else 1200.0, "MXN")]
    for r, mq in enumerate(maq_sample, 4):
        ws_maq.cell(r,1,mq[0]).alignment = Alignment(horizontal='center')
        wbs_formula(ws_maq, r)
        ws_maq.cell(r,3,mq[1]).font = font_bold
        ws_maq.cell(r,4,mq[2]); ws_maq.cell(r,5,mq[3]); ws_maq.cell(r,6,mq[4])
        ws_maq.cell(r,7,mq[5]).number_format = '$#,##0.00'
        ws_maq.cell(r,8,mq[6]).alignment = Alignment(horizontal='center')
        c_tot = ws_maq.cell(r,9, f"=IF(H{r}=\"USD\",E{r}*G{r}*PROY!$B$7,E{r}*G{r})")
        c_tot.number_format = '$#,##0.00'; c_tot.font = font_bold; c_tot.fill = fill_total
    for r in range(5, 200):
        wbs_formula(ws_maq, r)
        c_tot = ws_maq.cell(r,9, f"=IFERROR(IF(H{r}=\"USD\",E{r}*G{r}*PROY!$B$7,E{r}*G{r}),\"\")")
        c_tot.number_format = '$#,##0.00'; c_tot.fill = fill_total
    ws_maq.column_dimensions['A'].width = 12
    ws_maq.column_dimensions['B'].width = 38
    ws_maq.column_dimensions['C'].width = 42
    ws_maq.column_dimensions['G'].width = 16
    ws_maq.column_dimensions['I'].width = 22

    # ─────────────────────────────────────
    # 7.6. PLAN (Cronograma / Gantt)
    # ─────────────────────────────────────
    global ws_gantt
    ws_gantt = wb.create_sheet("PLAN")
    ws_gantt.cell(1,1,"PASO 3.6 — CRONOGRAMA DE GANTT Y PLAN DE PROYECTO").font = font_title
    ws_gantt.cell(2,1,"📌 Detalla las actividades del proyecto. La columna D muestra un calendario flotante. La columna G busca la partida y la columna I define predecesoras. El Gantt se dibuja automáticamente en columnas J en adelante.").font = Font(name="Segoe UI", size=9, italic=True, color="64748B")
    
    font_wbs_size7 = Font(name="Segoe UI", size=7, italic=True, color="475569")
    
    apply_header_row(ws_gantt, [
        "N° Tarea",
        "Tipo de Tarea",
        "Ing. Responsable / Rol",
        "Fecha de Inicio (AAAA-MM-DD)",
        "Duración (Días)",
        "Asociar a Partida N°",
        "Partida WBS (Auto)",
        "Descripción de la Actividad",
        "Predecesoras"
    ], row=3)

    # El Gantt ahora empieza en la columna J (Col 10)
    ws_gantt.cell(3, 10, "=D4") # Primera fecha
    ws_gantt.cell(3, 10).number_format = 'dd'
    ws_gantt.cell(3, 10).font = font_bold
    ws_gantt.cell(3, 10).fill = fill_orange
    ws_gantt.cell(3, 10).font = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
    ws_gantt.cell(3, 10).alignment = Alignment(horizontal='center')

    # Nombre del mes dinámico en la fila 2 en la columna J
    ws_gantt.cell(2, 10, '=UPPER(TEXT(J3, "MMMM YYYY"))').font = Font(name="Segoe UI", size=9, bold=True, color="FE8C29")

    import openpyxl.utils as openpyxl_utils
    for col_idx in range(11, 56): # Columnas K a BD (45 días de rango visual)
        col_letter = openpyxl_utils.get_column_letter(col_idx)
        prev_letter = openpyxl_utils.get_column_letter(col_idx - 1)
        cell = ws_gantt.cell(3, col_idx, f"={prev_letter}3+1")
        cell.number_format = 'dd'
        cell.font = font_bold
        cell.fill = fill_orange
        cell.font = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')
        
        # Etiqueta de mes si cambia el mes respecto a la celda anterior
        ws_gantt.cell(2, col_idx, f'=IF(MONTH({col_letter}3)<>MONTH({prev_letter}3), UPPER(TEXT({col_letter}3, "MMMM YYYY")), "")').font = Font(name="Segoe UI", size=9, bold=True, color="FE8C29")

    # Ajustar anchos de columnas específicos
    ws_gantt.column_dimensions['A'].width = 5.5
    ws_gantt.column_dimensions['B'].width = 12.0
    ws_gantt.column_dimensions['C'].width = 12.5
    ws_gantt.column_dimensions['D'].width = 16.0
    ws_gantt.column_dimensions['E'].width = 9.0
    ws_gantt.column_dimensions['F'].width = 9.0
    ws_gantt.column_dimensions['G'].width = 35.0
    ws_gantt.column_dimensions['H'].width = 45.0
    ws_gantt.column_dimensions['I'].width = 12.0

    for col_idx in range(10, 56):
        col_letter = openpyxl_utils.get_column_letter(col_idx)
        ws_gantt.column_dimensions[col_letter].width = 4.0

    # Llenar datos de muestra
    gantt_sample = [
        (1, "Actividad", "DS", date(2026, 8, 11), 5, 1, "Levantamiento y validación técnica en sitio", ""),
        (2, "Entregable", "DS", "", 7, 1, "Diseño tridimensional y diagramas eléctricos", 1),
        (3, "Reunión", "DS", "", 1, 2, "Reunión de revisión de ingeniería con cliente", 2),
        (4, "Hito", "DS", "", 1, 2, "Hito: Aprobación final de diseño para manufactura", 3),
    ]

    for r, g in enumerate(gantt_sample, 4):
        ws_gantt.cell(r, 1, g[0]).alignment = Alignment(horizontal='center')
        ws_gantt.cell(r, 2, g[1])
        ws_gantt.cell(r, 3, g[2])
        
        # Fecha de Inicio: row 4 es estática, row 5 en adelante es auto-calculada por predecesoras
        cell_d = ws_gantt.cell(r, 4)
        if r == 4:
            cell_d.value = g[3]
        else:
            cell_d.value = f'=IF(OR(ISBLANK(I{r}), I{r}=""), $D$4, INDEX(D$4:D$200, MATCH(I{r}, A$4:A$200, 0)) + INDEX(E$4:E$200, MATCH(I{r}, A$4:A$200, 0)))'
        cell_d.number_format = 'yyyy-mm-dd'
        cell_d.alignment = Alignment(horizontal='center')
        
        # Duración con formato personalizado: x.xx "días"
        cell_e = ws_gantt.cell(r, 5, g[4])
        cell_e.number_format = '0.00" días"'
        cell_e.alignment = Alignment(horizontal='center')
        
        # Asociar a Partida con formato personalizado: "Partida N° "0
        cell_f = ws_gantt.cell(r, 6, g[5])
        cell_f.number_format = '"Partida N° "0'
        cell_f.alignment = Alignment(horizontal='center')
        
        # Fórmula de búsqueda automática de la partida (WBS) - Tamaño de texto 7
        cell_wbs = ws_gantt.cell(r, 7, f"=IFERROR(VLOOKUP(F{r}, WBS!A$4:B$200, 2, FALSE), \"\")")
        cell_wbs.font = font_wbs_size7
        cell_wbs.fill = fill_auto
        
        # Descripción
        ws_gantt.cell(r, 8, g[6]).font = font_bold
        
        # Predecesora
        ws_gantt.cell(r, 9, g[7]).alignment = Alignment(horizontal='center')

    # Formatear filas vacías preventivamente
    for r in range(8, 200):
        # Fecha de Inicio auto por fórmula
        cell_d = ws_gantt.cell(r, 4, f'=IF(OR(ISBLANK(I{r}), I{r}=""), $D$4, INDEX(D$4:D$200, MATCH(I{r}, A$4:A$200, 0)) + INDEX(E$4:E$200, MATCH(I{r}, A$4:A$200, 0)))')
        cell_d.number_format = 'yyyy-mm-dd'
        cell_d.alignment = Alignment(horizontal='center')
        
        # Formatos por defecto
        ws_gantt.cell(r, 1, r - 3).alignment = Alignment(horizontal='center')
        ws_gantt.cell(r, 5).number_format = '0.00" días"'
        ws_gantt.cell(r, 6).number_format = '"Partida N° "0'
        
        # Lookup de WBS con tamaño de letra 7
        cell_wbs = ws_gantt.cell(r, 7, f"=IFERROR(VLOOKUP(F{r}, WBS!A$4:B$200, 2, FALSE), \"\")")
        cell_wbs.font = font_wbs_size7
        cell_wbs.fill = fill_auto
        
        ws_gantt.cell(r, 9).alignment = Alignment(horizontal='center')

    # Regla de Formato Condicional para colorear el Gantt por Tipo de Tarea
    from openpyxl.formatting.rule import FormulaRule
    fill_act = PatternFill(start_color="FE8C29", end_color="FE8C29", fill_type="solid") # Orange
    fill_ent = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid") # Green
    fill_reu = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid") # Blue
    fill_hit = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid") # Red

    rule_act = FormulaRule(formula=['AND(ISNUMBER($D4), ISNUMBER($E4), $B4="Actividad", J$3>=$D4, J$3<($D4+$E4))'], fill=fill_act)
    rule_ent = FormulaRule(formula=['AND(ISNUMBER($D4), ISNUMBER($E4), $B4="Entregable", J$3>=$D4, J$3<($D4+$E4))'], fill=fill_ent)
    rule_reu = FormulaRule(formula=['AND(ISNUMBER($D4), ISNUMBER($E4), $B4="Reunión", J$3>=$D4, J$3<($D4+$E4))'], fill=fill_reu)
    rule_hit = FormulaRule(formula=['AND(ISNUMBER($D4), ISNUMBER($E4), $B4="Hito", J$3>=$D4, J$3<($D4+$E4))'], fill=fill_hit)

    ws_gantt.conditional_formatting.add("J4:BD200", rule_act)
    ws_gantt.conditional_formatting.add("J4:BD200", rule_ent)
    ws_gantt.conditional_formatting.add("J4:BD200", rule_reu)
    ws_gantt.conditional_formatting.add("J4:BD200", rule_hit)

    # ─────────────────────────────────────
    # 8. RESUMEN DE COSTO (con prorrateo de GAS)
    # ─────────────────────────────────────
    ws_res = wb.create_sheet("RESUMEN DE COSTO")
    ws_res.cell(1,1,"PASO 4 — RESUMEN DE COSTO DIRECTO CON PRORRATEO DE GASTOS GENERALES").font = font_title
    ws_res.cell(2,1,"🟠 Los Gastos Generales (GAS) se prorratean entre partidas según su % del Costo Directo Base (MAT+MO+SUB+MAQ)").font = Font(name="Segoe UI", size=9, italic=True, color="E65100")

    # Encabezados: N°Partida | WBS | MAT | MO | SUB | MAQ | Costo Directo Base | % GAS | GAS Prorrateado | TOTAL CON GAS
    apply_header_row(ws_res, [
        "N° Partida",
        "Descripción Partida WBS",
        "Materiales (MXN)",
        "Mano de Obra (MXN)",
        "Subcontratos (MXN)",
        "Maquinaria (MXN)",
        "Costo Directo Base",
        "% sobre Total",
        "GAS Prorrateado (MXN)",
        "TOTAL CON GASTOS (MXN)"
    ])

    # Filas de partidas (4 a 6 → partidas 1 a 3)
    for r_idx in range(4, 7):
        p_num = r_idx - 3
        ws_res.cell(r_idx,1,p_num).alignment = Alignment(horizontal='center')
        ws_res.cell(r_idx,2, f"=WBS!B{r_idx}").font = font_bold

        # MAT, MO, SUB, MAQ por partida (SUMIF a hojas respectivas)
        ws_res.cell(r_idx,3, f"=SUMIF(MAT!A$4:A$200,A{r_idx},MAT!G$4:G$200)").number_format = '$#,##0.00'
        ws_res.cell(r_idx,4, f"=SUMIF(MO!A$4:A$200,A{r_idx},MO!J$4:J$200)").number_format = '$#,##0.00'
        ws_res.cell(r_idx,5, f"=SUMIF(SUB!A$4:A$200,A{r_idx},SUB!I$4:I$200)").number_format = '$#,##0.00'
        ws_res.cell(r_idx,6, f"=SUMIF(MAQ!A$4:A$200,A{r_idx},MAQ!I$4:I$200)").number_format = '$#,##0.00'

        # Costo Directo Base = MAT+MO+SUB+MAQ (sin GAS)
        c_cd = ws_res.cell(r_idx,7, f"=SUM(C{r_idx}:F{r_idx})")
        c_cd.number_format = '$#,##0.00'; c_cd.font = font_bold; c_cd.fill = fill_auto

        # % sobre el total (Costo Directo Base / Suma de todos los Costo Directo Base)
        c_pct = ws_res.cell(r_idx,8, f"=IFERROR(G{r_idx}/SUM($G$4:$G$6),0)")
        c_pct.number_format = '0.00%'; c_pct.fill = fill_auto
        c_pct.font = Font(name="Segoe UI", size=10, bold=True, color="1565C0")

        # GAS prorrateado = Total GAS de hoja GAS × % de esta partida
        c_gas = ws_res.cell(r_idx,9, f"=GAS!$F$200 * H{r_idx}")
        c_gas.number_format = '$#,##0.00'; c_gas.fill = fill_calc
        c_gas.font = Font(name="Segoe UI", size=10, bold=True, color="E65100")

        # Total con GAS = Costo Directo Base + GAS prorrateado
        c_tot = ws_res.cell(r_idx,10, f"=G{r_idx}+I{r_idx}")
        c_tot.number_format = '$#,##0.00'; c_tot.font = font_bold; c_tot.fill = fill_total

    # Fila de Totales
    row_tot = 7
    ws_res.cell(row_tot,2,"TOTALES:").font = font_total
    for col, formula in [
        (3,"=SUM(C4:C6)"), (4,"=SUM(D4:D6)"), (5,"=SUM(E4:E6)"),
        (6,"=SUM(F4:F6)"), (7,"=SUM(G4:G6)"),
    ]:
        c = ws_res.cell(row_tot, col, formula)
        c.number_format = '$#,##0.00'; c.font = font_total

    c_pct_tot = ws_res.cell(row_tot,8,"=SUM(H4:H6)")
    c_pct_tot.number_format = '0.00%'; c_pct_tot.font = font_total

    c_gas_tot = ws_res.cell(row_tot,9,"=GAS!$F$200")
    c_gas_tot.number_format = '$#,##0.00'; c_gas_tot.font = font_total; c_gas_tot.fill = fill_orange
    ws_res.cell(row_tot,9).font = Font(name="Segoe UI", size=12, bold=True, color="FFFFFF")

    c_gran = ws_res.cell(row_tot,10,"=SUM(J4:J6)")
    c_gran.number_format = '$#,##0.00'; c_gran.fill = fill_orange
    c_gran.font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")

    # Nota explicativa del prorrateo
    ws_res.cell(9,2,"* GAS!F200 = Celda TOTAL de la hoja GAS (suma de todos los Gastos Generales del proyecto)").font = Font(name="Segoe UI", size=8, italic=True, color="78909C")
    ws_res.cell(10,2,"* % sobre Total = Costo Directo Base de la partida / Total Costo Directo Base del proyecto").font = Font(name="Segoe UI", size=8, italic=True, color="78909C")
    ws_res.cell(11,2,"* GAS Prorrateado = Total GAS × %, distribuyendo el gasto general de forma proporcional al peso de cada partida").font = Font(name="Segoe UI", size=8, italic=True, color="78909C")

    ws_res.column_dimensions['A'].width = 12
    ws_res.column_dimensions['B'].width = 42
    for col_l in ['C','D','E','F']:
        ws_res.column_dimensions[col_l].width = 17
    ws_res.column_dimensions['G'].width = 20
    ws_res.column_dimensions['H'].width = 14
    ws_res.column_dimensions['I'].width = 22
    ws_res.column_dimensions['J'].width = 26

    # ─────────────────────────────────────
    # 9. BD (Catálogos Base — AL FINAL)
    # ─────────────────────────────────────
    ws_bd = wb.create_sheet("BD")
    ws_bd.cell(1,1,"PUESTOS_MO").font = Font(name="Segoe UI", bold=True, color="FFFFFF"); ws_bd.cell(1,1).fill = fill_header
    ws_bd.cell(1,2,"SUELDO_BASE").font = Font(name="Segoe UI", bold=True, color="FFFFFF"); ws_bd.cell(1,2).fill = fill_header
    ws_bd.cell(1,3,"FASAR").font = Font(name="Segoe UI", bold=True, color="FFFFFF"); ws_bd.cell(1,3).fill = fill_header
    ws_bd.cell(1,4,"GASTOS_OBRA").font = Font(name="Segoe UI", bold=True, color="FFFFFF"); ws_bd.cell(1,4).fill = fill_header
    ws_bd.cell(1,5,"SUBCONTRATOS").font = Font(name="Segoe UI", bold=True, color="FFFFFF"); ws_bd.cell(1,5).fill = fill_header
    ws_bd.cell(1,6,"MAQ_EQUIPO").font = Font(name="Segoe UI", bold=True, color="FFFFFF"); ws_bd.cell(1,6).fill = fill_header
    ws_bd.cell(1,7,"CLIENTES").font = Font(name="Segoe UI", bold=True, color="FFFFFF"); ws_bd.cell(1,7).fill = fill_header

    for r, tup in enumerate(mo_tuples, 2):
        ws_bd.cell(r,1,tup[0])
        ws_bd.cell(r,2,float(tup[1])).number_format = '$#,##0.00'
        ws_bd.cell(r,3,float(tup[2])).number_format = '0.00'

    for r, val in enumerate(gastos_cat, 2):
        ws_bd.cell(r,4,val)

    for r, val in enumerate(subs_cat, 2):
        ws_bd.cell(r,5,val)

    maq_cat_names = [r[0] for r in maq_cat]
    for r, val in enumerate(maq_cat_names, 2):
        ws_bd.cell(r,6,val)

    # Usar catálogo completo de gastos como maquinaria si es vacío
    if not maq_cat_names:
        for r, val in enumerate(gastos_cat[:10], 2):
            ws_bd.cell(r,6,val)

    for r, val in enumerate(clientes_cat, 2):
        ws_bd.cell(r,7,val)

    ws_bd.column_dimensions['A'].width = 30
    ws_bd.column_dimensions['B'].width = 22
    ws_bd.column_dimensions['C'].width = 10
    ws_bd.column_dimensions['D'].width = 42
    ws_bd.column_dimensions['E'].width = 38
    ws_bd.column_dimensions['F'].width = 42
    ws_bd.column_dimensions['G'].width = 32

    # ── Validaciones de Datos (Dropdowns) ──
    mo_end  = len(mo_tuples) + 1
    gas_end = len(gastos_cat) + 1
    sub_end = len(subs_cat) + 1
    maq_end = len(maq_cat_names) + 1 if maq_cat_names else 50
    cli_end = len(clientes_cat) + 1

    dv_cli    = DataValidation(type="list", formula1=f"BD!$G$2:$G${cli_end}", allow_blank=True)
    dv_mo     = DataValidation(type="list", formula1=f"BD!$A$2:$A${mo_end}",  allow_blank=True)
    dv_gastos = DataValidation(type="list", formula1=f"BD!$D$2:$D${gas_end}", allow_blank=True)
    dv_subs   = DataValidation(type="list", formula1=f"BD!$E$2:$E${sub_end}", allow_blank=True)
    dv_maq    = DataValidation(type="list", formula1=f"BD!$F$2:$F${maq_end}", allow_blank=True)
    dv_moneda = DataValidation(type="list", formula1='"MXN,USD"', allow_blank=False)
    dv_gantt_tipo = DataValidation(type="list", formula1='"Actividad,Entregable,Hito,Reunión"', allow_blank=False)

    ws_proy.add_data_validation(dv_cli);    dv_cli.add("B4")
    ws_mo.add_data_validation(dv_mo);       dv_mo.add("C4:C200")
    ws_gastos.add_data_validation(dv_gastos); dv_gastos.add("A4:A200")   # GAS: col A = Concepto
    ws_subs.add_data_validation(dv_subs);   dv_subs.add("C4:C200")
    ws_maq.add_data_validation(dv_maq);     dv_maq.add("C4:C200")
    ws_subs.add_data_validation(dv_moneda); dv_moneda.add("H4:H200")
    ws_maq.add_data_validation(dv_moneda);  dv_moneda.add("H4:H200")
    ws_gantt.add_data_validation(dv_gantt_tipo); dv_gantt_tipo.add("B4:B200")


    # Color de pestañas (tabs) según especificación visual
    ws_proy.sheet_properties.tabColor = "000000"       # PROY - Black
    ws_wbs.sheet_properties.tabColor  = "002060"       # WBS - Dark Navy Blue
    ws_mat.sheet_properties.tabColor  = "CCC0DA"       # MAT - Lilac / Light Purple
    ws_mo.sheet_properties.tabColor   = "92D050"       # MO - Lime Green
    ws_gastos.sheet_properties.tabColor = "00B0F0"     # GAS - Cyan / Light Blue
    ws_subs.sheet_properties.tabColor   = "FFC000"     # SUB - Gold / Yellow
    ws_maq.sheet_properties.tabColor    = "C4BD97"     # MAQ - Khaki
    ws_gantt.sheet_properties.tabColor  = "4F81BD"     # PLAN - Steel Blue
    ws_res.sheet_properties.tabColor    = "FF0000"     # RESUMEN DE COSTO - Red
    ws_bd.sheet_properties.tabColor     = "A6A6A6"     # BD - Grey

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# 2. AUDITORÍA Y PARSER DE ARCHIVOS EXCEL CARGADOS (ADAPTATIVO)
# ─────────────────────────────────────────────────────────────────────────────

def auditar_y_parsear_excel(file_bytes, filename):
    """
    Audita y valida celda por celda la estructura de la cotización cargada por el presupuestador.
    Soporta los nuevos nombres de hoja (PROY, WBS, 3_MATERIALES, 4_MANO_DE_OBRA, BD, etc.).
    Retorna: (is_valid, audit_df, parsed_data_dict, summary_msg)
    """
    init_db()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        return False, pd.DataFrame(), {}, f"Error al abrir el archivo Excel: {str(e)}"

    audit_rows = []
    parsed_data = {
        "generales": {},
        "partidas": [],
        "materiales": [],
        "mo": [],
        "gastos": [],
        "subcontratos": [],
        "maquinaria": []
    }

    # 1. DATOS GENERALES (HOJA PROY O SIMILAR)
    g_sheets = [s for s in wb.sheetnames if "PROY" in s.upper() or "GENERAL" in s.upper() or "DATOS" in s.upper()]
    if g_sheets:
        ws_g = wb[g_sheets[0]]
        for r in range(1, ws_g.max_row + 1):
            lbl = str(ws_g.cell(r, 1).value or "").strip()
            val = ws_g.cell(r, 2).value
            if "Cliente" in lbl: parsed_data["generales"]["cliente"] = str(val or "TREBOTTI").strip()
            elif "Proyecto" in lbl: parsed_data["generales"]["proyecto"] = str(val or "PROYECTO AUTOMATIZACIÓN").strip()
            elif "Ingeniero" in lbl or "Presupuestador" in lbl: parsed_data["generales"]["ingeniero"] = str(val or "DS").strip()
            elif "Cambio" in lbl or "USD" in lbl:
                try: parsed_data["generales"]["tc"] = float(val or 18.0)
                except: parsed_data["generales"]["tc"] = 18.0
    else:
        parsed_data["generales"] = {"cliente": "TREBOTTI", "proyecto": "PROYECTO AUTOMATIZACIÓN", "tc": 18.0}

    audit_rows.append({
        "Hoja / Sección": "PROY",
        "Elemento Auditado": "Datos del Proyecto",
        "Estado": "✅ CORRECTO",
        "Observaciones": f"Cliente: {parsed_data['generales'].get('cliente')} | Proyecto: {parsed_data['generales'].get('proyecto')}"
    })

    # 2. PARTIDAS WBS (HOJA WBS O SIMILAR)
    wbs_sheets = [s for s in wb.sheetnames if "WBS" in s.upper() or "PARTIDA" in s.upper() or "ESTRUCTURA" in s.upper()]
    if wbs_sheets:
        ws_p = wb[wbs_sheets[0]]
        start_row = 4 if ws_p.cell(3, 1).value else 2
        for r in range(start_row, ws_p.max_row + 1):
            np = ws_p.cell(r, 1).value
            desc = ws_p.cell(r, 2).value
            alcance = ws_p.cell(r, 3).value or ""
            if np and desc:
                try:
                    np = int(np)
                    parsed_data["partidas"].append({"numero_partida": np, "descripcion": str(desc).strip(), "alcance": str(alcance).strip()})
                    audit_rows.append({"Hoja / Sección": "WBS", "Elemento Auditado": f"Partida WBS {np}", "Estado": "✅ CORRECTO", "Observaciones": str(desc)[:40]})
                except Exception: pass

    # 3. MATERIALES (MAT / 3_MATERIALES)
    m_sheet_key = [s for s in wb.sheetnames if "MATERIA" in s.upper() or s.upper() == "MAT"]
    if m_sheet_key:
        ws_m = wb[m_sheet_key[0]]
        start_row = 4 if ws_m.cell(3, 1).value else 2
        col_offset = 1 if "Partida WBS" in str(ws_m.cell(3, 2).value or ws_m.cell(1, 2).value or "") or "Nombre Partida" in str(ws_m.cell(3, 2).value or "") else 0
        for r in range(start_row, ws_m.max_row + 1):
            np = ws_m.cell(r, 1).value
            conc = ws_m.cell(r, 2 + col_offset).value
            esp = ws_m.cell(r, 3 + col_offset).value or ""
            cant = ws_m.cell(r, 4 + col_offset).value or 1
            unid = ws_m.cell(r, 5 + col_offset).value or "PZA"
            pu = ws_m.cell(r, 6 + col_offset).value or 0.0
            mon = ws_m.cell(r, 7 + col_offset).value or "MXN"
            if np and conc:
                try:
                    np = int(np); cant = float(cant); pu = float(pu)
                    parsed_data["materiales"].append({"partida_num": np, "concepto": str(conc).strip(), "especificacion": str(esp).strip(), "cantidad": cant, "unidad": str(unid).strip(), "pu": pu, "moneda": str(mon).upper().strip()})
                    audit_rows.append({"Hoja / Sección": "3_MATERIALES", "Elemento Auditado": f"Material Partida {np}", "Estado": "✅ CORRECTO", "Observaciones": f"{conc} ({cant} {unid} @ ${pu:,.2f} {mon})"})
                except Exception: pass

    # 4. MANO DE OBRA (MO / 4_MANO_DE_OBRA)
    mo_sheet_key = [s for s in wb.sheetnames if "MANO" in s.upper() or s.upper() == "MO"]
    if mo_sheet_key:
        ws_mo = wb[mo_sheet_key[0]]
        start_row = 4 if ws_mo.cell(3, 1).value else 2
        col_offset = 1 if "Partida WBS" in str(ws_mo.cell(3, 2).value or ws_mo.cell(1, 2).value or "") or "Nombre Partida" in str(ws_mo.cell(3, 2).value or "") else 0
        for r in range(start_row, ws_mo.max_row + 1):
            np = ws_mo.cell(r, 1).value
            puesto = ws_mo.cell(r, 2 + col_offset).value
            pers = ws_mo.cell(r, 3 + col_offset).value or 1
            hrs = ws_mo.cell(r, 4 + col_offset).value or 8
            dias = ws_mo.cell(r, 5 + col_offset).value or 1
            if np and puesto:
                try:
                    np = int(np); pers = int(pers); hrs = float(hrs); dias = float(dias)
                    parsed_data["mo"].append({"partida_num": np, "puesto": str(puesto).strip(), "personas": pers, "horas": hrs, "dias": dias})
                    audit_rows.append({"Hoja / Sección": "4_MANO_DE_OBRA", "Elemento Auditado": f"MO Partida {np}", "Estado": "✅ CORRECTO", "Observaciones": f"{puesto} ({pers} pers x {dias}d)"})
                except Exception: pass

    # 5. GASTOS GENERALES (GAS / 5_GASTOS_GENERALES) — ahora son globales (sin partida)
    g_sheet_key = [s for s in wb.sheetnames if "GASTO" in s.upper() or s.upper() == "GAS"]
    if g_sheet_key:
        ws_g = wb[g_sheet_key[0]]
        # Detectar si la hoja es el nuevo formato (col A = Concepto, sin Partida)
        # o el formato viejo (col A = Partida N°, col B/C = concepto)
        h1_val = str(ws_g.cell(3, 1).value or ws_g.cell(1, 1).value or "").strip().upper()
        is_global_format = "CONCEPTO" in h1_val or "GASTO" in h1_val  # nuevo formato sin partida
        start_row = 4 if ws_g.cell(3, 1).value else 2

        if is_global_format:
            # Nuevo formato global: Concepto(A) | Cantidad(B) | Unidad(C) | Tiempo(D) | CostoUnitario(E) | Importe(F)
            for r in range(start_row, ws_g.max_row + 1):
                concepto = ws_g.cell(r, 1).value
                if concepto and str(concepto).strip() and "TOTAL" not in str(concepto).upper():
                    try:
                        cant  = float(ws_g.cell(r, 2).value or 1.0)
                        unid  = str(ws_g.cell(r, 3).value or "VJE").strip()
                        tiempo = float(ws_g.cell(r, 4).value or 1.0)
                        cu    = float(ws_g.cell(r, 5).value or 0.0)
                        # partida_num = None indica que es un gasto global a prorratear
                        parsed_data["gastos"].append({"partida_num": None, "concepto": str(concepto).strip(), "cantidad": cant, "unidad": unid, "tiempo": tiempo, "costo_unitario": cu})
                        audit_rows.append({"Hoja / Sección": "GAS", "Elemento Auditado": "Gasto General Global", "Estado": "✅ CORRECTO", "Observaciones": f"{concepto} (${cu:,.2f})"})
                    except Exception: pass
        else:
            # Formato viejo con Partida N° en col A
            col_offset = 1 if "Partida WBS" in str(ws_g.cell(3, 2).value or ws_g.cell(1, 2).value or "") else 0
            for r in range(start_row, ws_g.max_row + 1):
                np = ws_g.cell(r, 1).value
                concepto = ws_g.cell(r, 2 + col_offset).value
                cant  = ws_g.cell(r, 3 + col_offset).value or 1.0
                unid  = ws_g.cell(r, 4 + col_offset).value or "VJE"
                tiempo = ws_g.cell(r, 5 + col_offset).value or 1.0
                cu    = ws_g.cell(r, 6 + col_offset).value or 0.0
                if np and concepto:
                    try:
                        np = int(np); cant = float(cant); tiempo = float(tiempo); cu = float(cu)
                        parsed_data["gastos"].append({"partida_num": np, "concepto": str(concepto).strip(), "cantidad": cant, "unidad": str(unid).strip(), "tiempo": tiempo, "costo_unitario": cu})
                        audit_rows.append({"Hoja / Sección": "GAS", "Elemento Auditado": f"Gasto Partida {np}", "Estado": "✅ CORRECTO", "Observaciones": f"{concepto} (${cu:,.2f})"})
                    except Exception: pass


    # 6. SUBCONTRATOS (SUB / 6_SUBCONTRATOS)
    sub_sheet_key = [s for s in wb.sheetnames if "SUBCONTRATO" in s.upper() or s.upper() == "SUB"]
    if sub_sheet_key:
        ws_s = wb[sub_sheet_key[0]]
        start_row = 4 if ws_s.cell(3, 1).value else 2
        col_offset = 1 if "Partida WBS" in str(ws_s.cell(3, 2).value or ws_s.cell(1, 2).value or "") or "Nombre Partida" in str(ws_s.cell(3, 2).value or "") else 0
        for r in range(start_row, ws_s.max_row + 1):
            np = ws_s.cell(r, 1).value
            concepto = ws_s.cell(r, 2 + col_offset).value
            prov = ws_s.cell(r, 3 + col_offset).value or "General"
            cant = ws_s.cell(r, 4 + col_offset).value or 1.0
            unid = ws_s.cell(r, 5 + col_offset).value or "SERV"
            pu = ws_s.cell(r, 6 + col_offset).value or 0.0
            mon = ws_s.cell(r, 7 + col_offset).value or "MXN"
            if np and concepto:
                try:
                    np = int(np); cant = float(cant); pu = float(pu)
                    parsed_data["subcontratos"].append({"partida_num": np, "concepto": str(concepto).strip(), "proveedor": str(prov).strip(), "cantidad": cant, "unidad": str(unid).strip(), "pu": pu, "moneda": str(mon).upper().strip()})
                    audit_rows.append({"Hoja / Sección": "6_SUBCONTRATOS", "Elemento Auditado": f"Subcontrato Partida {np}", "Estado": "✅ CORRECTO", "Observaciones": f"{concepto} ({prov})"})
                except Exception: pass

    # 7. MAQUINARIA (MAQ / 7_MAQUINARIA)
    maq_sheet_key = [s for s in wb.sheetnames if "MAQUINARIA" in s.upper() or s.upper() == "MAQ"]
    if maq_sheet_key:
        ws_mq = wb[maq_sheet_key[0]]
        start_row = 4 if ws_mq.cell(3, 1).value else 2
        col_offset = 1 if "Partida WBS" in str(ws_mq.cell(3, 2).value or ws_mq.cell(1, 2).value or "") or "Nombre Partida" in str(ws_mq.cell(3, 2).value or "") else 0
        for r in range(start_row, ws_mq.max_row + 1):
            np = ws_mq.cell(r, 1).value
            concepto = ws_mq.cell(r, 2 + col_offset).value
            mod = ws_mq.cell(r, 3 + col_offset).value or ""
            tiempo = ws_mq.cell(r, 4 + col_offset).value or 1.0
            unid = ws_mq.cell(r, 5 + col_offset).value or "DÍA"
            pu = ws_mq.cell(r, 6 + col_offset).value or 0.0
            mon = ws_mq.cell(r, 7 + col_offset).value or "MXN"
            if np and concepto:
                try:
                    np = int(np); tiempo = float(tiempo); pu = float(pu)
                    parsed_data["maquinaria"].append({"partida_num": np, "concepto": str(concepto).strip(), "modelo": str(mod).strip(), "tiempo": tiempo, "unidad": str(unid).strip(), "pu": pu, "moneda": str(mon).upper().strip()})
                    audit_rows.append({"Hoja / Sección": "7_MAQUINARIA", "Elemento Auditado": f"Maquinaria Partida {np}", "Estado": "✅ CORRECTO", "Observaciones": f"{concepto} ({mod})"})
                except Exception: pass

    # 8. PLAN DE PROYECTO / CRONOGRAMA DE GANTT (PLAN_PROYECTO)
    gantt_sheet_key = [s for s in wb.sheetnames if "PLAN" in s.upper() or "CRONOGRAMA" in s.upper() or "GANTT" in s.upper()]
    parsed_data["gantt"] = []
    if gantt_sheet_key:
        ws_gantt = wb[gantt_sheet_key[0]]
        start_row = 4 if ws_gantt.cell(3, 1).value else 2
        for r in range(start_row, ws_gantt.max_row + 1):
            act = ws_gantt.cell(r, 8).value
            tipo = ws_gantt.cell(r, 2).value or "Actividad"
            resp = ws_gantt.cell(r, 3).value or "DS"
            f_ini = ws_gantt.cell(r, 4).value
            dur = ws_gantt.cell(r, 5).value or 1
            partida_num = ws_gantt.cell(r, 6).value
            pred = ws_gantt.cell(r, 9).value
            
            if act and str(act).strip():
                try:
                    # Limpieza robusta de duración
                    try:
                        dur_val = float(dur)
                    except:
                        import re
                        num_match = re.search(r'[\d\.]+', str(dur))
                        dur_val = float(num_match.group(0)) if num_match else 1.0
                    dur_val = int(dur_val)
                    
                    # Limpieza robusta de número de partida
                    try:
                        partida_val = int(partida_num)
                    except:
                        import re
                        num_match = re.search(r'\d+', str(partida_num or ""))
                        partida_val = int(num_match.group(0)) if num_match else None
                        
                    if isinstance(f_ini, (datetime, date)):
                        f_ini_str = f_ini.strftime("%Y-%m-%d")
                    else:
                        f_ini_str = str(f_ini or "").strip()
                        
                    # Limpieza de predecesora
                    try:
                        pred_val = int(pred) if pred else None
                    except:
                        import re
                        num_match = re.search(r'\d+', str(pred or ""))
                        pred_val = int(num_match.group(0)) if num_match else None
                        
                    parsed_data["gantt"].append({
                        "actividad": str(act).strip(),
                        "tipo": str(tipo).strip(),
                        "responsable": str(resp).strip(),
                        "fecha_inicio": f_ini_str,
                        "dias_duracion": dur_val,
                        "partida_num": partida_val,
                        "predecesora_id": pred_val
                    })
                    audit_rows.append({
                        "Hoja / Sección": "PLAN",
                        "Elemento Auditado": f"Tarea: {act}",
                        "Estado": "✅ CORRECTO",
                        "Observaciones": f"Duración: {dur_val}d | Predecesora: {pred_val or '—'}"
                    })
                except Exception:
                    pass

    audit_df = pd.DataFrame(audit_rows)
    is_valid = len(parsed_data["partidas"]) > 0
    msg = f"Auditoría exitosa para el archivo '{filename}'. Se detectaron {len(parsed_data['partidas'])} partida(s) listas para la gestión de ventas." if is_valid else "Error: No se detectaron partidas válidas en el archivo Excel."

    return is_valid, audit_df, parsed_data, msg


# ─────────────────────────────────────────────────────────────────────────────
# 3. GUARDADO DE COTIZACIÓN AUDITADA CON PARÁMETROS COMERCIALES DE VENTAS
# ─────────────────────────────────────────────────────────────────────────────

def guardar_cotizacion_auditada(parsed_data, comercial_params=None):
    """
    Inserta la cotización traída del presupuestador en SQLite e integra los parámetros comerciales de ventas
    (Margen %, Comisión %, Términos de Pago, Moneda, etc.), ejecutando el recálculo financiero de prorrateo.
    """
    init_db()
    conn = get_connection(); cur = conn.cursor()

    gen = parsed_data.get("generales", {})
    cliente_nombre = gen.get("cliente", "TREBOTTI").strip()
    proyecto = gen.get("proyecto", "PROYECTO AUTOMATIZACIÓN").strip()

    if not comercial_params: comercial_params = {}

    margen_pct = comercial_params.get("margen_pct", 0.30)
    comision_pct = comercial_params.get("comision_pct", 0.05)
    supervision_pct = comercial_params.get("supervision_pct", 0.30)
    herramienta_pct = comercial_params.get("herramienta_pct", 0.03)

    cond_pago = comercial_params.get("condiciones_pago", "CREDITO")
    entrega = comercial_params.get("tiempo_entrega", "2 SEMANAS")
    vigencia = comercial_params.get("vigencia_cotizacion", "15 días")
    moneda_cot = comercial_params.get("moneda_cotizacion", "MXN pesos mexicanos")

    cur.execute("SELECT id, acronimo FROM clientes WHERE nombre LIKE ?", (f"%{cliente_nombre}%",))
    r_cli = cur.fetchone()
    cliente_id = r_cli[0] if r_cli else None
    acronimo = (r_cli[1] if r_cli and r_cli[1] else cliente_nombre[:3].upper()) or "COT"
    if not cliente_id:
        cur.execute("INSERT INTO clientes (nombre, acronimo) VALUES (?, ?)", (cliente_nombre, acronimo))
        cliente_id = cur.lastrowid

    # Generación automática del Folio Oficial J&D según la nueva regla:
    # COT - [Contador Secuencial] - [Acrónimo Cliente 3 letras] - [Presupuestador 2 letras] - [Proyecto]_Cotizacion_Oficial
    cur.execute("SELECT COUNT(*) FROM cotizaciones")
    cnt = (cur.fetchone()[0] or 0) + 1
    
    cli_3 = acronimo[:3].upper() if acronimo else cliente_nombre[:3].upper()
    if not cli_3:
        cli_3 = "CLI"
        
    ing_init = gen.get("ingeniero", "DS")[:2].upper()
    if not ing_init:
        ing_init = "DS"
        
    proj_desc = proyecto.strip().upper() if proyecto else "PROYECTO"
    
    folio = f"COT-{cnt:03d}-{cli_3}-{ing_init}-{proj_desc}_Cotizacion_Oficial"

    hitos_json = comercial_params.get("hitos_pago_json", "[]")

    try:
        cur.execute("""
            INSERT INTO cotizaciones 
            (folio, cliente_id, proyecto, tipo_cambio_usd, margen_porcentaje, comision_porcentaje, supervision_porcentaje, herramienta_porcentaje, estatus, condiciones_pago, tiempo_entrega, vigencia_cotizacion, moneda_cotizacion, hitos_pago_json)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'Borrador', ?, ?, ?, ?, ?)
        """, (folio, cliente_id, proyecto, gen.get("tc", 18.0), margen_pct, comision_pct, supervision_pct, herramienta_pct, cond_pago, entrega, vigencia, moneda_cot, hitos_json))
        cot_id = cur.lastrowid
    except Exception:
        cur.execute("UPDATE cotizaciones SET cliente_id=?, proyecto=?, margen_porcentaje=?, comision_porcentaje=?, condiciones_pago=?, tiempo_entrega=?, vigencia_cotizacion=?, moneda_cotizacion=?, hitos_pago_json=? WHERE folio=?",
                    (cliente_id, proyecto, margen_pct, comision_pct, cond_pago, entrega, vigencia, moneda_cot, hitos_json, folio))
        cur.execute("SELECT id FROM cotizaciones WHERE folio=?", (folio,))
        cot_id = cur.fetchone()[0]

    partida_id_map = {}
    for p in parsed_data.get("partidas", []):
        cur.execute("""
            INSERT INTO cotizacion_partidas (cotizacion_id, numero_partida, descripcion)
            VALUES (?, ?, ?)
        """, (cot_id, p["numero_partida"], p["descripcion"]))
        partida_id_map[p["numero_partida"]] = cur.lastrowid

    first_pid = list(partida_id_map.values())[0] if partida_id_map else None

    for m in parsed_data.get("materiales", []):
        pid = partida_id_map.get(m["partida_num"], first_pid)
        if pid:
            pu_mxn = (m["pu"] * gen.get("tc", 18.0)) if m["moneda"] == "USD" else m["pu"]
            imp_mxn = m["cantidad"] * pu_mxn
            cur.execute("""
                INSERT INTO cotizacion_materiales_detalle
                (cotizacion_id, partida_id, descripcion, codigo, cantidad, unidad, precio_unitario_mxn, importe_mxn)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (cot_id, pid, m["concepto"], m["especificacion"], m["cantidad"], m["unidad"], pu_mxn, imp_mxn))

    for mo in parsed_data.get("mo", []):
        pid = partida_id_map.get(mo["partida_num"], first_pid)
        if pid:
            puesto = mo["puesto"]
            cur.execute("SELECT sueldo_base_semanal, fasar FROM catalogo_mano_obra WHERE categoria=?", (puesto,))
            row_p = cur.fetchone()
            s_base = float(row_p[0]) if row_p else 3500.0
            fasar = float(row_p[1]) if row_p else 1.45
            costo_h = (s_base * fasar / 48.0)
            hrs_tot = mo["personas"] * mo["horas"] * mo["dias"]
            semanas = (mo["horas"] * mo["dias"]) / 48.0
            total_mo = hrs_tot * costo_h
            cur.execute("""
                INSERT INTO cotizacion_mo_detalle
                (cotizacion_id, partida_id, categoria_nombre, cantidad_personal, sueldo_base_semanal, fasar, semanas, horas_hombre, importe_total)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (cot_id, pid, puesto, mo["personas"], s_base, fasar, semanas, hrs_tot, total_mo))

    for g in parsed_data.get("gastos", []):
        imp_g = g["cantidad"] * g["tiempo"] * g["costo_unitario"]
        # Si partida_num es None = gasto global a prorratear (nuevo formato)
        # Si partida_num tiene valor = gasto asignado a partida (formato viejo compatible)
        pid_g = partida_id_map.get(g["partida_num"]) if g["partida_num"] is not None else None
        cur.execute("""
            INSERT INTO cotizacion_gastos_detalle
            (cotizacion_id, nombre, cantidad, unidad, tiempo_valor, costo_unitario, importe_total)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (cot_id, g["concepto"], g["cantidad"], g["unidad"], g["tiempo"], g["costo_unitario"], imp_g))


    for s in parsed_data.get("subcontratos", []):
        pid = partida_id_map.get(s["partida_num"], first_pid)
        if pid:
            pu_s_mxn = (s["pu"] * gen.get("tc", 18.0)) if s["moneda"] == "USD" else s["pu"]
            imp_s = s["cantidad"] * pu_s_mxn
            cur.execute("""
                INSERT INTO cotizacion_subcontratos_detalle
                (cotizacion_id, partida_id, descripcion, subcontratista, cantidad, unidad, pu_mxn, importe_mxn)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (cot_id, pid, s["concepto"], s["proveedor"], s["cantidad"], s["unidad"], pu_s_mxn, imp_s))

    for mq in parsed_data.get("maquinaria", []):
        pid = partida_id_map.get(mq["partida_num"], first_pid)
        if pid:
            pu_mq_mxn = (mq["pu"] * gen.get("tc", 18.0)) if mq["moneda"] == "USD" else mq["pu"]
            imp_mq = mq["tiempo"] * pu_mq_mxn
            cur.execute("""
                INSERT INTO cotizacion_maquinaria_detalle
                (cotizacion_id, partida_id, nombre, clave, cantidad, unidad, costo_unitario, total_mxn)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (cot_id, pid, mq["concepto"], mq["modelo"], mq["tiempo"], mq["unidad"], pu_mq_mxn, imp_mq))

    # 8. Guardar cronograma de Gantt si existe en la importación
    cur.execute("DELETE FROM cotizacion_gantt WHERE cotizacion_id=?", (cot_id,))
    excel_task_map = {}
    
    # Primera pasada: insertar tareas y mapear N° Tarea de Excel -> SQLite ID insertado
    for idx, g in enumerate(parsed_data.get("gantt", []), 1):
        pid = partida_id_map.get(g["partida_num"]) if g.get("partida_num") else None
        f_ini_str = g.get("fecha_inicio", "")
        try:
            datetime.strptime(f_ini_str, "%Y-%m-%d")
        except:
            f_ini_str = datetime.now().strftime("%Y-%m-%d")
            
        cur.execute("""
            INSERT INTO cotizacion_gantt
            (cotizacion_id, partida_id, actividad, tipo, responsable, fecha_inicio, dias_duracion, orden)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (cot_id, pid, g["actividad"], g["tipo"], g["responsable"], f_ini_str, g["dias_duracion"], idx))
        
        inserted_id = cur.lastrowid
        excel_task_map[idx] = inserted_id

    # Segunda pasada: actualizar predecesora_id usando el mapeo de relaciones
    for idx, g in enumerate(parsed_data.get("gantt", []), 1):
        pred_idx = g.get("predecesora_id")
        if pred_idx and pred_idx in excel_task_map:
            pred_sqlite_id = excel_task_map[pred_idx]
            task_sqlite_id = excel_task_map[idx]
            cur.execute("UPDATE cotizacion_gantt SET predecesora_id=? WHERE id=?", (pred_sqlite_id, task_sqlite_id))

    conn.commit()
    conn.close()

    sync_cotizacion_totals(cot_id)
    return cot_id, folio


# ─────────────────────────────────────────────────────────────────────────────
# 4. INTERFAZ STREAMLIT PRINCIPAL DEL IMPORTADOR & GESTIÓN DE VENTAS
# ─────────────────────────────────────────────────────────────────────────────

def render_excel_importer():
    st.markdown(f"""
    <div class="jd-section-header">
        <h2>📥 Importador Excel & Gestión Comercial de Ventas</h2>
        <p>Flujo en 2 Roles: 1. Presupuestador trabaja 100% en la plantilla Excel con catálogos en hoja 'BD', sueldos FASAR y búsqueda automática de Partidas WBS. 2. Ventas carga el archivo y genera la cotización oficial.</p>
    </div>""", unsafe_allow_html=True)

    tab_imp1, tab_imp2 = st.tabs([
        "📥 1. Descargar Plantilla Oficial para el Presupuestador (.xlsx)",
        "💼 2. Cargar Excel del Presupuestador & Aplicar Visión de Ventas"
    ])

    with tab_imp1:
        st.markdown(f"""
        <div style="background:{BRAND_WHITE};border:1px solid {BRAND_BORDER_LIGHT};border-left:5px solid {BRAND_ORANGE};
                    border-radius:10px;padding:20px;margin-bottom:20px;font-family:'Montserrat',sans-serif;">
            <h3 style="font-size:16px;font-weight:800;color:{BRAND_CHARCOAL};margin:0 0 8px 0;">PLANTILLA EXCEL ROBUSTA DE PRESUPUESTOS (HOJAS: PROY, WBS, DETALLES Y BD)</h3>
            <p style="font-size:12px;color:{BRAND_CHARCOAL_MED};line-height:1.6;margin:0 0 14px 0;">
                • <b>1. PROY</b>: Selección de Cliente, Proyecto y Parámetros (sin folio manual).<br>
                • <b>2. WBS</b>: Definición de Partidas del proyecto.<br>
                • <b>3. Detalle por Partida</b>: Materiales, Mano de Obra (con Sueldo Base y FASAR), Gastos, Subcontratos y Maquinaria con fórmula automática VLOOKUP del Nombre Partida WBS.<br>
                • <b>4. BD (Al Final)</b>: Hoja de Catálogos Base (Puestos MO, Sueldos, FASAR, Gastos, Subcontratos y Clientes).
            </p>
        </div>
        """, unsafe_allow_html=True)

        plantilla_bytes = generar_plantilla_excel_oficial_bytes()
        time_tag = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Guardar copia fresca en disco local
        local_copy_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Plantilla_Oficial_Presupuestos_JND_ACTUALIZADA.xlsx")
        try:
            with open(local_copy_path, "wb") as f:
                f.write(plantilla_bytes)
        except Exception: pass

        # ─── 1. DESCARGA DE PLANTILLA EXCEL (FORMATO CORPORATIVO CON FOLIO SIGUIENTE + VERSIÓN) ───
        try:
            conn_fl = get_connection()
            cur_fl = conn_fl.cursor()
            cur_fl.execute("SELECT COUNT(*) FROM cotizaciones")
            next_num = (cur_fl.fetchone()[0] or 0) + 1
            conn_fl.close()
        except Exception:
            next_num = 1

        folio_next_str = f"{next_num:04d}"
        date_str = datetime.now().strftime("%Y%m%d")
        official_filename = f"JD-PLANTILLA-COT-{folio_next_str}-v2.0_{date_str}.xlsx"

        d1, d2 = st.columns([2, 1])
        with d1:
            st.download_button(
                label=f"📥 DESCARGAR PLANTILLA EXCEL ({official_filename})",
                data=plantilla_bytes,
                file_name=official_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
                key=f"btn_dl_plantilla_{folio_next_str}_{date_str}"
            )
        with d2:
            st.info(f"💡 Folio Siguiente sugerido: **COT-{folio_next_str}**. El Presupuestador llena el Excel localmente y entrega el archivo al Vendedor.")


        st.markdown("<div style='margin-top:16px;'></div>", unsafe_allow_html=True)

        # ─── 2. SECCIÓN DESTACADA DEL MANUAL OFICIAL EN PDF (COMO PIDIÓ EL USUARIO) ───
        try:
            from generate_pdf_manual import obtener_manual_pdf_bytes
            pdf_manual_data = obtener_manual_pdf_bytes()
        except Exception:
            pdf_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Manual_Usuario_Plantilla_Excel_JD.pdf")
            pdf_manual_data = open(pdf_path, "rb").read() if os.path.exists(pdf_path) else b""

        col_m1, col_m2 = st.columns([3.2, 1.2])
        with col_m1:
            st.markdown(f"""
            <div style="background:{BRAND_GRAY_BG};border:2px solid {BRAND_ORANGE};border-radius:10px;padding:14px 18px;">
                <h4 style="margin:0 0 6px 0;color:{BRAND_CHARCOAL};font-weight:900;font-size:14px;">📘 MANUAL OFICIAL EN PDF PARA PRESUPUESTADORES (PLANTILLA EXCEL V2.0)</h4>
                <p style="margin:0;color:{BRAND_CHARCOAL_MED};font-size:12px;line-height:1.4;">Documento técnico ilustrado paso a paso: reglas de captura, listas desplegables, sueldos FASAR automáticos y prorrateo de gastos.</p>
            </div>
            """, unsafe_allow_html=True)
        with col_m2:
            st.download_button(
                label="📄 DESCARGAR MANUAL EN PDF",
                data=pdf_manual_data,
                file_name="Manual_Usuario_Plantilla_Excel_JD.pdf",
                mime="application/pdf",
                type="primary",
                use_container_width=True,
                key=f"btn_dl_manual_imp_pdf_{time_tag}"
            )



        st.markdown(f"""
        <div style="background:#F8FAFC;border:1px solid #CBD5E1;border-radius:8px;padding:12px;margin-top:10px;">
            <p style="font-size:12px;color:#434E62;margin:0;">
                <b> ESTRUCTURA CONFIRMADA DEL LIBRO GENERADO:</b><br>
                1. <code>PROY</code> (Sin Folio manual) &bull; 2. <code>WBS</code> (Partidas) &bull; 3. <code>3_MATERIALES</code> (Partida WBS automatizada)<br>
                4. <code>4_MANO_DE_OBRA</code> (Partida WBS, Sueldo Base, FASAR % y Costo/Hora) &bull; 5. <code>5_GASTOS_GENERALES</code><br>
                6. <code>6_SUBCONTRATOS</code> &bull; 7. <code>7_MAQUINARIA</code> &bull; 8. <code>4_RESUMEN</code> (Fórmulas =SUMIF)<br>
                9. <code>BD</code> ( Catálogos Base al final del libro)
            </p>
        </div>
        """, unsafe_allow_html=True)


    with tab_imp2:
        uploaded_file = st.file_uploader(
            "Seleccionar archivo Excel completado por el Presupuestador (.xlsx)",
            type=["xlsx"],
            key="excel_uploader_audit"
        )

        if uploaded_file:
            file_bytes = uploaded_file.getvalue()
            filename = uploaded_file.name

            is_valid, audit_df, parsed_data, msg = auditar_y_parsear_excel(file_bytes, filename)

            if is_valid:
                st.success(f"✅ **{msg}**")
            else:
                st.error(f"❌ **{msg}**")

            st.markdown(f"<p style='font-size:12px;font-weight:800;color:{BRAND_CHARCOAL};margin:16px 0 8px 0;'>INSPECCIÓN CELDA POR CELDA DE COSTO DIRECTO (PRESUPUESTADOR)</p>", unsafe_allow_html=True)
            if not audit_df.empty:
                st.dataframe(audit_df, use_container_width=True, hide_index=True)

            if is_valid:
                st.divider()
                gen = parsed_data.get("generales", {})


                st.markdown(f"""
                <div style="background:{BRAND_GRAY_BG};border:2px solid {BRAND_ORANGE};border-radius:12px;padding:20px;margin-bottom:20px;">
                    <h3 style="font-size:16px;font-weight:900;color:{BRAND_CHARCOAL};margin:0 0 10px 0;">💼 CONFIGURACIÓN DE VISIÓN DE VENTAS & PARÁMETROS COMERCIALES</h3>
                    <p style="font-size:12px;color:{BRAND_CHARCOAL_MED};margin:0;">
                        Defina los márgenes de utilidad, comisión comercial y términos legales para generar la cotización de venta oficial.
                    </p>
                </div>
                """, unsafe_allow_html=True)

                # ── Recuperar condiciones del cliente (última cotización del mismo cliente) ──
                cliente_nombre_gen = gen.get("cliente", "").strip()
                _cond_default = "CONTADO"
                _vigencia_default = "30 días"
                try:
                    from database.models import get_connection as _gc
                    _c = _gc(); _cur = _c.cursor()
                    _cur.execute("""
                        SELECT c.condiciones_pago, c.vigencia_cotizacion
                        FROM cotizaciones c
                        JOIN clientes cl ON cl.id = c.cliente_id
                        WHERE cl.nombre LIKE ? AND c.condiciones_pago IS NOT NULL
                        ORDER BY c.id DESC LIMIT 1
                    """, (f"%{cliente_nombre_gen}%",))
                    _row = _cur.fetchone()
                    if _row and _row[0]:
                        _cond_default = _row[0].strip().upper()
                        _vigencia_default = _row[1] or "30 días"
                    _c.close()
                except Exception:
                    pass

                # ── Calcular Tiempo de Entrega desde semanas acumuladas de MO ──
                _semanas_total = 0.0
                _semanas_por_partida = {}
                for _mo in parsed_data.get("mo", []):
                    _p = _mo.get("partida_num", 0) or 0
                    _s = (_mo.get("horas", 8) * _mo.get("dias", 1)) / 48.0
                    _semanas_por_partida[_p] = max(_semanas_por_partida.get(_p, 0), _s)
                _semanas_total = sum(_semanas_por_partida.values())
                if _semanas_total <= 0:
                    _semanas_total = len(parsed_data.get("partidas", [])) * 2.0
                _sem_int = max(1, round(_semanas_total))
                _entrega_calculado = f"{_sem_int} SEMANA{'S' if _sem_int != 1 else ''}"

                # ──────────────────────────────────────────
                # FILA 1: MARGEN con botones de acceso rápido
                # ──────────────────────────────────────────
                st.markdown(f"<p style='font-size:13px;font-weight:800;color:{BRAND_CHARCOAL};margin:8px 0 4px 0;'>📈 Margen de Utilidad %</p>", unsafe_allow_html=True)

                if "margen_seleccionado" not in st.session_state:
                    st.session_state["margen_seleccionado"] = 35

                _opciones_margen = [20, 25, 30, 35, 40, 45, 50]
                _btn_cols = st.columns(len(_opciones_margen) + 1)
                for _i, _pct in enumerate(_opciones_margen):
                    _label = f"**{_pct}%**" if _pct == st.session_state["margen_seleccionado"] else f"{_pct}%"
                    _btn_type = "primary" if _pct == st.session_state["margen_seleccionado"] else "secondary"
                    if _btn_cols[_i].button(_label, key=f"btn_margen_{_pct}", type=_btn_type, use_container_width=True):
                        st.session_state["margen_seleccionado"] = _pct
                        st.rerun()

                margen_custom = _btn_cols[-1].number_input("Otro %", min_value=5, max_value=80, value=st.session_state["margen_seleccionado"], step=1, label_visibility="collapsed")
                if margen_custom != st.session_state["margen_seleccionado"]:
                    st.session_state["margen_seleccionado"] = margen_custom

                margen_sales = st.session_state["margen_seleccionado"] / 100.0

                st.markdown(f"""
                <div style="background:linear-gradient(90deg,#FE8C29 0%,#FF6B00 100%);border-radius:8px;padding:8px 16px;display:inline-block;margin-bottom:12px;">
                    <span style="color:white;font-weight:900;font-size:18px;">Margen Activo: {st.session_state['margen_seleccionado']}%</span>
                    &nbsp;&nbsp;<span style="color:#FFF3E0;font-size:12px;">→ Factor precio venta = 1 / (1 - {st.session_state['margen_seleccionado']/100:.2f})</span>
                </div>
                """, unsafe_allow_html=True)

                st.markdown("---")

                # ──────────────────────────────────────────
                # FILA 2: Comisión | Condiciones Pago | Tiempo Entrega | Vigencia
                # ──────────────────────────────────────────
                v_col1, v_col2, v_col3, v_col4 = st.columns([1.5, 2, 2, 1.5])

                with v_col1:
                    comision_pct_val = st.slider("Comisión Comercial %", min_value=0, max_value=20, value=5, step=1, format="%d%%")
                    comision_sales = comision_pct_val / 100.0
                    moneda_sales = st.selectbox("Moneda Cotización", ["MXN pesos mexicanos", "USD dólares americanos"], index=0)


                with v_col2:
                    # Condiciones de pago — precargadas del cliente, editables
                    _opciones_cond = ["CONTADO", "CRÉDITO 30 DÍAS", "CRÉDITO 60 DÍAS", "CRÉDITO 90 DÍAS", "50% ANTICIPO / 50% ENTREGA", "30% ANTICIPO / 70% CONTRA ENTREGA"]
                    _idx_cond = 0
                    for _i_c, _op in enumerate(_opciones_cond):
                        if _cond_default.upper() in _op or _op in _cond_default.upper():
                            _idx_cond = _i_c; break

                    pago_sales = st.selectbox(
                        "Condiciones de Pago",
                        _opciones_cond,
                        index=_idx_cond,
                        help=f"Precargado del historial del cliente: **{_cond_default}**. Puede cambiar para esta cotización."
                    )
                    pago_custom = st.text_input("O captura condición específica", placeholder="Ej: 70% Anticipo / 30% Entrega", label_visibility="visible")
                    if pago_custom.strip():
                        pago_sales = pago_custom.strip().upper()

                with v_col3:
                    # Tiempo de entrega — calculado automáticamente, editable
                    st.markdown(f"""
                    <div style="background:#E8F5E9;border:1px solid #4CAF50;border-radius:8px;padding:10px 12px;margin-bottom:8px;">
                        <div style="font-size:11px;color:#2E7D32;font-weight:700;">⏱ CALCULADO AUTOMÁTICAMENTE</div>
                        <div style="font-size:15px;font-weight:900;color:#1B5E20;">{_entrega_calculado}</div>
                        <div style="font-size:10px;color:#388E3C;">Suma acumulada de semanas MO: {_semanas_total:.1f} sem</div>
                        <div style="font-size:10px;color:#616161;">({len(_semanas_por_partida)} partida(s) en paralelo → máx. por partida)</div>
                    </div>
                    """, unsafe_allow_html=True)
                    _opciones_entrega = ["2 SEMANAS", "4 SEMANAS", "6 SEMANAS", "8 SEMANAS", "10 SEMANAS", "12 SEMANAS", "14 SEMANAS", "16 SEMANAS"]
                    _idx_entrega = 0
                    for _i_e, _op_e in enumerate(_opciones_entrega):
                        if _entrega_calculado.upper() in _op_e or _op_e in _entrega_calculado.upper():
                            _idx_entrega = _i_e
                            break

                    entrega_sales = st.selectbox(
                        "Ajustar Tiempo de Entrega",
                        _opciones_entrega,
                        index=_idx_entrega,
                        help="Calculado automáticamente como suma de semanas MO por partida. Selecciona las semanas necesarias para la entrega."
                    )

                with v_col4:
                    vigencia_sales = st.selectbox(
                        "Vigencia Cotización",
                        ["15 días", "30 días", "45 días", "60 días", "90 días"],
                        index=["15 días", "30 días", "45 días", "60 días", "90 días"].index(_vigencia_default) if _vigencia_default in ["15 días", "30 días", "45 días", "60 días", "90 días"] else 1
                    )

                # ──────────────────────────────────────────
                # SIMULACIÓN FINANCIERA EN TIEMPO REAL ($)
                # ──────────────────────────────────────────
                _cd_mat = sum(m.get("pu", 0) * m.get("cantidad", 1) for m in parsed_data.get("materiales", []))
                _cd_mo  = sum(mo.get("costo_hora", 0) * mo.get("horas", 8) * mo.get("dias", 1) * mo.get("personas", 1) for mo in parsed_data.get("mo", []))
                _cd_sup = _cd_mo * 0.30
                _cd_sub = sum(s.get("pu", 0) * s.get("cantidad", 1) for s in parsed_data.get("subcontratos", []))
                _cd_maq = sum(mq.get("pu", 0) * mq.get("tiempo", 1) for mq in parsed_data.get("maquinaria", []))
                _cd_hta = _cd_mo * 0.03
                _cd_gas = sum(g.get("costo_unitario", 0) * g.get("cantidad", 1) * g.get("tiempo", 1) for g in parsed_data.get("gastos", []))

                _cd_total = _cd_mat + _cd_mo + _cd_sup + _cd_sub + _cd_maq + _cd_hta + _cd_gas

                _pv_base = _cd_total / (1 - margen_sales) if margen_sales < 1 else _cd_total
                _monto_utilidad = _pv_base - _cd_total

                _pv_final = _pv_base / (1 - comision_sales) if comision_sales < 1 else _pv_base
                _monto_comision = _pv_final - _pv_base

                _tc_val = parsed_data.get("generales", {}).get("tc", 18.0) or 18.0
                _pv_usd = _pv_final / _tc_val if _tc_val > 0 else 0.0

                st.markdown(f"""
                <div style="background:#1E293B;border-radius:12px;padding:16px 20px;margin:16px 0;box-shadow:0 4px 12px rgba(0,0,0,0.15);font-family:'Montserrat',sans-serif;">
                    <div style="color:{BRAND_ORANGE};font-weight:900;font-size:13px;letter-spacing:0.5px;margin-bottom:12px;">
                        💰 SIMULACIÓN FINANCIERA EN TIEMPO REAL ($) — PROYECCIÓN COMERCIAL DE VENTAS
                    </div>
                    <div style="display:grid;grid-template-columns:repeat(5,1fr);gap:10px;text-align:center;">
                        <div style="background:#334155;border-radius:8px;padding:10px;">
                            <div style="color:#94A3B8;font-size:10px;font-weight:700;">COSTO DIRECTO</div>
                            <div style="color:white;font-size:15px;font-weight:900;">${_cd_total:,.2f}</div>
                            <div style="color:#CBD5E1;font-size:9px;">Base 100%</div>
                        </div>
                        <div style="background:#334155;border-radius:8px;padding:10px;">
                            <div style="color:#94A3B8;font-size:10px;font-weight:700;">UTILIDAD ({st.session_state['margen_seleccionado']}%)</div>
                            <div style="color:{BRAND_ORANGE};font-size:15px;font-weight:900;">${_monto_utilidad:,.2f}</div>
                            <div style="color:#CBD5E1;font-size:9px;">P.V. Base: ${_pv_base:,.2f}</div>
                        </div>
                        <div style="background:#334155;border-radius:8px;padding:10px;border:1px solid {BRAND_ORANGE};">
                            <div style="color:{BRAND_ORANGE};font-size:10px;font-weight:700;">COMISIÓN ({comision_pct_val}%)</div>
                            <div style="color:{BRAND_ORANGE};font-size:15px;font-weight:900;">${_monto_comision:,.2f}</div>
                            <div style="color:#CBD5E1;font-size:9px;">Monto Agente</div>
                        </div>
                        <div style="background:#0F172A;border-radius:8px;padding:10px;border:1.5px solid #10B981;">
                            <div style="color:#10B981;font-size:10px;font-weight:700;">PRECIO VENTA (MXN)</div>
                            <div style="color:#10B981;font-size:16px;font-weight:900;">${_pv_final:,.2f}</div>
                            <div style="color:#A7F3D0;font-size:9px;">IVA no incl.</div>
                        </div>
                        <div style="background:#0F172A;border-radius:8px;padding:10px;border:1.5px solid #3B82F6;">
                            <div style="color:#3B82F6;font-size:10px;font-weight:700;">PRECIO VENTA (USD)</div>
                            <div style="color:#60A5FA;font-size:15px;font-weight:900;">${_pv_usd:,.2f} USD</div>
                            <div style="color:#93C5FD;font-size:9px;">TC: ${_tc_val:,.2f}</div>
                        </div>
                    </div>
                </div>
                """, unsafe_allow_html=True)

                # ──────────────────────────────────────────
                # MATRIZ DE HITOS DE PAGO PARA ÓRDEN DE COMPRA (PO)
                # ──────────────────────────────────────────
                st.markdown("---")
                st.markdown(f"""
                <div style="background:#F8FAFC;border:1px solid #CBD5E1;border-left:4px solid {BRAND_ORANGE};border-radius:8px;padding:12px 16px;margin-bottom:14px;font-family:'Montserrat',sans-serif;">
                    <div style="font-size:13px;font-weight:800;color:{BRAND_CHARCOAL};">🎯 HITOS DE PAGO Y CRONOGRAMA PARA ÓRDEN DE COMPRA (PO)</div>
                    <div style="font-size:11px;color:{BRAND_CHARCOAL_MED};">Define los porcentajes %, descripciones y fechas estimadas para que el cliente genere su Purchase Order (PO) con claridad.</div>
                </div>
                """, unsafe_allow_html=True)

                # Extraer número total de semanas seleccionadas para la entrega
                import re
                _m_sem = re.search(r'\d+', entrega_sales)
                sem_tot = int(_m_sem.group()) if _m_sem else 8

                col_preset1, col_preset2 = st.columns([2.5, 2.5])
                with col_preset1:
                    _opciones_presets = [
                        "50% Anticipo / 50% Entrega (2 Hitos Estándar)",
                        "40% Anticipo / 40% Avance / 20% Entrega (3 Hitos Fabricación)",
                        "30% Anticipo / 40% Pruebas FAT / 30% SAT (3 Hitos Robótica)",
                        "30% PO / 30% Compras / 30% FAT / 10% SAT (4 Hitos Llave en Mano)",
                        "20% PO / 20% Diseño / 30% Fabricación / 20% FAT / 10% SAT (5 Hitos Integración Compleja)",
                        "60% Anticipo / 40% Contra Entrega (2 Hitos Equipos Importación)",
                        "70% Anticipo / 30% Entrega (2 Hitos Suministros Directos)",
                        "50% PO / 30% FAT J&D / 10% Embarque / 10% Puesta en Marcha (4 Hitos Celdas Ensamble)",
                        "100% Contado Al Colocar PO (1 Hito Pólizas & Servicios)",
                        "Personalizado Libre (5 Hitos en Blanco)"
                    ]
                    hitos_preset = st.selectbox(
                        "Plantilla Rápida de Hitos (10 Ejemplos Industriales)",
                        _opciones_presets,
                        index=0,
                        key="sel_hitos_preset"
                    )

                if "hitos_pago_state" not in st.session_state or st.session_state.get("last_preset") != hitos_preset or st.session_state.get("last_sem_tot") != sem_tot:
                    st.session_state["last_preset"] = hitos_preset
                    st.session_state["last_sem_tot"] = sem_tot

                    _sem_fat = max(1, round(sem_tot * 0.65))
                    _sem_compras = max(1, round(sem_tot * 0.35))
                    _sem_fat4 = max(1, round(sem_tot * 0.70))
                    _sem_diseno = max(1, round(sem_tot * 0.25))
                    _sem_fab = max(1, round(sem_tot * 0.50))
                    _sem_fat5 = max(1, round(sem_tot * 0.75))
                    _sem_emb = max(1, round(sem_tot * 0.85))

                    if "50% Anticipo / 50% Entrega" in hitos_preset:
                        st.session_state["hitos_pago_state"] = [
                            {"pct": 50, "desc": "Anticipo a la emisión de la Órden de Compra (PO)", "fecha": "Día 1 / Al colocar PO"},
                            {"pct": 50, "desc": "Finiquito contra entrega y recepción de equipos", "fecha": f"Semana {sem_tot} / Entrega"},
                            {"pct": 0, "desc": "", "fecha": ""},
                            {"pct": 0, "desc": "", "fecha": ""},
                            {"pct": 0, "desc": "", "fecha": ""}
                        ]
                    elif "40% Anticipo / 40% Avance / 20% Entrega" in hitos_preset:
                        st.session_state["hitos_pago_state"] = [
                            {"pct": 40, "desc": "Anticipo a la firma de la Órden de Compra (PO)", "fecha": "Día 1 / Colocación PO"},
                            {"pct": 40, "desc": "Pago por avance de fabricación y embarque", "fecha": f"Semana {max(1, round(sem_tot * 0.5))} / Embarque"},
                            {"pct": 20, "desc": "Finiquito tras entrega y recepción en planta", "fecha": f"Semana {sem_tot} / Entrega"},
                            {"pct": 0, "desc": "", "fecha": ""},
                            {"pct": 0, "desc": "", "fecha": ""}
                        ]
                    elif "30% Anticipo / 40% Pruebas FAT / 30% SAT" in hitos_preset:
                        st.session_state["hitos_pago_state"] = [
                            {"pct": 30, "desc": "Anticipo inicial para compra de componentes", "fecha": "Día 1 / Firma PO"},
                            {"pct": 40, "desc": "Pruebas de funcionamiento FAT en planta J&D", "fecha": f"Semana {_sem_fat} / FAT"},
                            {"pct": 30, "desc": "Aprobación final SAT y entrega en planta cliente", "fecha": f"Semana {sem_tot} / SAT"},
                            {"pct": 0, "desc": "", "fecha": ""},
                            {"pct": 0, "desc": "", "fecha": ""}
                        ]
                    elif "30% PO / 30% Compras / 30% FAT / 10% SAT" in hitos_preset:
                        st.session_state["hitos_pago_state"] = [
                            {"pct": 30, "desc": "Anticipo a la colocación de PO", "fecha": "Día 1 / Firma PO"},
                            {"pct": 30, "desc": "Arribo de componentes principales a taller J&D", "fecha": f"Semana {_sem_compras} / Compras"},
                            {"pct": 30, "desc": "Aceptación de Pruebas FAT en Taller J&D", "fecha": f"Semana {_sem_fat4} / FAT"},
                            {"pct": 10, "desc": "Aceptación Final SAT y Arranque en Planta Cliente", "fecha": f"Semana {sem_tot} / SAT"},
                            {"pct": 0, "desc": "", "fecha": ""}
                        ]
                    elif "20% PO / 20% Diseño / 30% Fabricación / 20% FAT / 10% SAT" in hitos_preset:
                        st.session_state["hitos_pago_state"] = [
                            {"pct": 20, "desc": "Anticipo inicial a la emisión de PO", "fecha": "Día 1 / Firma PO"},
                            {"pct": 20, "desc": "Liberación de Ingeniería de Diseño 3D en SolidWorks", "fecha": f"Semana {_sem_diseno} / Diseño"},
                            {"pct": 30, "desc": "Avance del 80% en Ensamble Mecánico y Tableros", "fecha": f"Semana {_sem_fab} / Fabricación"},
                            {"pct": 20, "desc": "Pruebas de Aceptación FAT en Taller J&D", "fecha": f"Semana {_sem_fat5} / FAT"},
                            {"pct": 10, "desc": "Aceptación Final SAT y Puesta en Marcha", "fecha": f"Semana {sem_tot} / SAT"}
                        ]
                    elif "60% Anticipo / 40% Contra Entrega" in hitos_preset:
                        st.session_state["hitos_pago_state"] = [
                            {"pct": 60, "desc": "Anticipo por compra e importación de componentes", "fecha": "Día 1 / Firma PO"},
                            {"pct": 40, "desc": "Finiquito al aviso de embarque y entrega", "fecha": f"Semana {sem_tot} / Entrega"},
                            {"pct": 0, "desc": "", "fecha": ""},
                            {"pct": 0, "desc": "", "fecha": ""},
                            {"pct": 0, "desc": "", "fecha": ""}
                        ]
                    elif "70% Anticipo / 30% Entrega" in hitos_preset:
                        st.session_state["hitos_pago_state"] = [
                            {"pct": 70, "desc": "Anticipo inicial para suministro de materiales", "fecha": "Día 1 / Firma PO"},
                            {"pct": 30, "desc": "Finiquito contra entrega de suministros", "fecha": f"Semana {sem_tot} / Entrega"},
                            {"pct": 0, "desc": "", "fecha": ""},
                            {"pct": 0, "desc": "", "fecha": ""},
                            {"pct": 0, "desc": "", "fecha": ""}
                        ]
                    elif "50% PO / 30% FAT J&D / 10% Embarque / 10% Puesta en Marcha" in hitos_preset:
                        st.session_state["hitos_pago_state"] = [
                            {"pct": 50, "desc": "Anticipo a la emisión de la Órden de Compra PO", "fecha": "Día 1 / Firma PO"},
                            {"pct": 30, "desc": "Aprobación de Pruebas FAT en Planta J&D", "fecha": f"Semana {_sem_fat} / FAT"},
                            {"pct": 10, "desc": "Salida de Embarque rumbo a planta del cliente", "fecha": f"Semana {_sem_emb} / Embarque"},
                            {"pct": 10, "desc": "Finiquito tras Puesta en Marcha exitosa", "fecha": f"Semana {sem_tot} / Puesta en Marcha"},
                            {"pct": 0, "desc": "", "fecha": ""}
                        ]
                    elif "100% Contado Al Colocar PO" in hitos_preset:
                        st.session_state["hitos_pago_state"] = [
                            {"pct": 100, "desc": "Pago de Contado al colocar la Órden de Compra (PO)", "fecha": "Día 1 / Colocación PO"},
                            {"pct": 0, "desc": "", "fecha": ""},
                            {"pct": 0, "desc": "", "fecha": ""},
                            {"pct": 0, "desc": "", "fecha": ""},
                            {"pct": 0, "desc": "", "fecha": ""}
                        ]
                    else:
                        st.session_state["hitos_pago_state"] = [
                            {"pct": 0, "desc": "", "fecha": ""},
                            {"pct": 0, "desc": "", "fecha": ""},
                            {"pct": 0, "desc": "", "fecha": ""},
                            {"pct": 0, "desc": "", "fecha": ""},
                            {"pct": 0, "desc": "", "fecha": ""}
                        ]

                    for _idx, _h_item in enumerate(st.session_state["hitos_pago_state"]):
                        st.session_state[f"hito_pct_{_idx}"] = int(_h_item["pct"])
                        st.session_state[f"hito_desc_{_idx}"] = _h_item["desc"]
                        st.session_state[f"hito_fecha_{_idx}"] = _h_item["fecha"]

                hitos_configurados = []
                hitos_pct_suma = 0

                st.markdown("<p style='font-size:12px;font-weight:800;color:#434E62;margin:10px 0 6px 0;'>CONFIGURACIÓN INDIVIDUAL DE HITOS (LOS 5 HITOS VISIBLES PARA CAPTURA LIBRE):</p>", unsafe_allow_html=True)

                for idx in range(5):
                    hc1, hc2, hc3, hc4 = st.columns([1, 2.5, 1.8, 1.5])
                    with hc1:
                        h_pct = st.number_input(f"% Hito {idx+1}", min_value=0, max_value=100, step=5, key=f"hito_pct_{idx}")
                    with hc2:
                        h_desc = st.text_input(f"Concepto / Descripción Hito {idx+1}", placeholder="Ej: Anticipo a la colocación de PO", key=f"hito_desc_{idx}")
                    with hc3:
                        h_fecha = st.text_input(f"Fecha Estimada / Plazo Hito {idx+1}", placeholder="Ej: Semana 3 / 15-Ago-2026", key=f"hito_fecha_{idx}")
                    with hc4:
                        h_monto = (_pv_final * (h_pct / 100.0))
                        st.markdown(f"<div style='margin-top:24px;font-size:12px;font-weight:900;color:{BRAND_ORANGE};'>${h_monto:,.2f} MXN</div>", unsafe_allow_html=True)

                    if h_pct > 0 or h_desc.strip():
                        hitos_configurados.append({
                            "hito_num": idx + 1,
                            "porcentaje": h_pct,
                            "descripcion": h_desc,
                            "fecha_estimada": h_fecha,
                            "monto_mxn": round(h_monto, 2)
                        })
                        hitos_pct_suma += h_pct


                if hitos_pct_suma == 100:
                    st.success(f"🟢 **Suma de Hitos de Pago: {hitos_pct_suma}% (Perfecto - 100% de la PO)**")
                elif hitos_pct_suma < 100:
                    st.warning(f"⚠️ **Suma de Hitos de Pago: {hitos_pct_suma}%** (Falta {100 - hitos_pct_suma}% para completar el 100% de la PO)")
                else:
                    st.error(f"❌ **Suma de Hitos de Pago: {hitos_pct_suma}%** (Supera por {hitos_pct_suma - 100}% el 100% de la PO)")

                import json
                hitos_pago_json_str = json.dumps(hitos_configurados, ensure_ascii=False)

                st.markdown("---")

                comercial_params = {
                    "margen_pct": margen_sales,
                    "comision_pct": comision_sales,
                    "moneda_cotizacion": moneda_sales,
                    "condiciones_pago": pago_sales,
                    "tiempo_entrega": entrega_sales,
                    "vigencia_cotizacion": vigencia_sales,
                    "hitos_pago_json": hitos_pago_json_str
                }


                # Resumen visual de la configuración antes de confirmar
                _factor_pv = 1 / (1 - margen_sales) if margen_sales < 1 else 1
                st.markdown(f"""
                <div style="background:{BRAND_CHARCOAL};border-radius:10px;padding:14px 20px;margin-bottom:16px;">
                    <div style="color:{BRAND_ORANGE};font-weight:900;font-size:13px;margin-bottom:8px;">📋 RESUMEN DE PARÁMETROS COMERCIALES SELECCIONADOS</div>
                    <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:8px;">
                        <div style="color:white;"><span style="color:#94A3B8;font-size:11px;">Margen Utilidad</span><br/><b style="font-size:15px;color:{BRAND_ORANGE};">{st.session_state['margen_seleccionado']}% (${_monto_utilidad:,.2f})</b></div>
                        <div style="color:white;"><span style="color:#94A3B8;font-size:11px;">Comisión Comercial</span><br/><b style="font-size:15px;color:{BRAND_ORANGE};">{comision_pct_val}% (${_monto_comision:,.2f})</b></div>
                        <div style="color:white;"><span style="color:#94A3B8;font-size:11px;">Precio Venta Final</span><br/><b style="font-size:15px;color:#10B981;">${_pv_final:,.2f} MXN</b></div>
                        <div style="color:white;"><span style="color:#94A3B8;font-size:11px;">Condiciones de Pago</span><br/><b style="font-size:13px;">{pago_sales}</b></div>
                        <div style="color:white;"><span style="color:#94A3B8;font-size:11px;">Tiempo de Entrega</span><br/><b style="font-size:13px;">{entrega_sales}</b></div>
                        <div style="color:white;"><span style="color:#94A3B8;font-size:11px;">Vigencia</span><br/><b style="font-size:13px;">{vigencia_sales}</b></div>
                    </div>
                </div>
                """, unsafe_allow_html=True)

                if st.button("🚀 CONFIRMAR Y CREAR COTIZACIÓN DE VENTAS EN EL SISTEMA", type="primary", use_container_width=True):
                    c_id, fol = guardar_cotizacion_auditada(parsed_data, comercial_params)
                    st.success(f"🎉 ¡Cotización **{fol}** procesada exitosamente con visión de ventas y guardada en la base de datos!")
                    st.info("Redireccionando al módulo de Cierre y Entrega...")
                    st.rerun()
