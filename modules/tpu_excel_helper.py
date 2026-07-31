def _xl_pct(value, base):
    return (value / base) if base > 0 else 0.0


def generate_tpu_dashboard_excel(cot_info, partidas):
    """
    Genera un Excel oficial multi-hoja con:
      - Hoja 1: Dashboard Resumen horizontal (formato matricial con colores J&D)
      - Hoja 2..N: Detalle de calculo TPU por cada Partida (1 hoja por partida)
      - Ultima hoja: Resumen General vertical (como prototipo del usuario)
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    C_ORANGE  = "FE8C29"
    C_NAVY    = "1E293B"
    C_AMBER_H = "FDE68A"
    C_AMBER_L = "FFF7ED"
    C_GREEN_D = "065F46"
    C_GREEN_L = "D1FAE5"
    C_BLUE_H  = "1E40AF"
    C_BLUE_L  = "DBEAFE"
    C_GRAY_L  = "F8FAFC"
    C_WHITE   = "FFFFFF"
    C_BLACK   = "0F172A"
    FMT_MXN   = '"$"#,##0.00'
    FMT_PCT   = '0.0%'

    def fill(hex_color):
        return PatternFill(fill_type='solid', fgColor=hex_color)

    def font(bold=False, color=C_BLACK, size=10):
        return Font(bold=bold, color=color, size=size, name='Calibri')

    def align(h='center', v='center', wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    thin_side   = Side(style='thin',   color='BFBFBF')
    thick_side  = Side(style='medium', color=C_NAVY)
    thin_border  = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    thick_border = Border(left=thick_side, right=thick_side, top=thick_side, bottom=thick_side)

    folio   = cot_info.get('folio', '---')
    proy    = cot_info.get('proyecto', '---')
    cliente = cot_info.get('cliente', '---')
    fecha   = datetime.now().strftime('%Y-%m-%d')

    wb = openpyxl.Workbook()

    # ==========================================================================
    # HOJA 1: DASHBOARD RESUMEN HORIZONTAL
    # ==========================================================================
    ws_dash = wb.active
    ws_dash.title = "Dashboard TPU"
    ws_dash.sheet_view.showGridLines = False
    ws_dash.freeze_panes = "B6"

    # Titulo principal
    ws_dash.merge_cells('A1:S1')
    ws_dash['A1'].value = f"RESUMEN DASHBOARD DE PRECIOS UNITARIOS (TPU) - {folio}"
    ws_dash['A1'].font = Font(bold=True, color=C_WHITE, size=13, name='Calibri')
    ws_dash['A1'].fill = fill(C_NAVY)
    ws_dash['A1'].alignment = align('left', 'center')
    ws_dash.row_dimensions[1].height = 22

    ws_dash.merge_cells('A2:E2')
    ws_dash['A2'].value = f"Proyecto: {proy}"
    ws_dash['A2'].font = font(bold=True, size=10)
    ws_dash['A2'].fill = fill(C_AMBER_L)
    ws_dash['A2'].alignment = align('left')
    ws_dash.row_dimensions[2].height = 16

    ws_dash.merge_cells('F2:J2')
    ws_dash['F2'].value = f"Cliente: {cliente}"
    ws_dash['F2'].font = font(size=10)
    ws_dash['F2'].fill = fill(C_AMBER_L)
    ws_dash['F2'].alignment = align('left')

    ws_dash.merge_cells('K2:S2')
    ws_dash['K2'].value = f"Fecha: {fecha}"
    ws_dash['K2'].font = font(size=9)
    ws_dash['K2'].fill = fill(C_AMBER_L)
    ws_dash['K2'].alignment = align('right')

    ws_dash.row_dimensions[3].height = 4

    # Sub-cabeceras con grupos
    group_headers = [
        ('A4', None,  'Partida'),
        ('B4', 'C4',  'Materiales Directos ($)'),
        ('D4', 'E4',  'Maquinaria y Equipo ($)'),
        ('F4', 'G4',  'MO + Factores ($)'),
        ('H4', 'I4',  'COSTO UNITARIO BASE ($)'),
        ('J4', 'K4',  'Indirecto de Campo ($)'),
        ('L4', 'M4',  'Indirecto Central ($)'),
        ('N4', 'O4',  'Utilidad ($)'),
        ('P4', 'Q4',  'PRECIO UNITARIO FINAL ($)'),
        ('R4', None,  'Total con IVA ($)'),
    ]
    for start, end, label in group_headers:
        if end:
            ws_dash.merge_cells(f'{start}:{end}')
        c = ws_dash[start]
        c.value = label
        c.font = Font(bold=True, color=C_WHITE, size=9, name='Calibri')
        c.fill = fill(C_NAVY)
        c.alignment = align('center', 'center', wrap=True)
        c.border = thin_border
    ws_dash.row_dimensions[4].height = 28

    sub_labels = ['','$ Importe','% Mat','$ Importe','% Maq','$ Importe','% MO',
                  '$ Importe','% Base','$ Importe','% Campo','$ Importe','% Central',
                  '$ Importe','% Util','$ Importe','% Final','$ c/IVA']
    for ci, lbl in enumerate(sub_labels, start=1):
        c = ws_dash.cell(row=5, column=ci, value=lbl)
        c.font = Font(bold=True, color=C_BLACK, size=8)
        c.fill = fill(C_AMBER_H)
        c.alignment = align('center')
        c.border = thin_border
    ws_dash.row_dimensions[5].height = 14

    col_widths_dash = [35,14,7,14,7,14,7,14,7,14,7,14,7,14,7,14,7,14]
    for ci, w in enumerate(col_widths_dash, start=1):
        from openpyxl.utils import get_column_letter
        ws_dash.column_dimensions[get_column_letter(ci)].width = w

    tot_mat=tot_maq=tot_mo=tot_base=tot_campo=tot_central=tot_util=tot_final=tot_iva=0.0
    all_tpu_data = []

    row_num = 6
    for idx, p in enumerate(partidas):
        p_id = p['id']
        conn = get_connection(); cur = conn.cursor()
        cur.execute("SELECT * FROM cotizacion_materiales_detalle WHERE partida_id=?", (p_id,))
        mats = [dict(r) for r in cur.fetchall()]
        cur.execute("SELECT * FROM cotizacion_mo_detalle WHERE partida_id=?", (p_id,))
        mo_rows = [dict(r) for r in cur.fetchall()]
        cur.execute("SELECT * FROM cotizacion_subcontratos_detalle WHERE partida_id=?", (p_id,))
        sub = [dict(r) for r in cur.fetchall()]
        cur.execute("SELECT * FROM cotizacion_maquinaria_detalle WHERE partida_id=?", (p_id,))
        maq = [dict(r) for r in cur.fetchall()]
        conn.close()

        tpu = calcular_tpu_partida(p, cot_info, mats, mo_rows, sub, maq)
        all_tpu_data.append((p, tpu, mats, mo_rows, maq))

        pu_f      = tpu['precio_unitario_final']
        m_mat     = tpu['costo_mat_unitario']
        m_maq     = tpu['costo_maq_unitario']
        m_base    = tpu['costo_unitario_base']
        m_campo   = tpu['monto_ind_campo']
        m_central = tpu['monto_ind_central']
        m_util    = tpu['monto_utilidad']
        m_mo      = m_base - m_mat - m_maq   # MO+Factores residual
        m_iva     = pu_f * 1.16

        tot_mat     += m_mat
        tot_maq     += m_maq
        tot_mo      += m_mo
        tot_base    += m_base
        tot_campo   += m_campo
        tot_central += m_central
        tot_util    += m_util
        tot_final   += pu_f
        tot_iva     += m_iva

        bg = C_GRAY_L if idx % 2 == 0 else C_WHITE
        p_name = f"Partida {p.get('numero_partida',1):04d}: {p.get('descripcion','')[:45]}"
        row_vals = [
            p_name,
            m_mat,   _xl_pct(m_mat,    pu_f),
            m_maq,   _xl_pct(m_maq,    pu_f),
            m_mo,    _xl_pct(m_mo,     pu_f),
            m_base,  _xl_pct(m_base,   pu_f),
            m_campo, _xl_pct(m_campo,  pu_f),
            m_central, _xl_pct(m_central, pu_f),
            m_util,  _xl_pct(m_util,   pu_f),
            pu_f,    1.0,
            m_iva
        ]
        for ci, val in enumerate(row_vals, start=1):
            c = ws_dash.cell(row=row_num, column=ci, value=val)
            c.fill = fill(bg)
            c.border = thin_border
            c.font = font(size=9)
            c.alignment = align('right') if ci > 1 else align('left')
            is_money = ci in (2,4,6,8,10,12,14,16,18)
            is_pct   = ci in (3,5,7,9,11,13,15,17)
            if is_money: c.number_format = FMT_MXN
            if is_pct:   c.number_format = FMT_PCT

        # COSTO BASE azul
        for ci in (8,9):
            c = ws_dash.cell(row=row_num, column=ci)
            c.fill = fill(C_BLUE_L)
            c.font = Font(bold=True, color=C_BLUE_H, size=9)
        # PRECIO FINAL verde
        for ci in (16,17):
            c = ws_dash.cell(row=row_num, column=ci)
            c.fill = fill(C_GREEN_L)
            c.font = Font(bold=True, color=C_GREEN_D, size=9)

        ws_dash.row_dimensions[row_num].height = 15
        row_num += 1

    # Fila TOTAL GENERAL
    tot_row = row_num
    tot_vals = [
        "TOTAL GENERAL",
        tot_mat,   _xl_pct(tot_mat,    tot_final),
        tot_maq,   _xl_pct(tot_maq,    tot_final),
        tot_mo,    _xl_pct(tot_mo,     tot_final),
        tot_base,  _xl_pct(tot_base,   tot_final),
        tot_campo, _xl_pct(tot_campo,  tot_final),
        tot_central, _xl_pct(tot_central, tot_final),
        tot_util,  _xl_pct(tot_util,   tot_final),
        tot_final, 1.0,
        tot_iva
    ]
    for ci, val in enumerate(tot_vals, start=1):
        c = ws_dash.cell(row=tot_row, column=ci, value=val)
        c.fill = fill(C_ORANGE)
        c.font = Font(bold=True, color=C_WHITE, size=10, name='Calibri')
        c.alignment = align('right') if ci > 1 else align('left')
        c.border = thick_border
        is_money = ci in (2,4,6,8,10,12,14,16,18)
        is_pct   = ci in (3,5,7,9,11,13,15,17)
        if is_money: c.number_format = FMT_MXN
        if is_pct:   c.number_format = FMT_PCT
    ws_dash.row_dimensions[tot_row].height = 20

    # Leyenda de formula
    ley_row = tot_row + 2
    ws_dash.merge_cells(f'A{ley_row}:R{ley_row}')
    ws_dash[f'A{ley_row}'].value = (
        "(*) MO + Factores = COSTO BASE - Materiales - Maquinaria  |  "
        "Mat% + Maq% + MO% + Ind.Campo% + Ind.Central% + Util% = 100%"
    )
    ws_dash[f'A{ley_row}'].font = Font(italic=True, color='64748B', size=8)

    # ==========================================================================
    # HOJAS 2..N: DETALLE DE CALCULO TPU POR PARTIDA
    # ==========================================================================
    for p, tpu, mats, mo_rows, maq_rows in all_tpu_data:
        sheet_name = f"P{p.get('numero_partida',1):04d}"[:31]
        ws_p = wb.create_sheet(title=sheet_name)
        ws_p.sheet_view.showGridLines = False
        for col_l, w in [('A',30),('B',10),('C',12),('D',14),('E',14),('F',10)]:
            ws_p.column_dimensions[col_l].width = w

        # Cabecera de cotizacion en cada hoja
        ws_p.merge_cells('A1:F1')
        ws_p['A1'].value = f"TARJETA DE PRECIO UNITARIO (TPU) -- {folio}"
        ws_p['A1'].font = Font(bold=True, color=C_WHITE, size=12)
        ws_p['A1'].fill = fill(C_NAVY)
        ws_p['A1'].alignment = align('left', 'center')
        ws_p.row_dimensions[1].height = 20

        ws_p.merge_cells('A2:C2')
        ws_p['A2'].value = f"Proyecto: {proy}"
        ws_p['A2'].font = font(bold=True, size=9)
        ws_p['A2'].fill = fill(C_AMBER_L)
        ws_p['A2'].alignment = align('left')

        ws_p.merge_cells('D2:F2')
        ws_p['D2'].value = f"Cliente: {cliente}  |  Rev: {cot_info.get('revision','R0')}  |  {fecha}"
        ws_p['D2'].font = font(size=9)
        ws_p['D2'].fill = fill(C_AMBER_L)
        ws_p['D2'].alignment = align('right')
        ws_p.row_dimensions[2].height = 15

        # Titulo de partida
        ws_p.merge_cells('A3:F3')
        ws_p['A3'].value = f"Partida {tpu['numero_partida']:04d}: {tpu['nombre_partida']}"
        ws_p['A3'].font = Font(bold=True, color=C_WHITE, size=11)
        ws_p['A3'].fill = fill(C_ORANGE)
        ws_p['A3'].alignment = align('left', 'center')
        ws_p.row_dimensions[3].height = 20

        ws_p.merge_cells('A4:F4')
        ws_p['A4'].value = f"Unidad: {tpu['unidad']}   |   H-H: {tpu['horas_hh_unitarias']:.4f} hrs   |   {tpu['descripcion'][:70]}"
        ws_p['A4'].font = Font(italic=True, size=8.5, color='475569')
        ws_p['A4'].alignment = align('left')
        ws_p.row_dimensions[4].height = 13

        r = 5

        def sh(labels, bg_hex):
            nonlocal r
            for ci, lbl in enumerate(labels, start=1):
                c = ws_p.cell(row=r, column=ci, value=lbl)
                c.font = Font(bold=True, color=C_WHITE, size=9)
                c.fill = fill(bg_hex)
                c.alignment = align('center')
                c.border = thin_border
            ws_p.row_dimensions[r].height = 16
            r += 1

        def dr(vals, bg=C_WHITE, fmts=None, bold=False, fg=C_BLACK):
            nonlocal r
            fmts = fmts or [''] * len(vals)
            for ci, (val, fmt) in enumerate(zip(vals, fmts), start=1):
                c = ws_p.cell(row=r, column=ci, value=val)
                c.font = Font(bold=bold, size=9, color=fg)
                c.fill = fill(bg)
                c.alignment = align('right') if isinstance(val, (int, float)) else align('left')
                c.border = thin_border
                if fmt:
                    c.number_format = fmt
            ws_p.row_dimensions[r].height = 14
            r += 1

        def subtotal_row(label, value, bg_hex, fg_hex=C_BLACK, pct=None):
            nonlocal r
            ws_p.merge_cells(f'A{r}:D{r}')
            ws_p[f'A{r}'].value = label
            ws_p[f'A{r}'].font = Font(bold=True, size=9, color=fg_hex)
            ws_p[f'A{r}'].fill = fill(bg_hex)
            ws_p[f'A{r}'].alignment = align('right')
            ws_p[f'A{r}'].border = thin_border
            ws_p[f'E{r}'].value = value
            ws_p[f'E{r}'].number_format = FMT_MXN
            ws_p[f'E{r}'].font = Font(bold=True, size=9, color=fg_hex)
            ws_p[f'E{r}'].fill = fill(bg_hex)
            ws_p[f'E{r}'].border = thin_border
            if pct is not None:
                ws_p[f'F{r}'].value = pct
                ws_p[f'F{r}'].number_format = FMT_PCT
                ws_p[f'F{r}'].font = Font(bold=True, size=9, color=fg_hex)
                ws_p[f'F{r}'].fill = fill(bg_hex)
                ws_p[f'F{r}'].border = thin_border
            ws_p.row_dimensions[r].height = 16
            r += 1

        # MATERIALES
        sh(['Material / Insumo','Unidad','Cantidad','Costo Unitario','Importe',''], C_NAVY)
        for m in tpu['mat_rows']:
            dr([m['material'], m['unidad'], m['cantidad'], m['costo'], m['importe'], ''],
               bg=C_GRAY_L, fmts=['','','0.000',FMT_MXN,FMT_MXN,''])
        subtotal_row("Total Materiales Directos", tpu['costo_mat_unitario'], C_AMBER_H)
        r += 1

        # MAQUINARIA (si aplica)
        if tpu['maq_rows']:
            sh(['Maquinaria / Equipo','Unidad','Cantidad','Costo Unitario','Importe',''], C_NAVY)
            for mq in tpu['maq_rows']:
                dr([mq['equipo'], mq['unidad'], mq['cantidad'], mq['costo'], mq['importe'], ''],
                   bg=C_GRAY_L, fmts=['','','0.000',FMT_MXN,FMT_MXN,''])
            subtotal_row("Total Maquinaria y Equipo", tpu['costo_maq_unitario'], C_AMBER_H)
            r += 1

        # MANO DE OBRA
        sh(['Categoria / Puesto','Cantidad','Horas H-H','Costo H-H','Importe',''], C_NAVY)
        for o in tpu['mo_rows']:
            dr([o['puesto'], o['cantidad'], o['horas'], o['costo_hh'], o['importe'], ''],
               bg=C_GRAY_L, fmts=['','0','0.000',FMT_MXN,FMT_MXN,''])
        subtotal_row("Total Mano de Obra (Base)", tpu['costo_mo_unitario'], C_AMBER_H)

        # Factores
        for lbl, monto in [
            (f"  + Herramienta ({tpu['hta_pct']:.2f}% sobre MO)", tpu['monto_herramienta']),
            (f"  + Supervision  ({tpu['sup_pct']:.2f}% sobre MO)",  tpu['monto_supervision']),
        ]:
            ws_p.merge_cells(f'A{r}:D{r}')
            ws_p[f'A{r}'].value = lbl
            ws_p[f'A{r}'].font = Font(italic=True, size=8.5, color='475569')
            ws_p[f'A{r}'].fill = fill(C_GRAY_L)
            ws_p[f'A{r}'].alignment = align('left')
            ws_p[f'A{r}'].border = thin_border
            ws_p[f'E{r}'].value = monto
            ws_p[f'E{r}'].number_format = FMT_MXN
            ws_p[f'E{r}'].font = Font(italic=True, size=8.5, color='475569')
            ws_p[f'E{r}'].fill = fill(C_GRAY_L)
            ws_p[f'E{r}'].border = thin_border
            ws_p.row_dimensions[r].height = 13
            r += 1

        subtotal_row("Total MO + Herramienta + Supervision",
                     tpu['precio_unitario_mo_factor'], C_BLUE_L, fg_hex=C_BLUE_H)
        r += 1

        # COSTO BASE
        subtotal_row("COSTO UNITARIO BASE",
                     tpu['costo_unitario_base'], C_BLUE_L, fg_hex=C_BLUE_H,
                     pct=_xl_pct(tpu['costo_unitario_base'], tpu['precio_unitario_final']))
        r += 1

        # INDIRECTOS
        for lbl, monto, pct_d in [
            (f"Indirecto de Campo ({tpu['ind_campo_pct']:.2f}%)",
             tpu['monto_ind_campo'],   _xl_pct(tpu['monto_ind_campo'],   tpu['precio_unitario_final'])),
            (f"Indirecto Central ({tpu['ind_central_pct']:.2f}%)",
             tpu['monto_ind_central'], _xl_pct(tpu['monto_ind_central'], tpu['precio_unitario_final'])),
            (f"Utilidad ({tpu['utilidad_pct']:.2f}%)",
             tpu['monto_utilidad'],    _xl_pct(tpu['monto_utilidad'],    tpu['precio_unitario_final'])),
        ]:
            subtotal_row(lbl, monto, C_GRAY_L, pct=pct_d)
        r += 1

        # PRECIO FINAL
        subtotal_row("PRECIO UNITARIO FINAL", tpu['precio_unitario_final'],
                     C_GREEN_D, fg_hex=C_WHITE, pct=1.0)
        subtotal_row("Total con IVA (16%)", tpu['precio_unitario_final'] * 1.16,
                     C_GREEN_L, fg_hex=C_GREEN_D)
        r += 1

        ws_p.merge_cells(f'A{r}:F{r}')
        ws_p[f'A{r}'].value = f"Son: {tpu['monto_letras']}"
        ws_p[f'A{r}'].font = Font(italic=True, size=8, color='64748B')
        ws_p[f'A{r}'].alignment = align('right')

    # ==========================================================================
    # ULTIMA HOJA: RESUMEN GENERAL VERTICAL
    # ==========================================================================
    ws_sum = wb.create_sheet(title="Resumen General")
    ws_sum.sheet_view.showGridLines = False
    for col_l, w in [('A',28),('B',4),('C',16),('D',10),('E',16),('F',10)]:
        ws_sum.column_dimensions[col_l].width = w

    ws_sum.merge_cells('A1:F1')
    ws_sum['A1'].value = f"RESUMEN DASHBOARD DE PRECIOS UNITARIOS (TPU) - {folio}"
    ws_sum['A1'].font = Font(bold=True, color=C_WHITE, size=12)
    ws_sum['A1'].fill = fill(C_NAVY)
    ws_sum['A1'].alignment = align('left', 'center')
    ws_sum.row_dimensions[1].height = 22

    ws_sum.merge_cells('A2:C2')
    ws_sum['A2'].value = f"Proyecto: {proy}"
    ws_sum['A2'].font = font(bold=True, size=9)
    ws_sum['A2'].fill = fill(C_AMBER_L)
    ws_sum['A2'].alignment = align('left')

    ws_sum['D2'].value = "Cliente:"
    ws_sum['D2'].font = font(size=9)
    ws_sum['D2'].fill = fill(C_AMBER_L)

    ws_sum.merge_cells('E2:F2')
    ws_sum['E2'].value = f"{cliente}  |  {fecha}"
    ws_sum['E2'].font = font(bold=True, size=9)
    ws_sum['E2'].fill = fill(C_AMBER_L)
    ws_sum['E2'].alignment = align('right')
    ws_sum.row_dimensions[2].height = 16

    def vrow(ws, row, label, value, pct, bg=C_WHITE, bold=False, fg=C_BLACK, show_pct=True):
        ws.merge_cells(f'A{row}:B{row}')
        c_l = ws[f'A{row}']
        c_l.value = label
        c_l.font = Font(bold=bold, size=10, color=fg, name='Calibri')
        c_l.fill = fill(bg)
        c_l.alignment = Alignment(horizontal='right', vertical='center')
        c_l.border = thin_border

        c_v = ws[f'C{row}']
        c_v.value = value
        c_v.number_format = FMT_MXN
        c_v.font = Font(bold=bold, size=10, color=fg)
        c_v.fill = fill(bg)
        c_v.alignment = Alignment(horizontal='right', vertical='center')
        c_v.border = thin_border

        if show_pct and pct is not None:
            c_p = ws[f'D{row}']
            c_p.value = pct
            c_p.number_format = FMT_PCT
            c_p.font = Font(bold=bold, size=10, color=fg)
            c_p.fill = fill(bg)
            c_p.alignment = Alignment(horizontal='right', vertical='center')
            c_p.border = thin_border

        ws.row_dimensions[row].height = 18

    rs = 4
    vrow(ws_sum, rs,   "Materiales Directos ($)",  tot_mat,     _xl_pct(tot_mat,    tot_final), C_GRAY_L)
    vrow(ws_sum, rs+1, "Maquinaria y Equipo ($)",  tot_maq,     _xl_pct(tot_maq,    tot_final), C_GRAY_L)
    vrow(ws_sum, rs+2, "Mano de Obra ($)",         tot_mo,      _xl_pct(tot_mo,     tot_final), C_GRAY_L)
    ws_sum.row_dimensions[rs+3].height = 5

    # COSTO BASE en columna E/F como en prototipo
    ws_sum.merge_cells(f'A{rs+4}:B{rs+4}')
    ws_sum[f'A{rs+4}'].value = "COSTO UNITARIO BASE ($)"
    ws_sum[f'A{rs+4}'].font = Font(bold=True, color=C_BLUE_H, size=11)
    ws_sum[f'A{rs+4}'].fill = fill(C_BLUE_L)
    ws_sum[f'A{rs+4}'].alignment = Alignment(horizontal='right', vertical='center')
    ws_sum[f'A{rs+4}'].border = thick_border
    ws_sum[f'E{rs+4}'].value = tot_base
    ws_sum[f'E{rs+4}'].number_format = FMT_MXN
    ws_sum[f'E{rs+4}'].font = Font(bold=True, color=C_BLUE_H, size=11)
    ws_sum[f'E{rs+4}'].fill = fill(C_BLUE_L)
    ws_sum[f'E{rs+4}'].border = thick_border
    ws_sum[f'F{rs+4}'].value = _xl_pct(tot_base, tot_final)
    ws_sum[f'F{rs+4}'].number_format = FMT_PCT
    ws_sum[f'F{rs+4}'].font = Font(bold=True, color=C_BLUE_H, size=11)
    ws_sum[f'F{rs+4}'].fill = fill(C_BLUE_L)
    ws_sum[f'F{rs+4}'].border = thick_border
    ws_sum.row_dimensions[rs+4].height = 22

    ws_sum.row_dimensions[rs+5].height = 5
    vrow(ws_sum, rs+6, "Indirecto de Campo ($)",  tot_campo,   _xl_pct(tot_campo,   tot_final), C_GRAY_L)
    vrow(ws_sum, rs+7, "Indirecto Central ($)",   tot_central, _xl_pct(tot_central, tot_final), C_GRAY_L)
    ws_sum.row_dimensions[rs+8].height = 5
    vrow(ws_sum, rs+9, "Utilidad ($)",            tot_util,    _xl_pct(tot_util,    tot_final), C_GRAY_L)
    ws_sum.row_dimensions[rs+10].height = 5

    # PRECIO FINAL
    ws_sum.merge_cells(f'A{rs+11}:B{rs+11}')
    ws_sum[f'A{rs+11}'].value = "PRECIO UNITARIO FINAL ($)"
    ws_sum[f'A{rs+11}'].font = Font(bold=True, color=C_WHITE, size=12)
    ws_sum[f'A{rs+11}'].fill = fill(C_GREEN_D)
    ws_sum[f'A{rs+11}'].alignment = Alignment(horizontal='right', vertical='center')
    ws_sum[f'A{rs+11}'].border = thick_border
    ws_sum[f'C{rs+11}'].value = tot_final
    ws_sum[f'C{rs+11}'].number_format = FMT_MXN
    ws_sum[f'C{rs+11}'].font = Font(bold=True, color=C_WHITE, size=12)
    ws_sum[f'C{rs+11}'].fill = fill(C_GREEN_D)
    ws_sum[f'C{rs+11}'].border = thick_border
    ws_sum.row_dimensions[rs+11].height = 24

    vrow(ws_sum, rs+12, "Total con IVA ($)", tot_iva, None, C_GREEN_L, bold=True, fg=C_GREEN_D, show_pct=False)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
