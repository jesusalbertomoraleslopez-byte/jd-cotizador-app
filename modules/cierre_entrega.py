import streamlit as st
import io
import os
import zipfile
from datetime import datetime, date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import pandas as pd

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas

from config import BRAND_ORANGE, BRAND_CHARCOAL, BRAND_CHARCOAL_MED, BRAND_GRAY_BG, BRAND_WHITE, BRAND_BORDER_LIGHT, get_brand_asset_path
from database.models import get_connection, init_db
from database.db_manager import get_cotizacion_detalles

# ─────────────────────────────────────────────────────────────────────────────
# 1. CANVAS PERSONALIZADO CON HOJA MEMBRETADA GENERAL (JD_HOJAMEMBRETADA_GENERAL)
# ─────────────────────────────────────────────────────────────────────────────

from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

def _get_jd_fonts():
    bold_font = "Helvetica-Bold"
    regular_font = "Helvetica"
    try:
        brand_dir = r"C:\Users\albertol\JD_Automation_Brand_Assets\Imagen Corporativa J&D\JD_ENTREGABLES\JD_TIPOGRAFIAS\nexa"
        heavy_path = os.path.join(brand_dir, "Nexa-Heavy.ttf")
        light_path = os.path.join(brand_dir, "Nexa-ExtraLight.ttf")

        if os.path.exists(heavy_path):
            if "Nexa-Heavy" not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont('Nexa-Heavy', heavy_path))
            bold_font = "Nexa-Heavy"

        if os.path.exists(light_path):
            if "Nexa-Light" not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont('Nexa-Light', light_path))
            regular_font = "Nexa-Light"
    except Exception:
        pass
    return bold_font, regular_font


# ─────────────────────────────────────────────────────────────────────────────
# 2. GENERADOR DE PDF CORPORATIVO (CON FONDO EXCLUSIVO Y NÚMERO DE PÁGINA)
# ─────────────────────────────────────────────────────────────────────────────

def _generate_cotizacion_pdf_oficial(cot_info, partidas, respuestas_tecnicas=None):
    """
    Genera el PDF oficial corporativo utilizando la hoja membretada oficial J&D
    con marca de agua del logo en diagonal en la portada y pie de página en todas las hojas.
    """
    buffer = io.BytesIO()

    # Título/Nombre de la cotización para el pie de página
    folio = cot_info.get('folio', '')
    concepto = cot_info.get('concepto') or cot_info.get('nombre_proyecto') or ''
    cot_title_footer = f"Cotización: {folio}" if folio else "Cotización J&D Automation"
    if concepto:
        cot_title_footer += f" | {concepto}"
    if len(cot_title_footer) > 65:
        cot_title_footer = cot_title_footer[:62] + "..."

    class JDFooterCanvas(canvas.Canvas):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self.pages = []
            
            # Dibujar fondo para la Página 1 inmediatamente en el constructor
            bg_path = get_brand_asset_path("hoja_membretada.png")
            if os.path.exists(bg_path):
                try:
                    self.saveState()
                    self.drawImage(bg_path, 0, 0, width=612, height=792)
                    self.restoreState()
                except Exception:
                    pass

        def _startPage(self):
            super()._startPage()
            # Fondo Membretado Oficial para páginas subsecuentes
            bg_path = get_brand_asset_path("hoja_membretada.png")
            if os.path.exists(bg_path):
                try:
                    self.saveState()
                    self.drawImage(bg_path, 0, 0, width=612, height=792)
                    self.restoreState()
                except Exception:
                    pass

        def showPage(self):
            self.pages.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            num_pages = len(self.pages)
            for page in self.pages:
                self.__dict__.update(page)
                self.draw_page_decorations(num_pages)
                super().showPage()
            super().save()

        def draw_page_decorations(self, total_pages):
            bold_font, regular_font = _get_jd_fonts()
            self.saveState()

            # 1. Nombre de la Cotización en el pie de página (cenefa naranja inferior izquierda)
            self.setFont(bold_font, 8)
            self.setFillColor(colors.white)
            self.drawString(36, 22, cot_title_footer)

            # 2. Número de Página en el pie de página (cenefa naranja inferior derecha)
            self.drawRightString(576, 22, f"Página {self._pageNumber} de {total_pages}")

            # 3. Folio/Acrónimo de la Cotización en el encabezado de todas las hojas (blanco alineado a la izquierda dentro de la cenefa naranja superior)
            self.setFont(bold_font, 8)
            self.setFillColor(colors.white)
            self.drawString(36, 742, f"FOLIO: {folio}")
            self.restoreState()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=36,
        rightMargin=36,
        topMargin=110,  # Espacio para respetar el encabezado de la hoja membretada
        bottomMargin=55   # Espacio para respetar la barra inferior de la hoja membretada
    )

    styles = getSampleStyleSheet()
    font_bold, font_regular = _get_jd_fonts()
    
    # ── PALETA DE COLORES INDUSTRIAL & TIPOGRAFÍA DE ALTO IMPACTO ──
    color_charcoal = colors.HexColor('#2C3442')      # Gris carbón industrial
    color_orange = colors.HexColor('#FE8C29')        # Naranja corporativo oficial J&D
    color_body = colors.HexColor('#2B2F38')          # Gris oscuro lectura confortable
    color_subtle = colors.HexColor('#5B6578')        # Gris secundario metadatos
    color_border = colors.HexColor('#CBD5E1')        # Gris para líneas de tabla
    color_bg_box = colors.HexColor('#F8FAFC')        # Plasta suave para cajas

    # Estilos Editoriales
    style_cover_title = ParagraphStyle('JDCoverTitle', parent=styles['Normal'], fontName=font_bold, fontSize=18, leading=22, textColor=color_charcoal)
    style_cover_sub = ParagraphStyle('JDCoverSub', parent=styles['Normal'], fontName=font_bold, fontSize=12, leading=15, textColor=color_orange)
    style_title = ParagraphStyle('JDFolio', parent=styles['Normal'], fontName=font_bold, fontSize=13, leading=15, textColor=color_charcoal)
    style_body = ParagraphStyle('JDBody', parent=styles['Normal'], fontName=font_regular, fontSize=8.5, leading=12.5, textColor=color_body)
    style_body_bold = ParagraphStyle('JDBodyBold', parent=styles['Normal'], fontName=font_bold, fontSize=8.5, leading=12.5, textColor=color_charcoal)
    style_heading = ParagraphStyle('JDHead', parent=styles['Normal'], fontName=font_bold, fontSize=11.5, leading=15, textColor=color_charcoal)
    style_anexo_title = ParagraphStyle('JDAnexoTitle', parent=styles['Normal'], fontName=font_bold, fontSize=10.5, leading=14, textColor=color_orange)
    style_anexo_body = ParagraphStyle('JDAnexoBody', parent=styles['Normal'], fontName=font_regular, fontSize=8, leading=12, textColor=color_body)
    
    # Estilos de Tabla
    style_table_header = ParagraphStyle('JDTableHeader', parent=styles['Normal'], fontName=font_bold, fontSize=8.5, leading=11, textColor=colors.white, alignment=1)
    style_table_header_l = ParagraphStyle('JDTableHeaderL', parent=styles['Normal'], fontName=font_bold, fontSize=8.5, leading=11, textColor=colors.white, alignment=0)
    style_table_header_r = ParagraphStyle('JDTableHeaderR', parent=styles['Normal'], fontName=font_bold, fontSize=8.5, leading=11, textColor=colors.white, alignment=2)
    style_table_cell = ParagraphStyle('JDTableCell', parent=styles['Normal'], fontName=font_regular, fontSize=8, leading=11, textColor=color_body, alignment=1)
    style_table_cell_l = ParagraphStyle('JDTableCellL', parent=styles['Normal'], fontName=font_regular, fontSize=8, leading=11, textColor=color_body, alignment=0)
    style_table_cell_r = ParagraphStyle('JDTableCellR', parent=styles['Normal'], fontName=font_regular, fontSize=8, leading=11, textColor=color_body, alignment=2)

    elements = []

    folio = cot_info.get('folio', 'YES-202607-090')
    fecha_str = cot_info.get('fecha_str') or datetime.now().strftime("%d %B %Y").upper()
    cliente_nombre = cot_info.get('cliente', 'YESERA MONTERREY')
    proyecto_nombre = cot_info.get('proyecto', 'INSTALACION ELECTRICA MOTORES')
    contacto_nombre = cot_info.get('nombre_contacto') or 'RICARDO GALLEGOS'

    # Términos Comerciales Dinámicos
    moneda_str = cot_info.get('moneda_cotizacion') or 'MXN pesos mexicanos'
    cond_pago_str = cot_info.get('condiciones_pago') or 'CREDITO'
    entrega_str = cot_info.get('tiempo_entrega') or '2 SEMANAS'
    vigencia_str = cot_info.get('vigencia_cotizacion') or '15 días'

    # ─────────────────────────────────────────────────────────────────────────
    # PÁGINA 1: PORTADA EJECUTIVA DE COTIZACIÓN Y METADATOS
    # ─────────────────────────────────────────────────────────────────────────
    # Espaciador para centrar verticalmente los metadatos debajo del encabezado de la hoja membretada
    elements.append(Spacer(1, 40))

    # Cuadro de Metadatos de la Cotización en la Portada (Réplica YES-202607-090)
    meta_box = [
        [Paragraph("Fecha:", style_body), Paragraph("Cotización:", style_body)],
        [Paragraph(f"<b>{fecha_str}</b>", style_body_bold), Paragraph(f"<b>{folio}</b>", style_body_bold)],
        [Paragraph(f"<b>{cliente_nombre.upper()}</b><br/>Planta Nazas<br/>Valle del Guadiana 37 esq. Piedras Negras<br/>Parque Industrial Lagunero", style_body),
         Paragraph("<b>J&D Automation, S.A. de C.V.</b><br/>Calle P # 352<br/>Eduardo Guerra<br/>Torreón, Coahuila, México<br/>Tel: (871) 1939690", style_body)]
    ]
    t_meta = Table(meta_box, colWidths=[266, 266])
    t_meta.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 0.5, color_border),
        ('INNERGRID', (0,0), (-1,-1), 0.5, color_border),
        ('BACKGROUND', (0,0), (-1,-1), color_bg_box),
        ('PADDING', (0,0), (-1,-1), 7),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    elements.append(t_meta)
    elements.append(Spacer(1, 14))

    # Párrafo institucional de Presentación J&D (Texto oficial de portada)
    intro_p = ("<b>J&D Automation</b> es una empresa especializada en proyectos llave en mano de robótica y automatización industrial "
               "para los sectores automotriz, metalúrgico y minero en México. Contamos con personal altamente calificado con más de 25 años "
               "de experiencia en la automatización de procesos, garantizando soluciones integrales que van desde el diseño y la fabricación "
               "hasta la implementación y el soporte técnico con componentes de vanguardia global.")

    elements.append(Paragraph(intro_p, style_body))
    elements.append(Spacer(1, 10))

    # Fortalezas de la Compañía (Con Ícono J&D y Tabulador Indentado)
    elements.append(Paragraph("<b>Fortalezas de la Compañía:</b>", style_body_bold))
    elements.append(Spacer(1, 4))
    
    bullet_icon_path = get_brand_asset_path("bullet_icon.png")
    fortalezas = [
        "Implementación de buenas prácticas CSIA (Control System Integrators Association) para la gestión de proyectos y administración del negocio.",
        "Capacitación constante por el PMI (Project Management Institute) para el manejo y administración de proyectos basándose en el PMBOK.",
        "Certificación como Recognized System Integrator por Rockwell Automation & ABB Robots.",
        "Integrador de sistemas por ABB Robótica.",
        "Licenciamiento para ingeniería de diseño mecánico en SolidWorks e Autodesk Inventor.",
        "Licenciamiento para ingeniería de diseño eléctrico en Eplan."
    ]
    for f in fortalezas:
        if os.path.exists(bullet_icon_path):
            try:
                img_bullet = Image(bullet_icon_path, width=9.5, height=6.5)
                t_bullet = Table([[img_bullet, Paragraph(f, style_body)]], colWidths=[18, 514])
                t_bullet.setStyle(TableStyle([
                    ('VALIGN', (0,0), (-1,-1), 'TOP'),
                    ('PADDING', (0,0), (-1,-1), 1),
                ]))
                elements.append(t_bullet)
            except Exception:
                elements.append(Paragraph(f"&bull; &nbsp;&nbsp;&nbsp;&nbsp;{f}", style_body))
        else:
            elements.append(Paragraph(f"&bull; &nbsp;&nbsp;&nbsp;&nbsp;{f}", style_body))
        elements.append(Spacer(1, 2.5))

    # Logotipos de Certificación y Tecnologías en el pie de la Portada (Nueva Tira Ultra Ancha 10.89:1)
    cert_logos_path = get_brand_asset_path("certificacion_logos.png")
    if os.path.exists(cert_logos_path):
        try:
            elements.append(Spacer(1, 10))
            img_cert = Image(cert_logos_path, width=460, height=42)
            img_cert.hAlign = 'CENTER'
            elements.append(img_cert)
        except Exception:
            pass

    # Salto obligatorio tras la Portada & Fortalezas en Página 1
    elements.append(PageBreak())

    # ── PÁGINA 2: COSTOS Y RESPUESTAS TÉCNICAS ──
    elements.append(Paragraph(f"<b>DIRIGIDO:</b> {contacto_nombre.upper()}", style_heading))
    elements.append(Spacer(1, 3))
    elements.append(Paragraph(f"<b>CONCEPTO:</b> {proyecto_nombre.upper()}", style_heading))
    elements.append(Spacer(1, 14))

    # ── SECCIÓN RESPUESTA TÉCNICA (FOTO Y ESPECIFICACIÓN DE EQUIPOS) ──
    if respuestas_tecnicas:
        elements.append(Paragraph("<b>RESPUESTA TÉCNICA Y ESPECIFICACIONES DE EQUIPOS:</b>", style_heading))
        elements.append(Spacer(1, 6))

        for rt in respuestas_tecnicas:
            img_w = None
            if rt.get('imagen_path') and os.path.exists(rt['imagen_path']):
                try:
                    img_w = Image(rt['imagen_path'], width=120, height=90)
                except Exception:
                    img_w = Paragraph("📷 <i>Imagen del equipo</i>", style_body)
            else:
                img_w = Paragraph("📸 <b>EQUIPO / COMPONENTE J&D</b>", style_body_bold)

            spec_text = (f"<b>Componente:</b> {rt.get('componente','TABLERO DE CONTROL')}<br/>"
                         f"<b>Partida N°:</b> {rt.get('partida_num', 1)}<br/>"
                         f"<b>Especificación Técnica:</b><br/>{rt.get('especificacion_tecnica','')}")

            t_rt = Table([[img_w, Paragraph(spec_text, style_body)]], colWidths=[130, 402])
            t_rt.setStyle(TableStyle([
                ('BOX', (0,0), (-1,-1), 1, color_orange),
                ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#FFF7ED')),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('PADDING', (0,0), (-1,-1), 6),
            ]))
            elements.append(t_rt)
            elements.append(Spacer(1, 10))

    elements.append(Paragraph("<b>COSTOS:</b>", style_heading))
    elements.append(Spacer(1, 8))

    # Tabla de Partidas y Costos (Réplica exact de YES-202607-090)
    cost_rows = [[
        Paragraph("<b>PARTIDA</b>", style_table_header),
        Paragraph("<b>DESCRIPCION</b>", style_table_header_l),
        Paragraph("<b>CANTIDAD</b>", style_table_header),
        Paragraph("<b>UNIDAD</b>", style_table_header),
        Paragraph("<b>PRECIO U.</b>", style_table_header_r),
        Paragraph("<b>SUBTOTAL</b>", style_table_header_r)
    ]]

    subtotal_acum = 0.0
    if not partidas:
        partidas = [{"numero_partida": 1, "descripcion": proyecto_nombre, "costo_directo_total": 107066.67}]

    for p in partidas:
        cd = float(p.get('costo_directo_total', 107066.67))
        pv = float(p.get('precio_venta') or (cd / 0.65))
        cant = p.get('cantidad') or 1
        unidad_str = p.get('unidad') or "SERV"
        sub_pv = pv * cant
        subtotal_acum += sub_pv

        cost_rows.append([
            Paragraph(str(p.get('numero_partida', 1)), style_table_cell),
            Paragraph(p.get('descripcion', proyecto_nombre).upper(), style_table_cell_l),
            Paragraph(str(cant), style_table_cell),
            Paragraph(str(unidad_str), style_table_cell),
            Paragraph(f"$ {pv:,.2f}", style_table_cell_r),
            Paragraph(f"$ {sub_pv:,.2f}", style_table_cell_r)
        ])

    iva_acum = subtotal_acum * 0.16
    total_general = subtotal_acum + iva_acum

    t_costos = Table(cost_rows, colWidths=[52, 230, 58, 50, 72, 70])
    t_costos.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), color_orange),
        ('GRID', (0,0), (-1,-1), 0.5, color_border),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING', (0,0), (-1,-1), 5),
    ]))
    elements.append(t_costos)
    elements.append(Spacer(1, 8))

    # Resumen Financiero Destacado (Plasta Gris Suave #F8FAFC, TOTAL en Negritas y Tamaño Superior en Naranja Óxido)
    totales_box = [
        [Paragraph("<b>SUBTOTAL:</b>", style_body_bold), Paragraph(f"<b>$ {subtotal_acum:,.2f}</b>", style_table_cell_r)],
        [Paragraph("<b>IVA 16%:</b>", style_body_bold), Paragraph(f"<b>$ {iva_acum:,.2f}</b>", style_table_cell_r)],
        [Paragraph("<font color='#D96B27' size=11><b>TOTAL:</b></font>", style_body_bold), Paragraph(f"<font color='#D96B27' size=11><b>$ {total_general:,.2f}</b></font>", style_table_cell_r)]
    ]
    t_tot = Table(totales_box, colWidths=[110, 110])
    t_tot.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 1, color_border),
        ('BACKGROUND', (0,0), (-1,-1), color_bg_box),
        ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
        ('PADDING', (0,0), (-1,-1), 6),
        ('LINEBELOW', (0,0), (-1,-2), 0.5, color_border),
    ]))

    t_wrapper = Table([["", t_tot]], colWidths=[312, 220])
    elements.append(t_wrapper)
    elements.append(Spacer(1, 16))

    # ── HITOS DE PAGO PARA ÓRDEN DE COMPRA (PO) ──
    hitos_json_raw = cot_info.get("hitos_pago_json", "")
    hitos_data = []
    if hitos_json_raw:
        try:
            import json
            hitos_data = json.loads(hitos_json_raw)
        except Exception:
            pass

    if hitos_data:
        elements.append(Paragraph("<b>ESQUEMA DE HITOS DE PAGO PARA ÓRDEN DE COMPRA (PO):</b>", style_heading))
        elements.append(Spacer(1, 6))
        
        t_hitos_rows = [[
            Paragraph("<b>Hito</b>", style_table_header),
            Paragraph("<b>% PO</b>", style_table_header),
            Paragraph("<b>Monto (MXN)</b>", style_table_header_r),
            Paragraph("<b>Fecha / Plazo Estimado</b>", style_table_header),
            Paragraph("<b>Concepto / Condición de Pago</b>", style_table_header),
        ]]
        for h in hitos_data:
            pct_val = h.get("porcentaje", 0)
            monto_val = total_general * (pct_val / 100.0)
            t_hitos_rows.append([
                Paragraph(f"Hito {h.get('hito_num', 1)}", style_table_cell),
                Paragraph(f"<b>{pct_val}%</b>", style_table_cell),
                Paragraph(f"${monto_val:,.2f}", style_table_cell_r),
                Paragraph(str(h.get("fecha_estimada", "-")), style_table_cell),
                Paragraph(str(h.get("descripcion", "-")), style_table_cell),
            ])

        t_hitos = Table(t_hitos_rows, colWidths=[45, 45, 95, 120, 227])
        t_hitos.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), color_charcoal),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('PADDING', (0,0), (-1,-1), 5.5),
            ('LINEBELOW', (0,0), (-1,-1), 0.5, color_border),
        ]))
        elements.append(t_hitos)
        elements.append(Spacer(1, 14))

    # Términos Comerciales Dinámicos
    elements.append(Paragraph("<b>TÉRMINOS COMERCIALES:</b>", style_heading))
    elements.append(Spacer(1, 6))
    tc_items = [
        f"• Los costos mencionados están indicados en {moneda_str}.",
        f"• Condiciones de pago:\n  o {cond_pago_str.upper()}",
        f"• Tiempo de entrega: {entrega_str.upper()}.",
        f"• Vigencia de la cotización: {vigencia_str}."
    ]
    for tc_i in tc_items:
        elements.append(Paragraph(tc_i.replace('\n','<br/>'), style_body))
        elements.append(Spacer(1, 2.5))

    elements.append(Spacer(1, 14))

    # ── CRONOGRAMA DE GANTT INTEGRADO EN EL PDF ──
    elements.append(PageBreak())
    try:
        from database.db_manager import get_connection
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT actividad, tipo, fecha_inicio, dias_duracion 
            FROM cotizacion_gantt 
            WHERE cotizacion_id=? 
            ORDER BY orden, id
        """, (cot_info["id"],))
        gantt_tasks = [dict(r) for r in cur.fetchall()]
        conn.close()
    except Exception:
        gantt_tasks = []

    if gantt_tasks:
        elements.append(Paragraph("<b>PLAN Y CRONOGRAMA DE PROYECTO (GANTT):</b>", style_heading))
        elements.append(Spacer(1, 4))
        
        # Calcular rango temporal
        start_dates = []
        end_dates = []
        for t in gantt_tasks:
            try:
                sd = datetime.strptime(str(t["fecha_inicio"]), "%Y-%m-%d").date()
                start_dates.append(sd)
                end_dates.append(sd + timedelta(days=int(t["dias_duracion"] or 1)))
            except Exception:
                pass
                
        if start_dates:
            min_start = min(start_dates)
            max_end = max(end_dates)
            total_days = (max_end - min_start).days
            if total_days <= 0:
                total_days = 30
        else:
            min_start = date.today()
            max_end = min_start + timedelta(days=30)
            total_days = 30
            
        col_width_days = max(1.0, total_days / 15.0)
        
        elements.append(Paragraph(f"<font color='#64748B'><i>Línea de tiempo del proyecto: {min_start.strftime('%d/%m/%Y')} al {max_end.strftime('%d/%m/%Y')}</i></font>", style_body))
        elements.append(Spacer(1, 6))

        # Crear cabecera de la tabla
        style_gantt_header = ParagraphStyle('JDGanttHeader', parent=styles['Normal'], fontName=font_bold, fontSize=8, leading=10, textColor=colors.white)
        style_gantt_header_c = ParagraphStyle('JDGanttHeaderC', parent=styles['Normal'], fontName=font_bold, fontSize=7, leading=9, textColor=colors.white, alignment=1)
        style_gantt_cell = ParagraphStyle('JDGanttCell', parent=styles['Normal'], fontName=font_regular, fontSize=8, leading=10, textColor=color_charcoal)
        
        t_gantt_rows = []
        header_row = [
            Paragraph("<b>Actividad / Hito</b>", style_gantt_header),
            Paragraph("<b>Inicio</b>", style_gantt_header),
            Paragraph("<b>Dur.</b>", style_gantt_header),
        ]
        
        for c in range(15):
            c_date = min_start + timedelta(days=int(c * col_width_days))
            header_row.append(Paragraph(c_date.strftime("%d"), style_gantt_header_c))
            
        t_gantt_rows.append(header_row)
        
        # Estilos de la tabla
        t_style = [
            ('BACKGROUND', (0,0), (-1,0), color_charcoal),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
            ('PADDING', (0,0), (-1,-1), 4),
        ]
        
        color_map = {
            "Actividad": colors.HexColor("#FE8C29"),   # Naranja
            "Entregable": colors.HexColor("#10B981"),  # Verde
            "Reunión": colors.HexColor("#3B82F6"),     # Azul
            "Hito": colors.HexColor("#EF4444")         # Rojo
        }
        
        for r_idx, t in enumerate(gantt_tasks, 1):
            t_row = [
                Paragraph(t["actividad"], style_gantt_cell),
                Paragraph(t["fecha_inicio"], style_gantt_cell),
                Paragraph(f"{t['dias_duracion']}d", style_gantt_cell),
            ]
            for c in range(15):
                t_row.append("")
            t_gantt_rows.append(t_row)
            
            # Calcular coloreado de celdas
            try:
                t_start = datetime.strptime(str(t["fecha_inicio"]), "%Y-%m-%d").date()
                t_end = t_start + timedelta(days=int(t["dias_duracion"] or 1))
            except Exception:
                t_start = min_start
                t_end = min_start + timedelta(days=1)
                
            for c in range(15):
                col_start = min_start + timedelta(days=c * col_width_days)
                col_end = min_start + timedelta(days=(c + 1) * col_width_days)
                
                # Si se cruzan temporalmente, pintar la celda
                if t_start < col_end and t_end > col_start:
                    task_color = color_map.get(t.get("tipo", "Actividad"), colors.HexColor("#FE8C29"))
                    t_style.append(('BACKGROUND', (3 + c, r_idx), (3 + c, r_idx), task_color))
                    
        # Anchos de columna
        col_widths = [180, 60, 40] + [17]*15
        t_gantt = Table(t_gantt_rows, colWidths=col_widths)
        t_gantt.setStyle(TableStyle(t_style))
        elements.append(t_gantt)
        elements.append(Spacer(1, 14))

    # ── BLOQUE DE FIRMA TOTALMENTE CARGADO A LA DERECHA ──
    style_firma_atentamente = ParagraphStyle('JDFirmaAtentamente', parent=styles['Normal'], fontName=font_bold, fontSize=9, leading=12, textColor=color_charcoal, alignment=2)
    style_firma_info = ParagraphStyle('JDFirmaInfo', parent=styles['Normal'], fontName=font_regular, fontSize=8, leading=11, textColor=color_body, alignment=2)
    style_firma_linea = ParagraphStyle('JDFirmaLinea', parent=styles['Normal'], fontName=font_regular, fontSize=8, leading=10, textColor=colors.HexColor('#94A3B8'), alignment=2)

    firma_elements = [
        Paragraph("<b>A t e n t a m e n t e</b>", style_firma_atentamente),
        Spacer(1, 4)
    ]

    firma_img_path = get_brand_asset_path("firma_david.png")
    if os.path.exists(firma_img_path):
        try:
            img_firma = Image(firma_img_path, width=160, height=38)
            img_firma.hAlign = 'RIGHT'
            firma_elements.append(img_firma)
            firma_elements.append(Spacer(1, 2))
        except Exception:
            firma_elements.append(Spacer(1, 15))
    else:
        firma_elements.append(Spacer(1, 15))

    firma_elements.append(Paragraph("___________________________________", style_firma_linea))
    firma_elements.append(Spacer(1, 3))
    firma_elements.append(Paragraph("<b>Ing. David Alaniz</b><br/><font color='#5B6578'>Gerente de Ventas &amp; Automatización</font><br/>Tel: 81 8176 8569 &bull; E-mail: ventas@jydautomation.mx", style_firma_info))

    t_firma_box = Table([[firma_elements]], colWidths=[240])
    t_firma_box.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('PADDING', (0,0), (-1,-1), 0),
    ]))

    t_firma_wrapper = Table([["", t_firma_box]], colWidths=[300, 240])
    t_firma_wrapper.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('PADDING', (0,0), (-1,-1), 0),
    ]))

    elements.append(t_firma_wrapper)

    # ── PÁGINA 3: ANEXOS Y CLÁUSULAS CONTRACTUALES COMPLETA EN 1 SOLA HOJA ──
    elements.append(PageBreak())
    
    style_anexo_head_single = ParagraphStyle('JDAnexoHeadSingle', parent=styles['Normal'], fontName=font_bold, fontSize=7.5, leading=9.5, textColor=color_charcoal)
    style_anexo_body_single = ParagraphStyle('JDAnexoBodySingle', parent=styles['Normal'], fontName=font_regular, fontSize=6.8, leading=8.8, textColor=color_body)
    
    elements.append(Paragraph("ANEXOS Y CONDICIONES CONTRACTUALES DE VENTA", style_anexo_title))
    elements.append(Spacer(1, 5))

    anexos_todos = [
        ("1. EMPAQUES Y TRANSPORTE",
         "Los productos embarcados por J&D AUTOMATION se protegen adecuadamente para asegurar su integridad. Precios cotizados en condición EXW (Torreón, Coahuila). Fletes foráneos, maniobras especiales y seguros adicionales corren por cuenta del cliente, salvo acuerdo previo por escrito."),

        ("2. GARANTÍA OFICIAL Y SERVICIO TÉCNICO",
         "Garantía de 1 año en ensambles y fabricaciones a partir de recepción en planta, y 3 meses en desarrollos de ingeniería (programación de PLC, Robots, HMI y Software). Se invalida por intervención de terceros sin visto bueno, mala aplicación, fluctuaciones eléctricas o uso fuera de especificación."),

        ("3. DEVOLUCIONES DE PRODUCTO Y LIBERACIÓN DE DIBUJOS",
         "Devoluciones únicamente aceptadas por discrepancias no conformes respecto al diseño liberado. Para todo servicio de ingeniería, es requisito indispensable la aprobación por escrito de planos y diagramas por el responsable asignado por el cliente."),

        ("4. SEGUROS, FIANZAS Y PROPIEDAD INTELECTUAL",
         "Fianzas y seguros específicos se cotizarán según RFQ. J&D AUTOMATION mantiene Seguro de Responsabilidad Civil General. Todos los desarrollos, diagramas y software integrados son propiedad intelectual exclusiva de J&D AUTOMATION, otorgándose licencia de uso exclusivo para operación interna."),

        ("5. ÉTICA PROFESIONAL Y CONFIDENCIALIDAD",
         "Información técnica y comercial tratada como confidencial. Queda estrictamente prohibida la contratación directa de personal que participe en la ejecución del proyecto entre ambas partes, durante el desarrollo y hasta 1 año posterior a la conclusión del contrato."),

        ("6. ERRORES, OMISIONES Y RESOLUCIÓN DE DISPUTAS",
         "Cualquier omisión involuntaria en la descripción de insumos será ajustada sin exceder los límites del alcance cotizado. Cualquier controversia será determinada bajo las leyes y jurisdicción de los tribunales del Estado de Coahuila."),

        ("7. RESTRICCIONES POR CAMBIOS, ACELERACIÓN Y CANCELACIÓN",
         "Tiempos extras, turnos nocturnos o paros de planta imputables al cliente se facturan por separado. Cambios de alcance requieren notificación escrita y emisión de Orden de Compra modificatoria. Cancelaciones cubren materiales comprados y mano de obra ejecutada, más 8% de administración."),

        ("8. FUERZA MAYOR Y EXCLUSIÓN DE PENALIZACIONES",
         "J&D AUTOMATION queda exento de responsabilidad por demoras derivadas de fuerza mayor, desastres naturales, huelgas, pandemias o regulaciones gubernamentales. Los plazos de entrega se extenderán en proporción directa a la demora sin penalización alguna.")
    ]

    for tit, txt in anexos_todos:
        elements.append(Paragraph(f"<b>{tit}</b>", style_anexo_head_single))
        elements.append(Spacer(1, 1))
        elements.append(Paragraph(txt, style_anexo_body_single))
        elements.append(Spacer(1, 3.5))

    # Construir PDF con el Canvas de Encabezado/Pie J&D
    doc.build(elements, canvasmaker=JDFooterCanvas)
    buffer.seek(0)
    return buffer.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# 3. GENERADOR DE PRESUPUESTO EN EXCEL (.XLSX) CON FÓRMULAS NATIVAS
# ─────────────────────────────────────────────────────────────────────────────

def _generate_presupuesto_excel(cot_info, partidas):
    """
    Genera un archivo Excel (.xlsx) estructurado financieramente con fórmulas nativas
    (=C*E, =SUMA, =SUBTOTAL*0.16, =TOTAL) editable por el equipo técnico.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PRESUPUESTO_FINANCIERO"
    ws.views.sheetView[0].showGridLines = True

    fill_head = PatternFill(start_color="434E62", end_color="434E62", fill_type="solid")
    font_head = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Segoe UI", size=10, bold=True)

    folio = cot_info.get('folio', 'YES-202607-089')
    cliente = cot_info.get('cliente', 'YESERA MONTERREY')
    proyecto = cot_info.get('proyecto', 'TABLERO DE CONTROL DE MOTORES')

    # Encabezado Corporativo
    ws.cell(1, 1, "J&D AUTOMATION INDUSTRIES S.A. DE C.V.").font = Font(name="Segoe UI", size=14, bold=True, color="434E62")
    ws.cell(2, 1, f"PRESUPUESTO FINANCIERO DE PROYECTO — FOLIO: {folio}").font = Font(name="Segoe UI", size=11, bold=True, color="FE8C29")
    ws.cell(4, 1, "Cliente:").font = font_bold; ws.cell(4, 2, cliente)
    ws.cell(5, 1, "Proyecto:").font = font_bold; ws.cell(5, 2, proyecto)
    ws.cell(6, 1, "Fecha:").font = font_bold; ws.cell(6, 2, datetime.now().strftime("%Y-%m-%d"))

    headers = ["PARTIDA", "DESCRIPCIÓN", "CANTIDAD", "UNIDAD", "PRECIO UNITARIO (MXN)", "SUBTOTAL (MXN)"]
    row_start = 9
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row_start, col_idx, h)
        cell.fill = fill_head; cell.font = font_head; cell.alignment = Alignment(horizontal='center')

    curr_row = row_start + 1
    if not partidas:
        partidas = [{"numero_partida": 1, "descripcion": proyecto, "costo_directo_total": 107066.67}]

    for p in partidas:
        cd = float(p.get('costo_directo_total', 107066.67))
        pv = float(p.get('precio_venta') or (cd / 0.65))

        ws.cell(curr_row, 1, p.get('numero_partida', 1)).alignment = Alignment(horizontal='center')
        ws.cell(curr_row, 2, p.get('descripcion', proyecto).upper()).font = font_bold
        ws.cell(curr_row, 3, 1).alignment = Alignment(horizontal='center')
        ws.cell(curr_row, 4, "SERV").alignment = Alignment(horizontal='center')
        ws.cell(curr_row, 5, pv).number_format = '$#,##0.00'

        cell_sub = ws.cell(curr_row, 6, f"=C{curr_row}*E{curr_row}")
        cell_sub.number_format = '$#,##0.00'
        cell_sub.font = font_bold
        curr_row += 1

    last_item_row = curr_row - 1

    curr_row += 1
    ws.cell(curr_row, 5, "SUBTOTAL:").font = font_bold
    cell_tot_sub = ws.cell(curr_row, 6, f"=SUM(F{row_start+1}:F{last_item_row})")
    cell_tot_sub.number_format = '$#,##0.00'
    cell_tot_sub.font = font_bold
    sub_row = curr_row

    curr_row += 1
    ws.cell(curr_row, 5, "IVA 16%:").font = font_bold
    cell_iva = ws.cell(curr_row, 6, f"=F{sub_row}*0.16")
    cell_iva.number_format = '$#,##0.00'
    cell_iva.font = font_bold
    iva_row = curr_row

    curr_row += 1
    ws.cell(curr_row, 5, "TOTAL PROYECTO:").font = Font(name="Segoe UI", size=11, bold=True, color="FE8C29")
    cell_total = ws.cell(curr_row, 6, f"=F{sub_row}+F{iva_row}")
    cell_total.number_format = '$#,##0.00'
    cell_total.font = Font(name="Segoe UI", size=12, bold=True, color="FE8C29")

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 24
    ws.column_dimensions['F'].width = 24

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# 4. GENERADOR DE CORREO NOTIFICACIÓN (.EML) MULTIPART/MIXED
# ─────────────────────────────────────────────────────────────────────────────

def _generate_correo_eml(cot_info, pdf_bytes, excel_bytes, partidas=None, custom_to=None, custom_cc=None):
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.application import MIMEApplication
    import base64
    import re
    import os
    
    msg = MIMEMultipart('mixed')
    msg['X-Unsent'] = '1'
    folio = cot_info.get('folio', 'COT-2026-001-JAI-DS')
    proyecto = cot_info.get('proyecto', 'AUTOMATIZACIÓN DE LÍNEA DE PROCESO')
    cliente = cot_info.get('cliente', 'YESERA MONTERREY')
    contacto = cot_info.get('nombre_contacto') or 'Ing. Ricardo Gallegos'
    fecha_rev = cot_info.get('fecha_str') or cot_info.get('fecha') or datetime.now().strftime("%d/%m/%Y")
    rev_str = cot_info.get('revision') or 'R0'
    
    if not (contacto.startswith('Ing.') or contacto.startswith('Lic.') or contacto.startswith('Arq.')):
        contacto_saludo = f"Ing. {contacto}"
    else:
        contacto_saludo = contacto

    clean_f = folio.replace('_Cotizacion_Oficial', '').strip()
    f_parts = clean_f.split('-')
    folio_corto = "-".join(f_parts[:3]) if len(f_parts) >= 3 else clean_f

    # Sanitizar el nombre del folio para nombres de archivo HTTP/MIME limpios sin caracteres especiales
    clean_folio_fname = re.sub(r'[^a-zA-Z0-9_-]', '_', folio_corto).strip('_')

    condiciones_pago = cot_info.get('condiciones_pago', '50% Anticipo | 30% Contra entrega de tableros | 20% Cierre de SAT')
    tiempo_entrega = cot_info.get('tiempo_entrega', '14 semanas')

    # Cálculo de Precio Total con IVA
    total_subtotal = 0.0
    if partidas:
        for p in partidas:
            cd = float(p.get('costo_directo_total', 0) or 0)
            pv = float(p.get('precio_venta') or (cd / 0.65 if cd > 0 else 0))
            total_subtotal += pv
    if total_subtotal <= 0:
        total_subtotal = 107066.67 / 0.65

    total_iva = total_subtotal * 0.16
    total_con_iva = total_subtotal + total_iva
    moneda_str = cot_info.get('moneda_cotizacion', 'MXN')
    moneda_code = 'USD' if 'USD' in str(moneda_str).upper() else 'MXN'
    precio_total_str = f"${total_con_iva:,.2f} {moneda_code} (IVA Incluido)"

    # Pre-cálculo y escalado del logotipo a Base64 manteniendo la proporción exacta al doble (280px)
    logo_w, logo_h = 280, 70
    logo_b64 = ""
    logo_path = get_brand_asset_path("logo_corporativo.png")
    if os.path.exists(logo_path):
        try:
            from PIL import Image
            im = Image.open(logo_path)
            im.thumbnail((280, 75), Image.Resampling.LANCZOS)
            logo_w, logo_h = im.size
            buf = io.BytesIO()
            im.save(buf, format="PNG")
            logo_b64 = base64.b64encode(buf.getvalue()).decode('utf-8')
        except Exception:
            pass

    msg['Subject'] = f"Propuesta Técnica y Comercial: {proyecto} | Ref: {folio_corto}"
    
    if custom_to and str(custom_to).strip():
        msg['To'] = str(custom_to).strip()
    else:
        client_email = cot_info.get('email_contacto') or cot_info.get('email_cliente') or cot_info.get('email') or cot_info.get('correo')
        if client_email and str(client_email).strip():
            msg['To'] = f"{contacto} <{str(client_email).strip()}>"
        else:
            msg['To'] = f"{contacto} <contacto.cliente@empresa.com>"

    if custom_cc and str(custom_cc).strip():
        msg['Cc'] = str(custom_cc).strip()
    else:
        msg['Cc'] = "alberto.morales@jydautomation.com.mx, david.alaniz@jydautomation.com.mx, ventas@jydautomation.com.mx"

    logo_img_tag = f'<div style="background:#FFFFFF; padding:8px 16px; border-radius:8px; display:inline-block; margin-bottom:12px;"><img src="data:image/png;base64,{logo_b64}" width="{logo_w}" height="{logo_h}" style="width:{logo_w}px; height:{logo_h}px; display:block; border:0;" alt="J&amp;D Automation"></div>' if logo_b64 else ''

    html_body = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <style>
        body {{ font-family: 'Segoe UI', Arial, sans-serif; color: #334155; margin: 0; padding: 0; background-color: #F8FAFC; }}
        .container {{ max-width: 680px; margin: 20px auto; background: #FFFFFF; border-radius: 12px; overflow: hidden; border: 1px solid #E2E8F0; box-shadow: 0 4px 12px rgba(0,0,0,0.05); }}
        .header {{ background-color: #1E293B; padding: 24px; text-align: left; border-bottom: 5px solid #FE8C29; }}
        .header h1 {{ color: #FFFFFF; margin: 0; font-size: 20px; font-weight: 800; letter-spacing: 1px; }}
        .header p {{ color: #FE8C29; margin: 4px 0 0 0; font-size: 12px; font-weight: 700; text-transform: uppercase; letter-spacing: 1.2px; }}
        .content {{ padding: 28px; line-height: 1.65; font-size: 13.5px; color: #334155; }}
        .section-title {{ font-weight: 800; font-size: 12px; text-transform: uppercase; color: #FE8C29; margin-top: 22px; margin-bottom: 8px; letter-spacing: 0.8px; }}
        .summary-box {{ background-color: #F8FAFC; border-left: 4px solid #FE8C29; padding: 14px 18px; margin: 12px 0 20px 0; border-radius: 4px; }}
        .summary-table {{ width: 100%; border-collapse: collapse; }}
        .summary-table td {{ padding: 5px 8px; font-size: 13px; vertical-align: top; }}
        .summary-label {{ font-weight: 700; color: #475569; width: 170px; }}
        .summary-val {{ color: #0F172A; }}
        .doc-list {{ background-color: #FFF7ED; border: 1px solid #FFEDD5; padding: 16px; border-radius: 8px; margin: 12px 0 20px 0; }}
        .doc-item {{ margin-bottom: 12px; }}
        .doc-item:last-child {{ margin-bottom: 0; }}
        .doc-name {{ font-weight: 700; color: #C2410C; font-size: 13px; }}
        .doc-desc {{ font-size: 12px; color: #64748B; margin-top: 2px; }}
        .signature {{ margin-top: 28px; padding-top: 18px; border-top: 1px solid #E2E8F0; font-size: 13px; color: #334155; }}
        .signature-title {{ color: #FE8C29; font-weight: 700; }}
        .footer {{ background-color: #F1F5F9; padding: 14px; font-size: 11px; color: #64748B; text-align: center; border-top: 1px solid #E2E8F0; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            {logo_img_tag}
            <h1>J&amp;D AUTOMATION INDUSTRIES</h1>
            <p>Propuesta Técnica &amp; Comercial de Automatización</p>
        </div>
        <div class="content">
            <p><b>Estimado(a) {contacto_saludo},</b></p>
            <p>
                A nombre de <b>J&amp;D Automation Industries</b>, le extendemos un cordial saludo. Nos dirigimos a usted con el agrado de presentarle formalmente nuestra propuesta técnica y económica para la <b>{proyecto}</b> en la planta de <b>{cliente}</b>, elaborada con base en el levantamiento de información y las necesidades operativas que nos han compartido.
            </p>
            <p>
                En J&amp;D nos comprometemos a ser un aliado estratégico en la evolución tecnológica de sus procesos de manufactura. La solución que ponemos a su consideración integra ingeniería de detalle, ensamble de tableros con componentes de marcas líderes a nivel global y una programación estructurada que garantiza operaciones seguras, eficientes y de fácil diagnóstico para sus operadores.
            </p>

            <div class="section-title">RESUMEN EJECUTIVO DE LA PROPUESTA:</div>
            <div class="summary-box">
                <table class="summary-table">
                    <tr><td class="summary-label">• Folio de Cotización:</td><td class="summary-val" style="color:#FE8C29; font-weight:bold;">{clean_f}</td></tr>
                    <tr><td class="summary-label">• Cliente:</td><td class="summary-val"><b>{cliente}</b></td></tr>
                    <tr><td class="summary-label">• Proyecto:</td><td class="summary-val"><b>{proyecto}</b></td></tr>
                    <tr><td class="summary-label">• Fecha / Revisión:</td><td class="summary-val"><b>{fecha_rev} ({rev_str})</b></td></tr>
                    <tr><td class="summary-label">• Monto Total Propuesta:</td><td class="summary-val" style="color:#059669; font-weight:bold; font-size:14px;">{precio_total_str}</td></tr>
                    <tr><td class="summary-label">• Tiempo de Entrega:</td><td class="summary-val">{tiempo_entrega} <span style="color:#64748B;">(A partir del anticipo e ingeniería base firmada)</span></td></tr>
                    <tr><td class="summary-label">• Esquema de Pago:</td><td class="summary-val">{condiciones_pago}</td></tr>
                </table>
            </div>

            <div class="section-title">DOCUMENTACIÓN ADJUNTA (Para su revisión):</div>
            <div class="doc-list">
                <div class="doc-item">
                    <div class="doc-name">1. {clean_folio_fname}_Propuesta_Tecnico_Comercial.pdf</div>
                    <div class="doc-desc">Documento membretado con el alcance de ingeniería, arquitectura de control (PLC/HMI), marcas propuestas y términos legales.</div>
                </div>
                <div class="doc-item">
                    <div class="doc-name">2. {clean_folio_fname}_Presupuesto_Financiero.xlsx</div>
                    <div class="doc-desc">Desglose económico transparente con fórmulas abiertas para su departamento de compras.</div>
                </div>
            </div>

            <p>
                Quedamos atentos para analizar juntos cada sección de este documento y agendar una sesión técnica si así lo requiere.
            </p>

            <div class="signature">
                <b>Atentamente,</b><br/><br/>
                <b style="font-size:14px; color:#0F172A;">Ing. David Alaniz</b><br/>
                <span class="signature-title">Área de Ingeniería Comercial &amp; Ventas</span><br/>
                <b>J&amp;D Automation Industries S.A. de C.V.</b><br/>
                Tel: 871 8176 8569 &bull; Cel: 871 795 4403<br/>
                Email: <a href="mailto:ventas@jdautomation.mx" style="color:#FE8C29;">ventas@jdautomation.mx</a> &bull; Web: <a href="https://www.jdautomation.mx" style="color:#FE8C29;">www.jdautomation.mx</a>
            </div>
        </div>
        <div class="footer">
            J&amp;D Automation Industries &bull; Calle F #382, Col. Eduardo Guerra, Torreón, Coahuila, México.
        </div>
    </div>
</body>
</html>"""

    # 1. Adjuntar únicamente el cuerpo HTML (con el logo incrustado en Base64)
    msg.attach(MIMEText(html_body, 'html', 'utf-8'))

    # 2. Adjuntar únicamente los 2 Archivos Oficiales (PDF y Excel)
    pdf_filename = f"{clean_folio_fname}_Propuesta_Tecnico_Comercial.pdf"
    excel_filename = f"{clean_folio_fname}_Presupuesto_Financiero.xlsx"

    part_pdf = MIMEApplication(pdf_bytes, _subtype="pdf")
    part_pdf.add_header("Content-Disposition", "attachment", filename=pdf_filename)
    msg.attach(part_pdf)

    part_excel = MIMEApplication(excel_bytes, _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part_excel.add_header("Content-Disposition", "attachment", filename=excel_filename)
    msg.attach(part_excel)

    return msg.as_bytes()


def _generate_zip_paquete(cot_info, pdf_bytes, excel_bytes, eml_bytes):
    clean_f = cot_info.get('folio', 'COT-2026-001-JAI-DS').replace('_Cotizacion_Oficial', '').strip()
    f_parts = clean_f.split('-')
    folio_corto = "-".join(f_parts[:3]) if len(f_parts) >= 3 else clean_f

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"{folio_corto}_Propuesta_Tecnico_Comercial.pdf", pdf_bytes)
        zf.writestr(f"{folio_corto}_Presupuesto_Financiero.xlsx", excel_bytes)
        zf.writestr(f"{folio_corto}_Correo_Notificacion.eml", eml_bytes)
    zip_buffer.seek(0)
    return zip_buffer.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# 5. DASHBOARD UI: CIERRE Y ENTREGA DE COTIZACIÓN CON TÉRMINOS Y RESPUESTAS TÉCNICAS
# ─────────────────────────────────────────────────────────────────────────────

def render_cierre_entrega():
    st.markdown(f"""
    <div class="jd-section-header">
        <h2>📦 Módulo de Cierre y Entrega de Cotización</h2>
        <p>Edición de Términos Comerciales, captura de Respuestas Técnicas con fotografías y descarga del paquete de entrega (.ZIP, PDF, .EML).</p>
    </div>""", unsafe_allow_html=True)

    # Selector de Cotización Aprobada y Congelada (únicamente las cerradas con candado)
    conn = get_connection(); cur = conn.cursor()
    cur.execute("""
        SELECT c.id, c.folio, c.proyecto, c.congelada, COALESCE(c.revision,'R0') as revision, COALESCE(cl.nombre,'TREBOTTI') as cliente
        FROM cotizaciones c LEFT JOIN clientes cl ON c.cliente_id = cl.id
        WHERE c.congelada = 1
        ORDER BY c.id DESC
    """)
    cots = [dict(r) for r in cur.fetchall()]
    conn.close()

    if not cots:
        st.info("🔒 Para generar el paquete oficial de cierre y entrega, la cotización debe estar **Aprobada y Congelada con Candado**. Ingresa a **'3. Modificador de Cotizaciones' (Paso 4: Análisis & Versiones)** y haz clic en **🔒 Aprobar y CONGELAR Cotización**.")
        return

    import re
    opt_cot = {
        f"🔒 {re.sub(r'\\s*\\(R\\d+\\)$', '', c['folio']).strip()} ({c.get('revision','R0')}) — {(c.get('proyecto') or '')[:40]} ({c.get('cliente','—')})": c['id']
        for c in cots
    }
    sel_label = st.selectbox("📌 Seleccionar Cotización Aprobada y Congelada para Cierre y Entrega", list(opt_cot.keys()))
    cot_id = opt_cot[sel_label]

    # Cargar detalles completos
    detalles = get_cotizacion_detalles(cot_id)
    cot_info = detalles.get('cotizacion', {})
    partidas = detalles.get('partidas', [])
    respuestas_tecnicas = detalles.get('respuestas_tecnicas', [])

    st.divider()

    # ── SECCIÓN 1: CONFIGURACIÓN DE TÉRMINOS COMERCIALES POR COTIZACIÓN ──
    with st.expander("📝 Configurar Términos Comerciales de la Cotización", expanded=False):
        t_col1, t_col2 = st.columns(2)
        with t_col1:
            moneda_input = st.selectbox(
                "Moneda de Cotización",
                ["MXN pesos mexicanos", "USD dólares americanos"],
                index=0 if "USD" not in cot_info.get("moneda_cotizacion","") else 1
            )
            cond_pago_input = st.text_input("Condiciones de Pago", value=cot_info.get("condiciones_pago","CREDITO"))
        with t_col2:
            _opciones_entrega = ["2 SEMANAS", "4 SEMANAS", "6 SEMANAS", "8 SEMANAS", "10 SEMANAS", "12 SEMANAS", "14 SEMANAS", "16 SEMANAS"]
            _val_entrega_actual = cot_info.get("tiempo_entrega","2 SEMANAS").upper()
            _idx_entrega = 0
            for _i_e, _op_e in enumerate(_opciones_entrega):
                if _val_entrega_actual in _op_e or _op_e in _val_entrega_actual:
                    _idx_entrega = _i_e
                    break

            entrega_input = st.selectbox(
                "Tiempo de Entrega",
                _opciones_entrega,
                index=_idx_entrega
            )
            vigencia_input = st.text_input("Vigencia de la Cotización", value=cot_info.get("vigencia_cotizacion","15 días"))

        if st.button("💾 Guardar Términos Comerciales en la Cotización", type="primary"):
            conn = get_connection()
            conn.execute("""
                UPDATE cotizaciones
                SET moneda_cotizacion=?, condiciones_pago=?, tiempo_entrega=?, vigencia_cotizacion=?
                WHERE id=?
            """, (moneda_input, cond_pago_input, entrega_input, vigencia_input, cot_id))
            conn.commit(); conn.close()
            st.success("Términos comerciales actualizados exitosamente en la cotización.")
            st.rerun()

    # ── SECCIÓN 2: RESPUESTAS TÉCNICAS (FOTO Y ESPECIFICACIÓN POR PARTIDA) ──
    with st.expander("📸 Agregar Respuesta Técnica / Especificación de Equipos con Foto", expanded=False):
        st.markdown("<p style='font-size:12px;color:#64748B;'>Integra fotografías y especificaciones técnicas detalladas de equipos que aparecerán dentro del documento PDF oficial.</p>", unsafe_allow_html=True)
        
        rt_col1, rt_col2 = st.columns([1, 2])
        with rt_col1:
            rt_partida_num = st.number_input("Partida N°", min_value=1, value=1, step=1)
            rt_comp = st.text_input("Componente / Equipo", "Tablero de Control de Motores NEMA 12")
            rt_foto = st.file_uploader("Fotografía del Equipo (.png, .jpg)", type=["png", "jpg", "jpeg"])

        with rt_col2:
            rt_spec = st.text_area(
                "Especificación Técnica Detallada",
                "Gabinete metálico Rittal 800x600x300mm NEMA 12. Incluye interruptor principal 3x70A NF, variador de frecuencia Allen-Bradley PowerFlex 525 10HP, relevadores de seguridad PILZ y clemas de conexión rápida Weidmuller.",
                height=130
            )

        if st.button("➕ Agregar Especificación Técnica a la Cotización", type="primary"):
            img_saved_path = ""
            if rt_foto:
                img_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "assets", "respuestas_tecnicas")
                os.makedirs(img_dir, exist_ok=True)
                img_saved_path = os.path.join(img_dir, f"rt_{cot_id}_{int(rt_partida_num)}_{rt_foto.name}")
                with open(img_saved_path, "wb") as f:
                    f.write(rt_foto.getbuffer())

            conn = get_connection()
            conn.execute("""
                INSERT INTO cotizacion_respuestas_tecnicas
                (cotizacion_id, partida_num, componente, especificacion_tecnica, imagen_path)
                VALUES (?, ?, ?, ?, ?)
            """, (cot_id, int(rt_partida_num), rt_comp.strip(), rt_spec.strip(), img_saved_path))
            conn.commit(); conn.close()
            st.success("Especificación técnica agregada a la cotización.")
            st.rerun()

        if respuestas_tecnicas:
            st.markdown("---")
            st.markdown("**Especificaciones Técnicas Registradas:**")
            for rt in respuestas_tecnicas:
                r_col1, r_col2 = st.columns([4, 1])
                r_col1.markdown(f"• **Partida {rt['partida_num']} — {rt['componente']}**: {rt['especificacion_tecnica'][:80]}...")
                if r_col2.button("🗑️ Borrar", key=f"del_rt_{rt['id']}"):
                    conn = get_connection()
                    conn.execute("DELETE FROM cotizacion_respuestas_tecnicas WHERE id=?", (rt['id'],))
                    conn.commit(); conn.close()
                    st.rerun()

    st.divider()

    # ── CONFIGURACIÓN DE DESTINATARIOS PARA EL CORREO .EML ──
    with st.expander("✉️ Personalizar Destinatario (Para) y Copia (Cc) del Correo (.EML)", expanded=False):
        st.markdown("<p style='font-size:12px;color:#475569;'>Puedes personalizar la dirección del cliente y las copias antes de descargar el borrador de correo (.EML) editable.</p>", unsafe_allow_html=True)
        col_eml1, col_eml2 = st.columns([1, 1])
        
        default_client_email = cot_info.get('email_contacto') or cot_info.get('email_cliente') or cot_info.get('email') or cot_info.get('correo') or 'contacto.cliente@empresa.com'
        contacto_name = cot_info.get('nombre_contacto') or 'Ing. Ricardo Gallegos'
        default_to_str = f"{contacto_name} <{default_client_email}>" if '<' not in str(default_client_email) else default_client_email
        
        with col_eml1:
            eml_to_custom = st.text_input("📩 Destinatario Principal (Para / To)", value=default_to_str, key="eml_to_input")
        with col_eml2:
            eml_cc_custom = st.text_input("📋 Correos en Copia (Cc)", value="alberto.morales@jydautomation.com.mx, david.alaniz@jydautomation.com.mx, ventas@jydautomation.com.mx", key="eml_cc_input")

    # Generar los entregables en memoria
    pdf_bytes = _generate_cotizacion_pdf_oficial(cot_info, partidas, respuestas_tecnicas)
    excel_bytes = _generate_presupuesto_excel(cot_info, partidas)
    eml_bytes = _generate_correo_eml(cot_info, pdf_bytes, excel_bytes, partidas=partidas, custom_to=eml_to_custom, custom_cc=eml_cc_custom)
    zip_bytes = _generate_zip_paquete(cot_info, pdf_bytes, excel_bytes, eml_bytes)

    # Sanitizar el folio para garantizar nombres de archivo HTTP 100% seguros sin acentos ni paréntesis
    import re
    folio = cot_info.get('folio', 'YES-202607-089')
    clean_folio_fname = re.sub(r'[^a-zA-Z0-9_-]', '_', folio).strip('_')

    # ── BOTÓN CENTRAL DE VALIDACIÓN Y DESCARGA ──
    st.markdown(f"""
    <div style="background:{BRAND_GRAY_BG};border:2px solid {BRAND_ORANGE};border-radius:12px;
                padding:22px;margin-bottom:24px;font-family:'Montserrat',sans-serif;box-shadow:0 4px 12px rgba(254,140,41,0.12);">
        <div style="display:flex;align-items:center;gap:12px;margin-bottom:10px;">
            <span style="font-size:26px;">📦</span>
            <span style="font-size:18px;font-weight:900;color:{BRAND_CHARCOAL};">PAQUETE COMPLETO DE ENTREGA Y CIERRE</span>
        </div>
        <p style="font-size:13px;color:{BRAND_CHARCOAL_MED};margin:0 0 16px 0;line-height:1.5;">
            Haz clic en cualquiera de los botones a continuación para descargar el <b>paquete comprimido (.ZIP)</b> o los entregables individuales (<b>PDF Oficial Membretado</b> y <b>Borrador de Correo .EML</b>) listos para enviar al cliente.
        </p>
    </div>
    """, unsafe_allow_html=True)

    # Auto-guardado persistente en estructura de carpetas (cotizaciones_guardadas/{folio}/)
    try:
        from database.storage_manager import save_cotizacion_to_folder
        save_cotizacion_to_folder(cot_info, pdf_bytes, excel_bytes, eml_bytes, zip_bytes)
    except Exception:
        pass

    st.markdown("""
    <style>
    .btn-correo-azul button {
        background: linear-gradient(135deg, #1E3A8A 0%, #0F172A 100%) !important;
        color: #FFFFFF !important;
        font-weight: 900 !important;
        font-size: 15px !important;
        border: 2px solid #2563EB !important;
        border-radius: 8px !important;
        box-shadow: 0 4px 14px rgba(30, 58, 138, 0.4) !important;
        transition: all 0.3s ease !important;
    }
    .btn-correo-azul button:hover {
        background: linear-gradient(135deg, #2563EB 0%, #1E3A8A 100%) !important;
        border-color: #60A5FA !important;
        color: #FFFFFF !important;
        transform: translateY(-2px);
        box-shadow: 0 6px 18px rgba(37, 99, 235, 0.6) !important;
    }
    .btn-correo-azul button p {
        color: #FFFFFF !important;
        font-weight: 900 !important;
        font-size: 15px !important;
    }
    </style>
    """, unsafe_allow_html=True)

    b_col1, b_col2, b_col3 = st.columns([2, 1, 1])
    with b_col1:
        st.download_button(
            label="📦 VALIDAR Y DESCARGAR PAQUETE COMPLETO (.ZIP)",
            data=zip_bytes,
            file_name=f"{clean_folio_fname}_Paquete_Completo_Entrega.zip",
            mime="application/zip",
            type="primary",
            use_container_width=True,
            key="btn_download_zip_main"
        )

    with b_col2:
        st.download_button(
            label="📄 DESCARGAR PDF (.PDF)",
            data=pdf_bytes,
            file_name=f"{clean_folio_fname}_Cotizacion_Oficial.pdf",
            mime="application/pdf",
            use_container_width=True,
            key="btn_download_pdf_main"
        )

    with b_col3:
        st.markdown('<div class="btn-correo-azul">', unsafe_allow_html=True)
        st.download_button(
            label="📬 ✉️ DESCARGAR CORREO (.EML)",
            data=eml_bytes,
            file_name=f"{clean_folio_fname}_Correo_Notificacion.eml",
            mime="message/rfc822",
            use_container_width=True,
            key="btn_download_eml_main"
        )
        st.markdown('</div>', unsafe_allow_html=True)

    st.divider()

    # ── PESTAÑAS DE PREVISUALIZACIÓN Y SUPERVISIÓN (UI) ──
    v_tab1, v_tab2, v_tab3 = st.tabs([
        "📄 1. Vista Previa PDF Corporativo",
        "✉️ 2. Vista Previa Cuerpo del Correo (.EML)",
        "📊 3. Datos del Presupuesto Excel"
    ])

    with v_tab1:
        st.markdown(f"""
        <div style="background:{BRAND_WHITE};border:1px solid {BRAND_BORDER_LIGHT};padding:18px;border-radius:10px;margin-bottom:12px;">
            <h4 style="color:{BRAND_CHARCOAL};margin:0 0 8px 0;">📄 ESTRUCTURA DEL PDF CORPORATIVO (REPLICA YES-202607-089 Y MEMBRETADO)</h4>
            <p style="font-size:12px;color:{BRAND_CHARCOAL_MED};">
                • <b>Hoja Membretada Oficial:</b> Fondo vectorial <code>assets/hoja_membretada.png</code> aplicado en todas las páginas.<br>
                • <b>Respuestas Técnicas Integradas:</b> {len(respuestas_tecnicas)} especificación(es) técnica(s) con fotografía.<br>
                • <b>Términos Comerciales Dinámicos:</b> Pago: <b>{cot_info.get('condiciones_pago','CREDITO')}</b> | Entrega: <b>{cot_info.get('tiempo_entrega','2 SEMANAS')}</b> | Vigencia: <b>{cot_info.get('vigencia_cotizacion','15 días')}</b>.<br>
                • <b>Anexos Legales:</b> 14 Cláusulas de ley completas en las páginas finales.
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        st.download_button(
            label="📄 DESCARGAR PDF OFICIAL DEDE ESTA VISTA",
            data=pdf_bytes,
            file_name=f"{clean_folio_fname}_Cotizacion_Oficial.pdf",
            mime="application/pdf",
            key="btn_dl_pdf_tab"
        )

    with v_tab2:
        contacto_v = cot_info.get('nombre_contacto') or 'Ing. Ricardo Gallegos'
        if not (contacto_v.startswith('Ing.') or contacto_v.startswith('Lic.') or contacto_v.startswith('Arq.')):
            contacto_saludo_v = f"Ing. {contacto_v}"
        else:
            contacto_saludo_v = contacto_v

        clean_f_v = folio.replace('_Cotizacion_Oficial', '').strip()
        f_parts_v = clean_f_v.split('-')
        folio_corto_v = "-".join(f_parts_v[:3]) if len(f_parts_v) >= 3 else clean_f_v

        fecha_rev_v = cot_info.get('fecha_str') or cot_info.get('fecha') or datetime.now().strftime("%d/%m/%Y")
        rev_str_v = cot_info.get('revision') or 'R0'

        total_sub_v = 0.0
        if partidas:
            for p in partidas:
                cd = float(p.get('costo_directo_total', 0) or 0)
                pv = float(p.get('precio_venta') or (cd / 0.65 if cd > 0 else 0))
                total_sub_v += pv
        if total_sub_v <= 0:
            total_sub_v = 107066.67 / 0.65

        total_con_iva_v = total_sub_v * 1.16
        moneda_v = cot_info.get('moneda_cotizacion', 'MXN')
        mon_v = 'USD' if 'USD' in str(moneda_v).upper() else 'MXN'
        precio_total_str_v = f"${total_con_iva_v:,.2f} {mon_v} (IVA Incluido)"

        logo_w_v, logo_h_v = 280, 70
        logo_b64_v = ""
        logo_path_v = get_brand_asset_path("logo_corporativo.png")
        if os.path.exists(logo_path_v):
            try:
                from PIL import Image
                im_v = Image.open(logo_path_v)
                im_v.thumbnail((280, 75), Image.Resampling.LANCZOS)
                logo_w_v, logo_h_v = im_v.size
                buf_v = io.BytesIO()
                im_v.save(buf_v, format="PNG")
                logo_b64_v = base64.b64encode(buf_v.getvalue()).decode('utf-8')
            except Exception:
                pass
        logo_img_html_v = f'<div style="background:#FFFFFF; padding:8px 16px; border-radius:8px; display:inline-block; margin-bottom:12px;"><img src="data:image/png;base64,{logo_b64_v}" width="{logo_w_v}" height="{logo_h_v}" style="width:{logo_w_v}px; height:{logo_h_v}px; display:block;" alt="J&amp;D Automation"></div>' if logo_b64_v else ''

        client_email_v = cot_info.get('email_contacto') or cot_info.get('email_cliente') or cot_info.get('email') or cot_info.get('correo') or 'contacto.cliente@empresa.com'

        st.markdown(f"""
        <div style="background:{BRAND_WHITE};border:1px solid {BRAND_BORDER_LIGHT};padding:22px;border-radius:10px;font-family:'Segoe UI',sans-serif;margin-bottom:12px;">
            <div style="background:#1E293B;padding:16px 20px;border-bottom:4px solid #FE8C29;border-radius:6px 6px 0 0;">
                {logo_img_html_v}
                <h3 style="color:#FFF;margin:0;font-size:16px;">J&amp;D AUTOMATION INDUSTRIES</h3>
                <p style="color:#FE8C29;margin:2px 0 0 0;font-size:11px;font-weight:bold;text-transform:uppercase;">Propuesta Técnica &amp; Comercial de Automatización</p>
            </div>
            <div style="padding:16px 0;font-size:13px;color:#334155;line-height:1.6;">
                <p><b>Para (To):</b> <code>{contacto_saludo_v} &lt;{client_email_v}&gt;</code></p>
                <p><b>Copia (Cc):</b> <code>alberto.morales@jydautomation.com.mx, david.alaniz@jydautomation.com.mx, ventas@jydautomation.com.mx</code></p>
                <p><b>Asunto:</b> <code>Propuesta Técnica y Comercial: {cot_info.get('proyecto','AUTOMATIZACIÓN DE LÍNEA DE PROCESO')} | Ref: {folio_corto_v}</code></p>
                <hr style="border:none;border-top:1px dashed #CBD5E1;margin:12px 0;">
                <p><b>Estimado(a) {contacto_saludo_v},</b></p>
                <p>A nombre de <b>J&amp;D Automation Industries</b>, le extendemos un cordial saludo. Nos dirigimos a usted con el agrado de presentarle formalmente nuestra propuesta técnica y económica para la <b>{cot_info.get('proyecto','AUTOMATIZACIÓN DE LÍNEA DE PROCESO')}</b> en la planta de <b>{cot_info.get('cliente','YESERA MONTERREY')}</b>, elaborada con base en el levantamiento de información y las necesidades operativas que nos han compartido.</p>
                <p>En J&amp;D nos comprometemos a ser un aliado estratégico en la evolución tecnológica de sus procesos de manufactura. La solución que ponemos a su consideración integra ingeniería de detalle, ensamble de tableros con componentes de marcas líderes a nivel global y una programación estructurada que garantiza operaciones seguras, eficientes y de fácil diagnóstico para sus operadores.</p>
                
                <p style="font-weight:800;font-size:11px;text-transform:uppercase;color:#FE8C29;margin-top:14px;margin-bottom:4px;">RESUMEN EJECUTIVO DE LA PROPUESTA:</p>
                <div style="background:#F8FAFC;border-left:4px solid #FE8C29;padding:10px 14px;border-radius:4px;margin-bottom:12px;">
                    <p style="margin:2px 0;">• <b>Folio de Cotización:</b> <span style="color:#FE8C29;font-weight:bold;">{clean_f_v}</span></p>
                    <p style="margin:2px 0;">• <b>Cliente:</b> {cot_info.get('cliente','YESERA MONTERREY')}</p>
                    <p style="margin:2px 0;">• <b>Proyecto:</b> {cot_info.get('proyecto','AUTOMATIZACIÓN DE LÍNEA DE PROCESO')}</p>
                    <p style="margin:2px 0;">• <b>Fecha / Revisión:</b> <b>{fecha_rev_v} ({rev_str_v})</b></p>
                    <p style="margin:2px 0;">• <b>Monto Total Propuesta:</b> <span style="color:#059669;font-weight:bold;font-size:14px;">{precio_total_str_v}</span></p>
                    <p style="margin:2px 0;">• <b>Tiempo de Entrega:</b> {cot_info.get('tiempo_entrega','14 semanas')} (A partir del anticipo e ingeniería base firmada)</p>
                    <p style="margin:2px 0;">• <b>Esquema de Pago:</b> {cot_info.get('condiciones_pago','50% Anticipo | 30% Contra entrega de tableros | 20% Cierre de SAT')}</p>
                </div>

                <p style="font-weight:800;font-size:11px;text-transform:uppercase;color:#FE8C29;margin-top:14px;margin-bottom:4px;">DOCUMENTACIÓN ADJUNTA (Para su revisión):</p>
                <div style="background:#FFF7ED;padding:12px;border-radius:6px;border:1px solid #FFEDD5;margin-bottom:12px;">
                    <p style="margin:2px 0;color:#C2410C;"><b>1. [{folio_corto_v}] Propuesta_Tecnico_Comercial.pdf</b></p>
                    <p style="margin:0 0 8px 0;font-size:11px;color:#64748B;">Documento membretado con el alcance de ingeniería, arquitectura de control (PLC/HMI), marcas propuestas y términos legales.</p>
                    <p style="margin:2px 0;color:#C2410C;"><b>2. [{folio_corto_v}] Presupuesto_Financiero.xlsx</b></p>
                    <p style="margin:0;font-size:11px;color:#64748B;">Desglose económico transparente con fórmulas abiertas para su departamento de compras.</p>
                </div>

                <p>Quedamos atentos para analizar juntos cada sección de este documento y agendar una sesión técnica si así lo requiere.</p>

                <div style="margin-top:16px;padding-top:12px;border-top:1px solid #E2E8F0;">
                    <b>Atentamente,</b><br/><br/>
                    <b style="font-size:14px;color:#0F172A;">Ing. David Alaniz</b><br/>
                    <span style="color:#FE8C29;font-weight:bold;">Área de Ingeniería Comercial &amp; Ventas</span><br/>
                    <b>J&amp;D Automation Industries S.A. de C.V.</b><br/>
                    Tel: 871 8176 8569 &bull; Cel: 871 795 4403<br/>
                    Email: <a href="mailto:ventas@jdautomation.mx" style="color:#FE8C29;">ventas@jdautomation.mx</a> &bull; Web: <a href="https://www.jdautomation.mx" style="color:#FE8C29;">www.jdautomation.mx</a><br>
                    <small style="color:#94A3B8;">J&amp;D Automation Industries &bull; Calle F #382, Col. Eduardo Guerra, Torreón, Coahuila, México.</small>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown('<div class="btn-correo-azul">', unsafe_allow_html=True)
        st.download_button(
            label="📬 ✉️ DESCARGAR BORRADOR DE CORREO (.EML) DESDE ESTA VISTA",
            data=eml_bytes,
            file_name=f"{clean_folio_fname}_Correo_Notificacion.eml",
            mime="message/rfc822",
            key="btn_dl_eml_tab"
        )
        st.markdown('</div>', unsafe_allow_html=True)

    with v_tab3:
        st.markdown(f"**Estructura Financiera del Presupuesto en Excel (`{folio}_Presupuesto_Financiero.xlsx`)**")
        df_prev = pd.DataFrame([
            {"Partida": p.get('numero_partida',1), "Descripción": p.get('descripcion','TABLERO DE CONTROL').upper(), "Cantidad": 1, "Unidad": "SERV", "Precio Unitario (MXN)": f"${p.get('costo_directo_total',107066.67)/0.65:,.2f}", "Subtotal (MXN) [Fórmula]": f"=C{10+i}*E{10+i}"}
            for i, p in enumerate(partidas if partidas else [{"numero_partida": 1, "descripcion": cot_info.get('proyecto','TABLERO DE CONTROL'), "costo_directo_total": 107066.67}])
        ])
        st.dataframe(df_prev, use_container_width=True, hide_index=True)
        
        st.download_button(
            label="📊 DESCARGAR PRESUPUESTO EXCEL (.XLSX) DESDE ESTA VISTA",
            data=excel_bytes,
            file_name=f"{clean_folio_fname}_Presupuesto_Financiero.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="btn_dl_excel_tab"
        )
