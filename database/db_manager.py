import sqlite3
import os
import openpyxl
import re
from .models import get_connection, init_db

def sync_cotizacion_totals(cotizacion_id):
    """
    Recalcula automáticamente los subtotales de cada partida (MAT, MO, SUP, SUB, MAQ, HTA, GASTOS)
    a partir del detalle línea por línea, prorratea Gastos y Herramienta, y actualiza la tabla cotizaciones.
    """
    conn = get_connection()
    cursor = conn.cursor()

    # 1. Obtener partidas de la cotización
    cursor.execute("SELECT id FROM cotizacion_partidas WHERE cotizacion_id = ?", (cotizacion_id,))
    partidas = cursor.fetchall()

    # 2. Recalcular Subtotal Gastos del Proyecto
    cursor.execute("SELECT SUM(importe_total) FROM cotizacion_gastos_detalle WHERE cotizacion_id = ?", (cotizacion_id,))
    gastos_totales = cursor.fetchone()[0] or 0.0

    # Actualizar gastos_indirectos en cotizaciones
    cursor.execute("UPDATE cotizaciones SET gastos_indirectos = ? WHERE id = ?", (gastos_totales, cotizacion_id))

    # Recalcular cada partida
    sum_cost1 = 0.0
    partidas_cost1 = {}

    for p in partidas:
        pid = p['id']
        
        # Total Materiales por Partida
        cursor.execute("SELECT SUM(importe_mxn) FROM cotizacion_materiales_detalle WHERE partida_id = ?", (pid,))
        tot_mat = cursor.fetchone()[0] or 0.0

        # Total MO por Partida
        cursor.execute("SELECT SUM(importe_total) FROM cotizacion_mo_detalle WHERE partida_id = ?", (pid,))
        tot_mo = cursor.fetchone()[0] or 0.0

        tot_sup = tot_mo * 0.30  # Supervisión predeterminada 30% MO

        # Total Subcontratos por Partida
        cursor.execute("SELECT SUM(importe_mxn) FROM cotizacion_subcontratos_detalle WHERE partida_id = ?", (pid,))
        tot_sub = cursor.fetchone()[0] or 0.0

        # Total Maquinaria por Partida
        cursor.execute("SELECT SUM(total_mxn) FROM cotizacion_maquinaria_detalle WHERE partida_id = ?", (pid,))
        tot_maq = cursor.fetchone()[0] or 0.0

        c1 = tot_mat + tot_mo + tot_sup
        partidas_cost1[pid] = {
            'mat': tot_mat,
            'mo': tot_mo,
            'sup': tot_sup,
            'sub': tot_sub,
            'maq': tot_maq,
            'cost1': c1
        }
        sum_cost1 += c1

    tot_mo_global = sum(p['mo'] for p in partidas_cost1.values())
    tot_hta_global = tot_mo_global * 0.03 # Herramienta 3% MO

    # Actualizar cada partida con prorrateo de HTA y GASTOS
    for pid, vals in partidas_cost1.items():
        pct = (vals['cost1'] / sum_cost1) if sum_cost1 > 0 else 0.0
        tot_hta_partida = tot_hta_global * pct
        tot_gastos_partida = gastos_totales * pct
        cd_partida = vals['cost1'] + vals['sub'] + vals['maq'] + tot_hta_partida + tot_gastos_partida

        cursor.execute("""
            UPDATE cotizacion_partidas
            SET costo_mat = ?, costo_mo = ?, costo_sup = ?, costo_sub = ?, costo_maq = ?,
                costo_hta = ?, costo_gastos = ?, costo_directo_total = ?
            WHERE id = ?
        """, (vals['mat'], vals['mo'], vals['sup'], vals['sub'], vals['maq'], tot_hta_partida, tot_gastos_partida, cd_partida, pid))

    conn.commit()
    conn.close()

def get_cotizacion_detalles(cotizacion_id):
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM cotizaciones WHERE id = ?", (cotizacion_id,))
    cot = dict(cursor.fetchone() or {})

    cursor.execute("SELECT * FROM cotizacion_partidas WHERE cotizacion_id = ? ORDER BY numero_partida", (cotizacion_id,))
    partidas = [dict(r) for r in cursor.fetchall()]

    cursor.execute("SELECT m.*, p.numero_partida FROM cotizacion_materiales_detalle m JOIN cotizacion_partidas p ON m.partida_id = p.id WHERE m.cotizacion_id = ?", (cotizacion_id,))
    materiales = [dict(r) for r in cursor.fetchall()]

    cursor.execute("SELECT mo.*, p.numero_partida FROM cotizacion_mo_detalle mo JOIN cotizacion_partidas p ON mo.partida_id = p.id WHERE mo.cotizacion_id = ?", (cotizacion_id,))
    mano_obra = [dict(r) for r in cursor.fetchall()]

    cursor.execute("SELECT s.*, p.numero_partida FROM cotizacion_subcontratos_detalle s JOIN cotizacion_partidas p ON s.partida_id = p.id WHERE s.cotizacion_id = ?", (cotizacion_id,))
    subcontratos = [dict(r) for r in cursor.fetchall()]

    cursor.execute("SELECT mq.*, p.numero_partida FROM cotizacion_maquinaria_detalle mq JOIN cotizacion_partidas p ON mq.partida_id = p.id WHERE mq.cotizacion_id = ?", (cotizacion_id,))
    maquinaria = [dict(r) for r in cursor.fetchall()]

    cursor.execute("SELECT * FROM cotizacion_gastos_detalle WHERE cotizacion_id = ?", (cotizacion_id,))
    gastos = [dict(r) for r in cursor.fetchall()]

    cursor.execute("SELECT * FROM cotizacion_respuestas_tecnicas WHERE cotizacion_id = ? ORDER BY partida_num, id", (cotizacion_id,))
    respuestas_tecnicas = [dict(r) for r in cursor.fetchall()]

    conn.close()
    return {
        'cotizacion': cot,
        'partidas': partidas,
        'materiales': materiales,
        'mano_obra': mano_obra,
        'subcontratos': subcontratos,
        'maquinaria': maquinaria,
        'gastos': gastos,
        'respuestas_tecnicas': respuestas_tecnicas
    }


def seed_initial_catalogs(excel_folder_path=None):
    """Siembra datos base de catálogos de MO, Clientes y Materiales"""
    init_db()
    conn = get_connection()
    cursor = conn.cursor()

    roles_base = [
        ("Diseñador de Controles", 5500.0, 1.45, 1.0, 0.0, 0.0),
        ("Ingeniero de Control", 5500.0, 1.45, 1.0, 0.0, 0.0),
        ("Programador PLC / Robot", 4500.0, 1.45, 1.0, 0.0, 0.0),
        ("Tablerista", 4500.0, 1.45, 1.0, 0.0, 0.0),
        ("Eléctricos", 3800.0, 1.45, 1.0, 0.0, 0.0),
        ("Diseñador Mecánico", 3500.0, 1.45, 1.0, 0.0, 0.0),
        ("Mecánicos", 3000.0, 1.45, 1.0, 0.0, 0.0),
        ("Montadores", 2800.0, 1.45, 1.0, 0.0, 0.0),
        ("Dibujante", 2500.0, 1.45, 1.0, 0.0, 0.0),
        ("Ayudantes", 2500.0, 1.45, 1.0, 0.0, 0.0),
        ("Supervisor de Obra", 6000.0, 1.45, 1.0, 0.0, 0.0),
    ]
    for r in roles_base:
        cursor.execute("""
            INSERT OR IGNORE INTO catalogo_mano_obra
            (categoria, sueldo_base_semanal, fasar, sobre_sueldo, bonos, viaticos_semanal)
            VALUES (?, ?, ?, ?, ?, ?)
        """, r)

    clientes_base = [
        ("DRAXTON SAN LUIS", "Contacto", "", "", "", "San Luis Potosí"),
        ("YESA - YESO Y MOLDURAS", "Contacto", "", "", "", "México"),
        ("OH - GABINETES Y AUTOMATIZACION", "Contacto", "", "", "", "México"),
        ("J&D AUTOMATION INTERNO", "Administración", "", "", "", "México"),
    ]
    for c in clientes_base:
        cursor.execute("""
            INSERT OR IGNORE INTO clientes (nombre, contacto, email, telefono, rfc, direccion)
            VALUES (?, ?, ?, ?, ?, ?)
        """, c)

    if excel_folder_path and os.path.exists(excel_folder_path):
        files = [f for f in os.listdir(excel_folder_path) if f.endswith('.xlsx') and not f.startswith('~$')]
        for f in files:
            p = os.path.join(excel_folder_path, f)
            try:
                wb = openpyxl.load_workbook(p, data_only=True)
                for sheet in wb.sheetnames:
                    if 'MATER' in sheet.upper():
                        ws = wb[sheet]
                        tc = 18.0
                        tc_val = ws.cell(1, 3).value
                        if isinstance(tc_val, (int, float)) and tc_val > 5:
                            tc = float(tc_val)
                        for r in range(3, ws.max_row + 1):
                            codigo = str(ws.cell(r, 1).value or '').strip()
                            desc = ws.cell(r, 2).value
                            pu_usd = ws.cell(r, 5).value or 0.0
                            pu_mxn = ws.cell(r, 6).value or 0.0
                            unidad = str(ws.cell(r, 4).value or 'PZA').strip().upper()
                            if desc and isinstance(desc, str) and len(desc.strip()) > 3:
                                try:
                                    pu_usd = float(pu_usd) if isinstance(pu_usd, (int, float)) else 0.0
                                    pu_mxn = float(pu_mxn) if isinstance(pu_mxn, (int, float)) else 0.0
                                    if pu_mxn == 0 and pu_usd > 0:
                                        pu_mxn = pu_usd * tc
                                    elif pu_usd == 0 and pu_mxn > 0:
                                        pu_usd = pu_mxn / tc
                                    if pu_mxn > 0:
                                        cursor.execute("""
                                            INSERT OR IGNORE INTO catalogo_materiales
                                            (codigo, descripcion, unidad, precio_unitario_usd, precio_unitario_mxn, categoria)
                                            VALUES (?, ?, ?, ?, ?, 'Importado Excel')
                                        """, (codigo, desc.strip(), unidad, pu_usd, pu_mxn))
                                except Exception:
                                    pass
            except Exception:
                pass

    conn.commit()
    conn.close()


def get_catalogo_mano_obra():
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, categoria, sueldo_base_semanal, fasar, sobre_sueldo, bonos, viaticos_semanal,
               (sueldo_base_semanal * fasar * sobre_sueldo) + bonos + viaticos_semanal AS costo_semanal,
               ((sueldo_base_semanal * fasar * sobre_sueldo) + bonos + viaticos_semanal) / 7.0 AS costo_diario_real
        FROM catalogo_mano_obra WHERE activo = 1 ORDER BY categoria
    """)
    rows = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return rows


def get_catalogo_materiales(filtro=""):
    conn = get_connection()
    cursor = conn.cursor()
    if filtro:
        cursor.execute("SELECT * FROM catalogo_materiales WHERE activo = 1 AND descripcion LIKE ? ORDER BY descripcion LIMIT 200", (f"%{filtro}%",))
    else:
        cursor.execute("SELECT * FROM catalogo_materiales WHERE activo = 1 ORDER BY descripcion LIMIT 200")
    rows = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return rows


def add_material(codigo, descripcion, unidad, pu_usd, pu_mxn, categoria="General"):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO catalogo_materiales (codigo, descripcion, unidad, precio_unitario_usd, precio_unitario_mxn, categoria)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (codigo, descripcion, unidad, pu_usd, pu_mxn, categoria))
    conn.commit()
    conn.close()


def update_mano_obra(role_id, sueldo_base, fasar=1.45, sobre_sueldo=1.0, bonos=0.0, viaticos=0.0):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE catalogo_mano_obra
        SET sueldo_base_semanal=?, fasar=?, sobre_sueldo=?, bonos=?, viaticos_semanal=?
        WHERE id=?
    """, (sueldo_base, fasar, sobre_sueldo, bonos, viaticos, role_id))
    conn.commit()
    conn.close()


# ─── CATÁLOGO BASE DE GASTOS GENERALE ───
def get_catalogo_gastos(filtro=""):
    conn = get_connection(); cursor = conn.cursor()
    if filtro:
        cursor.execute("""SELECT * FROM catalogo_gastos WHERE activo=1 AND
                          (concepto LIKE ? OR clave LIKE ? OR categoria LIKE ?)
                          ORDER BY id""",
                       (f"%{filtro}%", f"%{filtro}%", f"%{filtro}%"))
    else:
        cursor.execute("SELECT * FROM catalogo_gastos WHERE activo=1 ORDER BY id")
    rows = [dict(r) for r in cursor.fetchall()]; conn.close()
    return rows

def add_gasto_base(concepto, unidad="VJE", costo_default=0.0, categoria="Generales", clave="", uso=""):
    conn = get_connection(); cursor = conn.cursor()
    cursor.execute("""INSERT INTO catalogo_gastos
                      (clave, concepto, unidad, costo_unitario_default, categoria, uso_descripcion, activo)
                      VALUES (?,?,?,?,?,?,1)
                      ON CONFLICT(concepto) DO UPDATE SET
                      clave=excluded.clave, unidad=excluded.unidad,
                      costo_unitario_default=excluded.costo_unitario_default,
                      categoria=excluded.categoria, uso_descripcion=excluded.uso_descripcion""",
                   (clave.strip(), concepto.strip(), unidad, costo_default, categoria, uso))
    conn.commit(); conn.close()

def delete_gasto_base(gasto_id):
    conn = get_connection(); cursor = conn.cursor()
    cursor.execute("UPDATE catalogo_gastos SET activo=0 WHERE id=?", (gasto_id,))
    conn.commit(); conn.close()



# ─── CATÁLOGO BASE DE SUBCONTRATOS ───
def get_catalogo_subcontratos(filtro=""):
    conn = get_connection(); cursor = conn.cursor()
    if filtro:
        cursor.execute("SELECT * FROM catalogo_subcontratos WHERE activo=1 AND concepto LIKE ? ORDER BY concepto", (f"%{filtro}%",))
    else:
        cursor.execute("SELECT * FROM catalogo_subcontratos WHERE activo=1 ORDER BY concepto")
    rows = [dict(r) for r in cursor.fetchall()]; conn.close()
    return rows

def add_subcontrato_base(concepto, unidad="SERV", costo_ref=0.0, proveedor="", categoria="Especializados"):
    conn = get_connection(); cursor = conn.cursor()
    cursor.execute("INSERT OR REPLACE INTO catalogo_subcontratos (concepto, unidad, costo_referencia, proveedor_habitual, categoria, activo) VALUES (?,?,?,?,?,1)",
                   (concepto.strip(), unidad, costo_ref, proveedor, categoria))
    conn.commit(); conn.close()

def delete_subcontrato_base(sub_id):
    conn = get_connection(); cursor = conn.cursor()
    cursor.execute("UPDATE catalogo_subcontratos SET activo=0 WHERE id=?", (sub_id,))
    conn.commit(); conn.close()


def duplicar_cotizacion_nueva_revision(parent_cot_id):
    """
    Clona íntegramente una cotización congelada (R0, R1, etc.) creando un nuevo registro
    en 'cotizaciones' con revision='R(N+1)', congelada=0, estatus='Borrador'.
    Preserva intacta la versión congelada anterior y duplica todas sus partidas, materiales,
    mano de obra, subcontratos, maquinaria, gastos generales y tareas de Gantt asociadas.
    """
    conn = get_connection()
    cur = conn.cursor()

    # 1. Obtener cotización origen
    cur.execute("SELECT * FROM cotizaciones WHERE id=?", (parent_cot_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return None, "R0"
    parent = dict(row)

    curr_rev = parent.get('revision', 'R0') or 'R0'
    rev_num = int(re.sub(r'\D', '', curr_rev) or '0') + 1
    clean_base_folio = re.sub(r'\s*\(R\d+\)$', '', parent['folio']).strip()

    # Garantizar un folio único sin colisiones en SQLite
    while True:
        new_rev = f"R{rev_num}"
        candidate_folio = f"{clean_base_folio} ({new_rev})"
        cur.execute("SELECT id FROM cotizaciones WHERE folio=?", (candidate_folio,))
        if not cur.fetchone():
            new_folio = candidate_folio
            break
        rev_num += 1

    # 2. Insertar nueva cotización clonada
    historial_orig = str(parent.get('historial_modificaciones') or "").strip()
    nuevo_header_log = f"--- REVISIÓN {new_rev} CREADA (Copia editable de {curr_rev}) ---"
    nuevo_historial = f"{nuevo_header_log}\n{historial_orig}" if historial_orig else nuevo_header_log

    cur.execute("""
        INSERT INTO cotizaciones (
            folio, cliente_id, nombre_contacto, correo_contacto, telefono_contacto,
            proyecto, fecha, estatus, revision, congelada, tipo_cambio_usd,
            margen_porcentaje, comision_porcentaje, supervision_porcentaje, herramienta_porcentaje,
            condiciones_pago, tiempo_entrega, vigencia_cotizacion, moneda_cotizacion, hitos_pago_json, historial_modificaciones
        ) VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, 'Borrador', ?, 0, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        new_folio, parent.get('cliente_id'), parent.get('nombre_contacto'),
        parent.get('correo_contacto'), parent.get('telefono_contacto'),
        parent['proyecto'], new_rev, parent.get('tipo_cambio_usd', 18.0),
        parent.get('margen_porcentaje', 0.30), parent.get('comision_porcentaje', 0.05),
        parent.get('supervision_porcentaje', 0.30), parent.get('herramienta_porcentaje', 0.03),
        parent.get('condiciones_pago', 'CREDITO'), parent.get('tiempo_entrega', '2 SEMANAS'),
        parent.get('vigencia_cotizacion', '15 días'), parent.get('moneda_cotizacion', 'MXN pesos mexicanos'),
        parent.get('hitos_pago_json'), nuevo_historial
    ))
    new_cot_id = cur.lastrowid

    # 3. Duplicar Gastos Generales
    cur.execute("SELECT * FROM cotizacion_gastos_detalle WHERE cotizacion_id=?", (parent_cot_id,))
    gastos = [dict(r) for r in cur.fetchall()]
    for g in gastos:
        cur.execute("""
            INSERT INTO cotizacion_gastos_detalle
            (cotizacion_id, nombre, cantidad, unidad, tiempo_valor, tiempo_unidad, costo_unitario, importe_total)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (new_cot_id, g['nombre'], g['cantidad'], g['unidad'], g.get('tiempo_valor', 1.0), g.get('tiempo_unidad', 'DIAS'), g['costo_unitario'], g['importe_total']))

    # 3b. Duplicar Respuestas Técnicas / Especificaciones con Foto
    cur.execute("SELECT * FROM cotizacion_respuestas_tecnicas WHERE cotizacion_id=?", (parent_cot_id,))
    respuestas = [dict(r) for r in cur.fetchall()]
    for r in respuestas:
        cur.execute("""
            INSERT INTO cotizacion_respuestas_tecnicas
            (cotizacion_id, partida_num, componente, especificacion_tecnica, imagen_path)
            VALUES (?, ?, ?, ?, ?)
        """, (new_cot_id, r.get('partida_num', 1), r['componente'], r['especificacion_tecnica'], r.get('imagen_path')))

    # 4. Duplicar Partidas y sus detalles (Materiales, MO, Subcontratos, Maquinaria)
    cur.execute("SELECT * FROM cotizacion_partidas WHERE cotizacion_id=? ORDER BY numero_partida", (parent_cot_id,))
    partidas = [dict(r) for r in cur.fetchall()]

    for p in partidas:
        old_pid = p['id']
        cur.execute("""
            INSERT INTO cotizacion_partidas
            (cotizacion_id, numero_partida, descripcion, costo_mat, costo_mo, costo_sup, costo_sub, costo_maq, costo_hta, costo_gastos, costo_directo_total, precio_venta_partida)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (new_cot_id, p['numero_partida'], p['descripcion'], p['costo_mat'], p['costo_mo'], p['costo_sup'], p['costo_sub'], p['costo_maq'], p['costo_hta'], p['costo_gastos'], p['costo_directo_total'], p.get('precio_venta_partida', 0.0)))
        new_pid = cur.lastrowid

        # Copy Materiales
        cur.execute("SELECT * FROM cotizacion_materiales_detalle WHERE partida_id=?", (old_pid,))
        for m in cur.fetchall():
            m_d = dict(m)
            cur.execute("""
                INSERT INTO cotizacion_materiales_detalle
                (cotizacion_id, partida_id, codigo, descripcion, cantidad, unidad, precio_unitario_usd, precio_unitario_mxn, importe_mxn)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (new_cot_id, new_pid, m_d.get('codigo'), m_d['descripcion'], m_d['cantidad'], m_d['unidad'], m_d.get('precio_unitario_usd', 0.0), m_d['precio_unitario_mxn'], m_d['importe_mxn']))

        # Copy MO
        cur.execute("SELECT * FROM cotizacion_mo_detalle WHERE partida_id=?", (old_pid,))
        for mo in cur.fetchall():
            mo_d = dict(mo)
            cur.execute("""
                INSERT INTO cotizacion_mo_detalle
                (cotizacion_id, partida_id, categoria_nombre, cantidad_personal, sueldo_base_semanal, fasar, sobre_sueldo, bonos, viaticos_semanal, semanas, horas_hombre, importe_total)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (new_cot_id, new_pid, mo_d['categoria_nombre'], mo_d['cantidad_personal'], mo_d['sueldo_base_semanal'], mo_d.get('fasar', 1.45), mo_d.get('sobre_sueldo', 1.0), mo_d.get('bonos', 0.0), mo_d.get('viaticos_semanal', 0.0), mo_d['semanas'], mo_d['horas_hombre'], mo_d['importe_total']))

        # Copy Subcontratos
        cur.execute("SELECT * FROM cotizacion_subcontratos_detalle WHERE partida_id=?", (old_pid,))
        for s in cur.fetchall():
            s_d = dict(s)
            cur.execute("""
                INSERT INTO cotizacion_subcontratos_detalle
                (cotizacion_id, partida_id, descripcion, cantidad, unidad, pu_mxn, importe_mxn, subcontratista)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (new_cot_id, new_pid, s_d['descripcion'], s_d['cantidad'], s_d['unidad'], s_d['pu_mxn'], s_d['importe_mxn'], s_d.get('subcontratista')))

        # Copy Maquinaria
        cur.execute("SELECT * FROM cotizacion_maquinaria_detalle WHERE partida_id=?", (old_pid,))
        for mq in cur.fetchall():
            mq_d = dict(mq)
            cur.execute("""
                INSERT INTO cotizacion_maquinaria_detalle
                (cotizacion_id, partida_id, clave, nombre, cantidad, unidad, costo_unitario, total_mxn)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (new_cot_id, new_pid, mq_d.get('clave'), mq_d['nombre'], mq_d['cantidad'], mq_d['unidad'], mq_d['costo_unitario'], mq_d['total_mxn']))

    # 5. Copy Gantt tareas si existen
    cur.execute("SELECT * FROM cotizacion_gantt WHERE cotizacion_id=?", (parent_cot_id,))
    gantt_tasks = [dict(r) for r in cur.fetchall()]
    for gt in gantt_tasks:
        cur.execute("""
            INSERT INTO cotizacion_gantt
            (cotizacion_id, partida_id, actividad, tipo, responsable, fecha_inicio, dias_duracion, orden)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (new_cot_id, gt.get('partida_id'), gt['actividad'], gt.get('tipo', 'Actividad'), gt.get('responsable'), gt.get('fecha_inicio'), gt.get('dias_duracion', 1), gt.get('orden', 0)))

    conn.commit()
    conn.close()

    sync_cotizacion_totals(new_cot_id)
    return new_cot_id, new_rev


def backup_database_zip():
    """
    Empaqueta y genera un archivo .ZIP en memoria conteniendo la base de datos cotizador.db.
    """
    import io
    import zipfile
    from datetime import datetime

    db_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "database", "cotizador.db")
    zip_buffer = io.BytesIO()

    time_tag = datetime.now().strftime("%Y%m%d_%H%M%S")
    zip_filename = f"Backup_BD_Cotizador_JD_{time_tag}.zip"

    with zipfile.ZipHelper if hasattr(zipfile, 'ZipHelper') else zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        if os.path.exists(db_path):
            zf.write(db_path, arcname=f"cotizador_backup_{time_tag}.db")

    zip_buffer.seek(0)
    return zip_buffer.getvalue(), zip_filename


def factory_reset_database():
    """
    Borrado de Fábrica: Elimina todos los registros de cotizaciones, partidas y desgloses
    manteniendo 100% intactas las tablas de catálogos base (clientes, mano de obra, gastos, subcontratos, maquinaria).
    """
    conn = get_connection()
    cur = conn.cursor()
    
    tablas_transaccionales = [
        "cotizacion_materiales_detalle",
        "cotizacion_mo_detalle",
        "cotizacion_subcontratos_detalle",
        "cotizacion_maquinaria_detalle",
        "cotizacion_gastos_detalle",
        "cotizacion_partidas",
        "proyecto_gantt_tareas",
        "cotizaciones"
    ]
    
    for tbl in tablas_transaccionales:
        try:
            cur.execute(f"DELETE FROM {tbl}")
        except Exception:
            pass
            
    conn.commit()
    conn.close()


def bulk_delete_cotizaciones(cotizacion_ids):
    """
    Elimina un conjunto de cotizaciones (por lista de IDs) con borrado seguro en cascada.
    """
    if not cotizacion_ids:
        return
        
    conn = get_connection()
    cur = conn.cursor()
    placeholders = ",".join(["?"] * len(cotizacion_ids))
    
    tablas_detalle = [
        "cotizacion_materiales_detalle",
        "cotizacion_mo_detalle",
        "cotizacion_subcontratos_detalle",
        "cotizacion_maquinaria_detalle",
        "cotizacion_gastos_detalle",
        "cotizacion_partidas",
        "proyecto_gantt_tareas",
        "cotizaciones"
    ]
    
    for tbl in tablas_detalle:
        try:
            cur.execute(f"DELETE FROM {tbl} WHERE cotizacion_id IN ({placeholders})", cotizacion_ids)
        except Exception:
            try:
                cur.execute(f"DELETE FROM {tbl} WHERE id IN ({placeholders})", cotizacion_ids)
            except Exception:
                pass
                
    conn.commit()
    conn.close()


def bulk_update_cotizaciones_estatus(cotizacion_ids, nuevo_estatus):
    """
    Actualiza el estatus comercial de un conjunto de cotizaciones.
    """
    if not cotizacion_ids or not nuevo_estatus:
        return
        
    conn = get_connection()
    cur = conn.cursor()
    placeholders = ",".join(["?"] * len(cotizacion_ids))
    
    params = [nuevo_estatus] + list(cotizacion_ids)
    cur.execute(f"UPDATE cotizaciones SET estatus = ? WHERE id IN ({placeholders})", params)
    conn.commit()
    conn.close()



